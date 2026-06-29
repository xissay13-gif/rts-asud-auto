"""
resolutions.py — Выдача резолюций по обращениям из реестра.

Запускается под учёткой Есиной М.В., автоматически переключается
на учётку Халецкой Ю.В. (через выпадашку в шапке профиля).
Заходит в раздел "Исполнение" (там у Халецкой лежат документы которым
нужно выдать резолюцию — id раздела CABINET_MENU__RECEIVED__ALL_ACTIVE__TO_EXECUTION),
для каждой строки реестра находит соответствующий документ в АСУД и
выдаёт резолюцию начальнику абонентского отдела (по округу из колонки
ao/fio в реестре).

Реестр: Лист2 с колонками Link, Subject, TextBody, Тема, To, LS, ao, fio.
Матч документа в АСУД: по номеру обращения из TextBody (regex
"обращение № NNNNNN"), fallback на подстроку первых 60 символов TextBody.
"""

import os
import re
import sys
import time
import signal
import argparse
import logging
from datetime import date, datetime, timedelta

import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    WebDriverException, InvalidSessionIdException,
    NoSuchWindowException, NoSuchElementException,
)

import config as cfg
from ui import (click, wait_and_click, find_input_near_label,
                wait_asud_loaded, wait_modal_closed, close_open_modals, js_set_value)
from correspondent import match_correspondent
from xlsx_status import mark_status, get_done_asud_ids, COL_HALETSKAYA, COL_OKRUG
from xlsx_lock import is_xlsx_busy
from deadline import parse_ddmmyyyy, add_working_days, compute_deadline

_log_console = logging.StreamHandler()
_log_console.setLevel(logging.INFO)
_log_console.setFormatter(logging.Formatter(
    '%(asctime)s [%(levelname)s] %(message)s', datefmt='%H:%M:%S'))

logging.basicConfig(level=logging.DEBUG, handlers=[_log_console])
log = logging.getLogger("asud.res")
start_time = time.monotonic()


def _attach_file_logger(base_dir, rotate_daily=False):
    """Подключает FileHandler с DEBUG: <base_dir>/Logs/resolutions_<timestamp>.log.

    rotate_daily=True (для --watch): TimedRotatingFileHandler с ротацией
    в полночь, бэкап 7 дней. Имя файла без timestamp: resolutions.log
    (а старые получают суффикс .YYYY-MM-DD).
    """
    try:
        logs_dir = os.path.join(base_dir, "Logs")
        os.makedirs(logs_dir, exist_ok=True)
        if rotate_daily:
            from logging.handlers import TimedRotatingFileHandler
            path = os.path.join(logs_dir, "resolutions.log")
            fh = TimedRotatingFileHandler(
                path, when='midnight', backupCount=7, encoding='utf-8')
        else:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = os.path.join(logs_dir, f"resolutions_{ts}.log")
            fh = logging.FileHandler(path, encoding='utf-8')
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(logging.Formatter(
            '%(asctime)s.%(msecs)03d [%(levelname)s] %(name)s: %(message)s',
            datefmt='%H:%M:%S'))
        logging.getLogger().addHandler(fh)
        log.info(f"Подробный лог пишется в: {path}"
                 + (" (ротация ежесуточно, бэкап 7д)" if rotate_daily else ""))
        return path
    except Exception as e:
        log.warning(f"Не удалось создать файл лога: {e}")
        return None

settings = {}


# ============================================================
# EXCEL
# ============================================================

APPEAL_RE = re.compile(r'обращени[ея]\s*№?\s*(\d{4,10})', re.IGNORECASE)


# Lookup street_name → set of (house, okrug_short) — индекс для матчинга
# адресов из TextBody с адресной БД. Загружается лениво.
_STREET_INDEX = None
_ALL_STREETS_SORTED = None  # отсортированы по длине (длинные первыми)


def _addresses_csv_path():
    """Возвращает путь к addresses.csv: внутри exe (через _MEIPASS)
    или рядом с .py-скриптом в dev-режиме."""
    # PyInstaller --onefile распаковывает данные в sys._MEIPASS
    meipass = getattr(sys, '_MEIPASS', None)
    if meipass:
        path = os.path.join(meipass, 'addresses.csv')
        if os.path.exists(path):
            return path
    # Fallback: рядом с exe / скриптом
    base = cfg.get_base_dir()
    path = os.path.join(base, 'addresses.csv')
    if os.path.exists(path):
        return path
    return None


# --- Нормализация для парсинга адресов ---

_PREFIX_RE = re.compile(
    r'^(?:г\s*омск\s*,?\s*)?'
    r'(?:улица|ул\.?|проспект|пр[-]?т\.?|пр[-]?кт\.?|пр\.?|переулок|пер\.?|'
    r'бульвар|б[-]?р\.?|площадь|пл\.?|шоссе|ш\.?|набережная|наб\.?|'
    r'линия|тупик|проезд|пр[-]?д\.?|микрорайон|мкр\.?)\s*',
    re.IGNORECASE)

_HOUSE_RE = re.compile(r'(\d+[а-я]?)(?:[/\\-](\d+[а-я]?))?', re.IGNORECASE)


def _norm_text_for_match(s):
    """Нормализация для сопоставления: lower, ё→е, без пунктуации, single space.
    Также '3-я Молодежная' → '3 Молодежная'."""
    if not s:
        return ''
    s = str(s).lower().replace('ё', 'е')
    s = re.sub(r'[«»"\'`]', '', s)
    s = re.sub(r'[.,;:()\\/]+', ' ', s)
    s = re.sub(r'(\d+)[\s-]*я(?=\s)', r'\1', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _norm_street_name(s):
    """Только название улицы без префиксов."""
    return _PREFIX_RE.sub('', _norm_text_for_match(s)).strip()


def _norm_house_main(s):
    """Возвращает основной номер дома (первое число)."""
    if not s: return ''
    m = _HOUSE_RE.search(str(s).lower().replace('ё', 'е'))
    return m.group(1) if m else ''


def _build_street_index():
    """Парсит addresses.csv в индекс {street_norm: set((house, okrug_short))}.
    Используется один раз (кешируется в _STREET_INDEX)."""
    global _STREET_INDEX, _ALL_STREETS_SORTED
    if _STREET_INDEX is not None:
        return _STREET_INDEX, _ALL_STREETS_SORTED
    path = _addresses_csv_path()
    if not path:
        log.warning("addresses.csv не найден — авто-определение округа отключено")
        _STREET_INDEX = {}
        _ALL_STREETS_SORTED = []
        return _STREET_INDEX, _ALL_STREETS_SORTED

    # CSV: LS;okrug
    # Старая версия CSV содержит только LS+okrug, но для адрес-парсинга
    # нужны улицы из исходного xlsx. Если у нас только CSV без улиц —
    # парсер не сработает; идём через колонку ao реестра.
    # В будущем addresses.csv можно расширить колонкой улицы.
    try:
        import csv
        from collections import defaultdict
        idx = defaultdict(set)
        with open(path, encoding='utf-8') as f:
            reader = csv.reader(f, delimiter=';')
            header = next(reader, None)
            # Пытаемся определить колонки
            if not header or 'okrug' not in [h.lower() for h in header]:
                _STREET_INDEX = {}
                _ALL_STREETS_SORTED = []
                return _STREET_INDEX, _ALL_STREETS_SORTED
            # Если в CSV есть street/house — используем
            cols = {h.lower(): i for i, h in enumerate(header)}
            if 'street' in cols and 'house' in cols:
                for row in reader:
                    if not row or len(row) <= max(cols['street'], cols['house'], cols['okrug']):
                        continue
                    street = _norm_street_name(row[cols['street']])
                    house = _norm_house_main(row[cols['house']])
                    okrug = row[cols['okrug']].strip()
                    if street and house and okrug:
                        idx[street].add((house, okrug))
                        # Алиасы: 'молодежная 3-я' → '3-я молодежная',
                        # '3 молодежная', 'молодежная 3'
                        m = re.match(r'^(.+?)\s+(\d+)[\s-]*я$', street)
                        if m:
                            base, num = m.group(1).strip(), m.group(2)
                            idx[f"{base} {num}"].add((house, okrug))
                            idx[f"{num}-я {base}"].add((house, okrug))
                            idx[f"{num} {base}"].add((house, okrug))
        log.info(f"Street index: {len(idx)} улиц")
        _STREET_INDEX = idx
        _ALL_STREETS_SORTED = sorted(idx.keys(), key=lambda s: -len(s))
    except Exception as e:
        log.warning(f"Ошибка построения street index: {e}")
        _STREET_INDEX = {}
        _ALL_STREETS_SORTED = []
    return _STREET_INDEX, _ALL_STREETS_SORTED


def _find_street_house(text, idx, sorted_streets):
    """Ищет известную улицу в нормализованном тексте, потом дом рядом."""
    norm = _norm_text_for_match(text)
    for street in sorted_streets:
        if len(street) < 3:
            continue
        # Поиск как целое слово
        pos = 0
        while True:
            i = norm.find(street, pos)
            if i < 0: break
            left_ok = i == 0 or not norm[i-1].isalnum()
            end = i + len(street)
            right_ok = end == len(norm) or not norm[end].isalnum()
            if left_ok and right_ok:
                # Дом в окне 50 символов после улицы
                tail = norm[end:end+50]
                m = re.search(r'\b(\d+[а-я]?)', tail)
                if m:
                    return (street, m.group(1))
            pos = i + 1
    return (None, None)


def _okrug_from_textbody(textbody):
    """Извлекает округ из TextBody.

    Стратегия: суть обращения → почтовый адрес → весь текст.
    Возвращает короткий код округа ('КАО', 'ЦАО', ...) или None.
    """
    if not textbody:
        return None
    idx, sorted_streets = _build_street_index()
    if not idx:
        return None
    text = str(textbody)
    fragments = []
    # 1) Суть обращения
    m = re.search(r'суть\s+обращени[яе]\s*:?\s*([\s\S]+?)(?:\n\s*\n|$)',
                  text, re.IGNORECASE)
    if m:
        fragments.append(('суть', m.group(1)))
    # 2) Почтовый адрес
    m = re.search(r'почтов[а-я]+\s+адрес[а-я]*\s*:\s*([^\n]+)',
                  text, re.IGNORECASE)
    if m:
        fragments.append(('почт', m.group(1)))
    # 3) Весь текст как fallback
    fragments.append(('весь', text))

    for name, frag in fragments:
        street, house = _find_street_house(frag, idx, sorted_streets)
        if street and house:
            # '15г' → '15' для сравнения (литерные суффиксы — те же дома)
            house_digits = re.match(r'\d+', house)
            house_num = house_digits.group(0) if house_digits else house
            for h, o in idx[street]:
                h_digits = re.match(r'\d+', h)
                h_num = h_digits.group(0) if h_digits else h
                if h == house or h_num == house_num:
                    log.info(f"  адрес [{name}]: {street} {house} → {o}")
                    return o
    return None


def _resolve_executor(ao, fio, ls=None, textbody=None):
    """Возвращает ФИО начальника по приоритету:
    1. fio из реестра (если заполнено)
    2. ao из реестра → DEFAULT_OKRUG_MAP
    3. адрес из TextBody → addresses.csv → DEFAULT_OKRUG_MAP

    Параметр ls пока не используется — формат LS в реестре не совпадает
    с адресной БД (11 цифр vs 6), нужно правило конвертации.
    """
    # 1. fio напрямую (вручную проставленный)
    if fio and str(fio).strip():
        return str(fio).strip()
    # 2. ao из реестра (вручную проставленный)
    if ao:
        key = str(ao).strip()
        v = cfg.DEFAULT_OKRUG_MAP.get(key)
        if v:
            return v
    # 3. Парсинг адреса из TextBody
    if textbody:
        ao_short = _okrug_from_textbody(textbody)
        if ao_short:
            v = cfg.DEFAULT_OKRUG_MAP.get(ao_short)
            if v:
                return v
    return None


def load_excel(file_path):
    """Читает таблицу резолюций. Ожидаемые колонки (любой порядок,
    распознаются по заголовку): ОПТС, Округ, ФИО, Link.

    Если файл — старый формат (Лист2 с TextBody) — fallback на адрес-парсер.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Заголовки
    header = [str(c.value or '').strip() for c in next(ws.iter_rows(max_row=1))]
    header_lower = [h.lower() for h in header]
    log.info(f"Заголовки: {header}")

    # Если это _резолюции.xlsx (есть колонка ОПТС/asud_id)
    asud_keys = ('опт', 'орт', 'асуд', 'asud', 'регистрацион', 'номер')
    fio_keys = ('фио', 'исполнит')
    okrug_keys = ('округ', 'ао')
    link_keys = ('link', 'ссылк')
    planned_keys = ('планиру',)
    receipt_keys = ('дата получ', 'получено', 'получения')

    def _col(predicate_keys):
        for i, h in enumerate(header_lower):
            for k in predicate_keys:
                if k in h:
                    return i
        return None

    asud_col = _col(asud_keys)
    if asud_col is not None:
        # Формат _резолюции.xlsx
        fio_col = _col(fio_keys)
        okrug_col = _col(okrug_keys)
        link_col = _col(link_keys)
        planned_col = _col(planned_keys)
        receipt_col = _col(receipt_keys)
        log.info(f"Формат _резолюции: ОПТС=col{asud_col}, "
                 f"ФИО=col{fio_col}, Округ=col{okrug_col}, "
                 f"Link=col{link_col}, ПланДата=col{planned_col}, "
                 f"ДатаПолуч=col{receipt_col}")

        def _cell_date(row, col):
            """Значение ячейки-даты → DD.MM.YYYY (datetime→strftime, иначе str)."""
            if col is None or col >= len(row) or not row[col]:
                return ''
            pv = row[col]
            return pv.strftime("%d.%m.%Y") if hasattr(pv, 'strftime') else str(pv).strip()

        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row:
                continue
            asud_id = str(row[asud_col]).strip() if row[asud_col] else ''
            fio = str(row[fio_col]).strip() if (fio_col is not None and row[fio_col]) else ''
            ao = str(row[okrug_col]).strip() if (okrug_col is not None and row[okrug_col]) else ''
            link = row[link_col] if (link_col is not None) else None
            planned = _cell_date(row, planned_col)
            receipt = _cell_date(row, receipt_col)

            # Если ФИО пуст, но Округ есть — пытаемся через мапу
            if not fio and ao:
                fio = cfg.DEFAULT_OKRUG_MAP.get(ao, '')

            if not asud_id and not link:
                continue

            rows.append({
                "row_idx": row_idx,
                "asud_id": asud_id,
                "executor": fio,
                "ao": ao,
                "link": link,
                "subject": "",
                "textbody": "",
                "appeal_no": None,
                "ls": "",
                "planned_date": planned,
                "receipt_date": receipt,
            })
        wb.close()
        log.info(f"Загружено: {len(rows)} строк (формат _резолюции)")
        no_asud = sum(1 for r in rows if not r['asud_id'])
        no_fio = sum(1 for r in rows if not r['executor'])
        if no_asud:
            log.warning(f"  без ОПТС/ОРТС: {no_asud} (попробую матч по тексту)")
        if no_fio:
            log.warning(f"  без ФИО: {no_fio} (будут пропущены)")
        return rows

    # FALLBACK: старый формат — Почта_ТЭС.xlsx Лист2
    log.warning("Колонка ОПТС не найдена — пробую старый формат Лист2")
    if 'Лист2' in wb.sheetnames:
        ws = wb['Лист2']

    rows = []
    skipped = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
        if not row or len(row) < 8:
            skipped += 1
            continue
        link = row[0]
        subject = row[1]
        textbody = row[2] or ''
        ls = row[5] if len(row) > 5 else None
        ao = row[6] if len(row) > 6 else None
        fio = row[7] if len(row) > 7 else None
        if not link and not subject:
            skipped += 1
            continue
        executor = _resolve_executor(ao, fio, ls, textbody)
        appeal_no = _extract_appeal_no(textbody)
        rows.append({
            "row_idx": row_idx,
            "asud_id": "",
            "link": link,
            "subject": str(subject or '').strip(),
            "textbody": str(textbody),
            "ls": _norm_ls(ls),
            "ao": str(ao or '').strip(),
            "executor": executor,
            "appeal_no": appeal_no,
            "planned_date": "",
            "receipt_date": "",
        })
    wb.close()
    log.info(f"Загружено: {len(rows)} (старый формат), пропущено: {skipped}")
    return rows


# ============================================================
# UI: переключение учётки
# ============================================================

# ============================================================
# Account switching — DOM-based (НЕ пиксельные координаты)
# ============================================================
#
# АСУД-якоря (стабильные id/атрибуты, не зависят от viewport/headless):
#
#   #show-another-user-icon         — стрелка вниз (16×16, title="Выбрать
#                                     учетную запись"). Клик по НЕЙ открывает
#                                     выпадашку со списком учёток.
#
#   #app-acc-icon-myprofile         — аватарка пользователя (45×45,
#                                     title="Учетная запись"). НЕ кликабельна
#                                     для смены учётки — это просто иконка.
#
#   span#app-acc-info[x-account-name="..."]  — отображает ФИО учётки.
#                                     ЕСТЬ И В ШАПКЕ (текущая), И В ВЫПАДАШКЕ
#                                     (доступные для переключения).
#
#   Разница: спан в шапке лежит без position:absolute родителя,
#   а спан в выпадашке — внутри `<div style="...position: absolute;
#   z-index: 1948;...">`. По наличию z-index в style любого предка
#   надёжно отличаем «в шапке» от «в выпадашке».
#
# ============================================================

# XPath-предикат «предок имеет z-index в inline-style» — попап
_IN_POPUP = "ancestor::div[contains(@style, 'z-index')]"


def _current_account_text(driver):
    """ФИО текущей учётки из шапки. Шапка содержит ОДИН видимый span с
    id='app-acc-info', НЕ внутри popup'а (нет z-index родителя)."""
    try:
        spans = driver.find_elements(By.XPATH,
            f"//span[@id='app-acc-info' and not({_IN_POPUP})]")
        for s in spans:
            try:
                if s.is_displayed():
                    return (s.text or '').strip()
            except Exception:
                continue
    except Exception:
        pass
    return ''


def _account_active(driver, target_substring, timeout=0):
    """Проверяет, что в шапке отображается учётка с target_substring.

    timeout > 0 — поллит каждые 0.5с, пока шапка не отрендерится. Под Edge 148
    АСУД иногда показывает ФИО юзера в шапке через 5-8с после wait_asud_loaded
    — без ожидания _account_active отдаст False и прога ложно решит, что
    надо переключаться (а в шапке уже та учётка).
    """
    end = time.monotonic() + max(timeout, 0)
    first_pass = True
    while first_pass or time.monotonic() < end:
        first_pass = False
        cur = _current_account_text(driver)
        if cur and target_substring.lower() in cur.lower():
            return True
        if time.monotonic() >= end:
            break
        time.sleep(0.5)
    return False


def _find_account_in_popup(driver, target_substring):
    """Возвращает span с учёткой target_substring внутри открытой выпадашки
    (находится внутри div с z-index в inline-style). None если не найден."""
    try:
        spans = driver.find_elements(By.XPATH,
            f"//span[@id='app-acc-info' and contains(., '{target_substring}') "
            f"and {_IN_POPUP}]")
        for s in spans:
            try:
                if s.is_displayed():
                    return s
            except Exception:
                continue
    except Exception:
        pass
    return None


def _popup_accounts(driver):
    """Список текстов всех учёток в открытой выпадашке. Для диагностики."""
    out = []
    try:
        spans = driver.find_elements(By.XPATH,
            f"//span[@id='app-acc-info' and {_IN_POPUP}]")
        for s in spans:
            try:
                if s.is_displayed():
                    out.append((s.text or '').strip())
            except Exception:
                continue
    except Exception:
        pass
    return out


def switch_account(driver, target_substring):
    """Переключение учётки через выпадашку профиля.

    1. _account_active(timeout=10) — если уже на нужной учётке, выходим
    2. Клик по #show-another-user-icon (стрелка-триггер выпадашки)
    3. Ждём появления target_substring в выпадашке (in_popup)
    4. Клик по спану учётки
    5. Пост-верификация: _account_active(timeout=30)
    """
    log.info(f"Переключение на учётку: {target_substring}")

    if _account_active(driver, target_substring, timeout=10):
        log.info(f"Уже под учёткой '{target_substring}' — пропускаю переключение")
        return True

    current = _current_account_text(driver)
    log.info(f"Текущая учётка в шапке: '{current or '?'}'")

    # 1. Клик по стрелке-триггеру выпадашки. ВАЖНО: НЕ #app-acc-icon-myprofile
    # (та — просто аватарка пользователя, клик ничего не открывает).
    try:
        trigger = driver.find_element(By.ID, "show-another-user-icon")
    except Exception:
        log.error("Триггер #show-another-user-icon не найден — старый АСУД?")
        return False
    if not trigger.is_displayed():
        log.error("Триггер #show-another-user-icon невидим")
        return False
    click(driver, trigger, "стрелка «Выбрать учётную запись»")

    # 2. Ждём появления нужной учётки в выпадашке
    try:
        WebDriverWait(driver, 10).until(
            lambda d: _find_account_in_popup(d, target_substring) is not None)
    except Exception:
        available = _popup_accounts(driver)
        log.error(f"Учётка '{target_substring}' не найдена в выпадашке. "
                  f"Доступные в выпадашке: {available or '(пусто — возможно, выпадашка не открылась)'}")
        return False

    # 3. Клик по найденной учётке
    target = _find_account_in_popup(driver, target_substring)
    click(driver, target, f"учётка {target_substring}")
    log.info("Клик по учётке, жду перезагрузку АСУД")
    _wait_profile_loaded(driver)

    # 4. Пост-верификация
    if _account_active(driver, target_substring, timeout=30):
        log.info(f"Учётка переключена на '{target_substring}'")
        return True
    log.error(f"Учётка НЕ переключилась на '{target_substring}'. "
              f"В шапке сейчас: '{_current_account_text(driver) or '?'}'")
    return False


# ============================================================
# UI: сайдбар → "На резолюцию"
# ============================================================

# Маппинг названий sidebar-пунктов на стабильные id (из HTML-дампа).
# Pattern: id="CABINET_MENU__<category>__<subcategory>" на <table>-элементе.
SIDEBAR_IDS = {
    "На резолюцию":               "CABINET_MENU__RECEIVED__ALL_ACTIVE__TO_RESOLUTION",
    "Исполнение":                 "CABINET_MENU__RECEIVED__ALL_ACTIVE__TO_EXECUTION",
    "На исполнение":              "CABINET_MENU__RECEIVED__ALL_ACTIVE__TO_EXECUTION",
    "Согласование":               "CABINET_MENU__RECEIVED__ALL_ACTIVE__APPROVAL",
    "Подпись":                    "CABINET_MENU__RECEIVED__ALL_ACTIVE__SIGN",
    "Регистрация":                "CABINET_MENU__RECEIVED__ALL_ACTIVE__REGISTRATION",
    "Утверждение":                "CABINET_MENU__RECEIVED__ALL_ACTIVE__CONFIRMATION",
    "Контроль исполнения":        "CABINET_MENU__RECEIVED__ALL_ACTIVE__EXECUTION_CONTROL",
    "Снятие с контроля":          "CABINET_MENU__RECEIVED__ALL_ACTIVE__REMOVE_FROM_CONTROL",
    "Доработка":                  "CABINET_MENU__RECEIVED__ALL_ACTIVE__REVISION",
    "Проверка оформления":        "CABINET_MENU__RECEIVED__ALL_ACTIVE__TO_DECORATION_CHECK",
    "Внесение":                   "CABINET_MENU__RECEIVED__ALL_ACTIVE__TO_INTRODUCTION",
    "Рассмотрение":               "CABINET_MENU__RECEIVED__ALL_ACTIVE__TO_REVIEW",
    "Голосование":                "CABINET_MENU__RECEIVED__ALL_ACTIVE__TO_VOTING",
    "Подготовка рекомендаций":    "CABINET_MENU__RECEIVED__ALL_ACTIVE__PREPARATION_OF_RECOMMENDATIONS",
    "Подтверждение идентичности": "CABINET_MENU__RECEIVED__ALL_ACTIVE__PROOF_IDENTITY",
    "Запросы на аннулирование":   "CABINET_MENU__RECEIVED__ALL_ACTIVE__ANNULATION_REQUESTS",
    "Верификация":                "CABINET_MENU__RECEIVED__ALL_ACTIVE__VERIFICATION",
    "Печать":                     "CABINET_MENU__RECEIVED__ALL_ACTIVE__PRINT",
    "Прикрепление оригинала":     "CABINET_MENU__RECEIVED__ALL_ACTIVE__ATTACHING_ORIGINAL",
    "Завершённые":                "CABINET_MENU__RECEIVED__FINISHED",
    "Завершенные":                "CABINET_MENU__RECEIVED__FINISHED",
    "Черновики":                  "CABINET_MENU__MY_FOLDER__DRAFTS",
    "Избранное":                  "CABINET_MENU__MY_FOLDER__FAVORITES",
    "В работе":                   "CABINET_MENU__MY_FOLDER__IN_WORK",
    "Просмотренные":              "CABINET_MENU__MY_FOLDER__VIEWED",
    "Шаблоны документов":         "CABINET_MENU__ANONGRP__DOC_TEMPLATES",
    "Корзина":                    "CABINET_MENU__ANONGRP__TRASH",
}


def click_sidebar_section(driver, section_text):
    """Клик по пункту в левом сайдбаре.

    Приоритет — стабильный id из SIDEBAR_IDS (по HTML-дампу). Fallback —
    text-search по xpath. Поллит до 10 секунд.
    """
    log.info(f"Сайдбар → '{section_text}'")
    target = None
    end = time.monotonic() + 10

    sidebar_id = SIDEBAR_IDS.get(section_text)
    if sidebar_id:
        while time.monotonic() < end and not target:
            try:
                el = driver.find_element(By.ID, sidebar_id)
                if el.is_displayed():
                    target = el
                    log.debug(f"  найден по id: {sidebar_id}")
                    break
            except Exception:
                pass
            time.sleep(0.3)

    if not target:
        items = driver.find_elements(By.XPATH,
            f"//*[normalize-space(text())='{section_text}']")
        for it in items:
            try:
                if it.is_displayed():
                    target = it
                    break
            except Exception:
                continue
    if not target:
        log.error(f"Пункт сайдбара '{section_text}' не найден")
        return False
    click(driver, target, f"сайдбар: {section_text}")
    try:
        WebDriverWait(driver, 10).until(
            lambda d: len(d.find_elements(By.XPATH, DATA_ROW_XPATH)) > 0)
    except Exception:
        log.debug("Грид пустой за 10s — может быть нормально (нет задач)")
    return True


# ============================================================
# UI: поиск и открытие документа в списке "На резолюцию"
# ============================================================

# GXT-сетка АСУД: <tr> с обоими классами obj-list-rec и obj-list-task —
# это строки данных. Заголовки/служебные tr этих классов не имеют.
# Раньше использовался id "CABINET_MENU__RECEIVED__ALL_ACTIVE__TO_RESOLUTION",
# но это id сайдбар-пункта, а не таблицы → находили tr меню вместо данных.
DATA_ROW_XPATH = ("//tr[contains(concat(' ',normalize-space(@class),' '),' obj-list-rec ')"
                  " and contains(concat(' ',normalize-space(@class),' '),' obj-list-task ')]")

# Фильтр-input под колонкой "Номер" — стабильный id из DOM
NUMBER_FILTER_INPUT_ID = "FCPC_Номер-input"
NUMBER_FILTER_CONTAINER_ID = "FCPC_Номер"


def _set_filter_value(driver, container_id, input_id, value):
    """JS-ввод в фильтр колонки — без эмуляции клавиатуры (фон-friendly)."""
    log.debug(f"_set_filter_value: ищу input id={input_id!r}")
    inp = None
    try:
        inp = driver.find_element(By.ID, input_id)
        log.debug(f"  input найден по ID")
    except Exception:
        log.debug(f"  по ID не нашёл, пробую внутри container={container_id!r}")
        try:
            container = driver.find_element(By.ID, container_id)
            inp = container.find_element(By.CSS_SELECTOR, "input[type='text']")
            log.debug(f"  input найден внутри container")
        except Exception as e:
            log.warning(f"Фильтр {container_id}: input не найден ({e})")
            return False
    try:
        driver.execute_script("""
            var el = arguments[0], v = arguments[1];
            el.focus();
            el.value = v;
            el.dispatchEvent(new Event('input', {bubbles:true}));
            el.dispatchEvent(new Event('keyup', {bubbles:true}));
            el.dispatchEvent(new Event('change', {bubbles:true}));
        """, inp, value)
        log.debug(f"  JS dispatch выполнен; value={value!r}")
        return True
    except Exception as e:
        log.warning(f"JS-ввод в фильтр упал: {e}")
        return False


def filter_by_number(driver, asud_id):
    """Вбивает ОПТС/ОРТС-номер в фильтр колонки 'Номер'.
    Возвращает True если удалось ввести."""
    log.info(f"Фильтр 'Номер' = {asud_id}")
    return _set_filter_value(driver, NUMBER_FILTER_CONTAINER_ID,
                              NUMBER_FILTER_INPUT_ID, asud_id)


def clear_filter(driver):
    """Очищает фильтр колонки 'Номер'."""
    _set_filter_value(driver, NUMBER_FILTER_CONTAINER_ID,
                       NUMBER_FILTER_INPUT_ID, "")
    time.sleep(0.5)


def find_doc_row(driver, doc, timeout=8):
    """Находит <tr> в таблице списка после применения фильтра.

    Стратегия:
      1. Если есть doc['asud_id'] — вбить в фильтр 'Номер' → взять первую строку
      2. Иначе — поиск по тексту (appeal_no / TextBody / Subject) как раньше
    """
    asud_id = doc.get('asud_id') or ''
    log.debug(f"find_doc_row: asud_id={asud_id!r}, appeal_no={doc.get('appeal_no')!r}, "
             f"timeout={timeout}s")

    # Главная стратегия: фильтр по точному номеру
    if asud_id:
        log.info(f"[find] стратегия 1 (фильтр 'Номер'): {asud_id}")
        if filter_by_number(driver, asud_id):
            log.debug(f"  фильтр введён, жду дебаунс 1.5s")
            time.sleep(1.5)  # дебаунс GWT-фильтра
            end = time.monotonic() + timeout
            tick = 0
            while time.monotonic() < end:
                tick += 1
                try:
                    rows = driver.find_elements(By.XPATH, DATA_ROW_XPATH)
                    visible_rows = [r for r in rows if r.is_displayed()
                                    and (r.text or '').strip()]
                    log.debug(f"  tick#{tick}: всего obj-list-rec.obj-list-task <tr>="
                             f"{len(rows)}, видимых-с-текстом={len(visible_rows)}")
                    if visible_rows:
                        first_text = (visible_rows[0].text or '').replace('\n', ' ')[:80]
                        log.info(f"[find] МАТЧ по фильтру: {asud_id} → {len(visible_rows)} строк")
                        log.debug(f"  первая строка: {first_text!r}")
                        return visible_rows[0]
                except Exception as e:
                    log.debug(f"  tick#{tick}: исключение {e}")
                time.sleep(0.5)
            log.warning(f"[find] фильтр {asud_id} → 0 строк за {timeout}s")
            return None
        else:
            log.warning(f"[find] не получилось ввести фильтр для {asud_id}")

    # Fallback на текстовый поиск (без фильтра — сканируем что в DOM)
    log.info(f"[find] стратегия 2 (fallback по тексту)")
    end = time.monotonic() + timeout
    tick = 0
    while time.monotonic() < end:
        tick += 1
        if doc.get('appeal_no'):
            try:
                row = driver.find_element(By.XPATH,
                    f"{DATA_ROW_XPATH}[contains(., '{doc['appeal_no']}')]")
                if row.is_displayed():
                    log.info(f"[find] МАТЧ (fallback): appeal № {doc['appeal_no']}")
                    return row
            except Exception:
                pass
        body = (doc.get('textbody') or '').replace('\xa0', ' ').strip()
        snippet = re.sub(r'\s+', ' ', body)[:60].strip().replace("'", "")
        if len(snippet) >= 20:
            try:
                row = driver.find_element(By.XPATH,
                    f"{DATA_ROW_XPATH}[contains(., \"{snippet}\")]")
                if row.is_displayed():
                    log.info(f"[find] МАТЧ (fallback): TextBody {snippet!r}")
                    return row
            except Exception:
                pass
        log.debug(f"  fallback tick#{tick}: не нашёл")
        time.sleep(0.5)
    log.warning(f"[find] не найдено ни одной стратегией за {timeout}s")
    return None


def _card_opened(driver, timeout):
    """Признак открывшейся карточки — кнопка 'Создать резолюцию'."""
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, "header-action-btn-add_resolution")))
        return True
    except Exception:
        return False


def _refresh_first_row(driver):
    """Возвращает свежий <tr> первой видимой строки таблицы (после фильтра).

    Хитрость GXT: после фильтра старая ссылка на <tr> часто становится
    stale — пересортировка перерисовывает грид.
    """
    try:
        rows = driver.find_elements(By.XPATH, DATA_ROW_XPATH)
        log.debug(f"_refresh_first_row: data <tr>={len(rows)}")
        for idx, r in enumerate(rows):
            try:
                if r.is_displayed() and (r.text or '').strip():
                    cls_preview = (r.get_attribute('class') or '')[:80]
                    log.debug(f"  → свежий ref на data-<tr>[{idx}], class={cls_preview!r}")
                    return r
            except Exception:
                continue
        log.debug(f"  → ни одной видимой data-<tr> с текстом")
    except Exception as e:
        log.debug(f"  ошибка поиска data-<tr>: {e}")
    return None


def _meaningful_cell(row):
    """Берёт ячейку с текстом (subject/тип) — не чекбокс, не иконку."""
    try:
        tds = row.find_elements(By.XPATH, ".//td")
    except Exception as e:
        log.debug(f"_meaningful_cell: не нашёл <td>: {e}")
        return None
    log.debug(f"_meaningful_cell: всего <td>={len(tds)}")
    best = None
    for idx, td in enumerate(tds):
        try:
            if not td.is_displayed():
                continue
            txt = (td.text or '').strip()
            w = td.size.get('width', 0)
            if len(txt) > 10:  # пропускаем чекбоксы/иконки
                log.debug(f"  → выбрана <td>[{idx}] w={w} text={txt[:40]!r}")
                return td
            if best is None and w > 50:
                best = td
                log.debug(f"  кандидат <td>[{idx}] w={w} (без текста)")
        except Exception:
            continue
    log.debug(f"  → fallback на best={'найден' if best else 'нет'}")
    return best


def open_doc_card(driver, row):
    """Открывает карточку документа.

    GXT-сетка обычно реагирует так: одиночный клик выделяет строку
    (добавляет класс rowSelected), а двойной открывает карточку.
    ActionChains.double_click по разным причинам часто не срабатывает,
    поэтому пробуем несколько стратегий по очереди.
    """
    log.info("[open] начинаю открывать карточку документа")
    # Свежая ссылка — пред. могла стать stale
    fresh = _refresh_first_row(driver) or row
    log.debug(f"[open] fresh={'свежая ссылка' if fresh is not row else 'та же что передали'}")

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", fresh)
        time.sleep(0.3)
        log.debug(f"[open] scrollIntoView OK")
    except Exception as e:
        log.debug(f"[open] scrollIntoView err: {e}")

    # Берём содержательную ячейку (Subject/Тип) — не чекбокс
    target_cell = _meaningful_cell(fresh) or fresh
    log.debug(f"[open] target_cell={'<td>' if target_cell is not fresh else '<tr>'}")

    # ── Стратегия 1: single click + Enter (самый надёжный для GXT)
    log.info("[open] strat1: single click + Enter")
    try:
        ActionChains(driver).move_to_element(target_cell).pause(0.15).click().perform()
        log.debug(f"  click выполнен")
        try:
            cls = (fresh.get_attribute('class') or '')
            log.debug(f"  class у <tr> после клика: {cls!r}")
        except Exception:
            pass
        time.sleep(0.4)
        ActionChains(driver).send_keys(Keys.ENTER).perform()
        log.debug(f"  Enter отправлен")
    except Exception as e:
        log.debug(f"  strat1 err: {e}")
    if _card_opened(driver, timeout=5):
        log.info("[open] УСПЕХ: strat1 click + Enter")
        return True
    log.info("[open] strat1 не сработал, пробую strat2")

    # ── Стратегия 2: ActionChains double-click по содержательной ячейке
    log.info("[open] strat2: ActionChains dblclick по <td>")
    try:
        fresh2 = _refresh_first_row(driver) or fresh
        cell2 = _meaningful_cell(fresh2) or fresh2
        ActionChains(driver).move_to_element(cell2).pause(0.2).double_click().perform()
        log.debug(f"  double_click выполнен")
    except Exception as e:
        log.debug(f"  strat2 err: {e}")
    if _card_opened(driver, timeout=4):
        log.info("[open] УСПЕХ: strat2 ActionChains dblclick")
        return True
    log.info("[open] strat2 не сработал, пробую strat3")

    # ── Стратегия 3: JS — полная mouse-event цепочка (mousedown/up/click x2 + dblclick)
    log.info("[open] strat3: JS mouse-event chain")
    try:
        fresh3 = _refresh_first_row(driver) or fresh
        cell3 = _meaningful_cell(fresh3) or fresh3
        driver.execute_script("""
            const el = arguments[0];
            const r = el.getBoundingClientRect();
            const x = r.left + r.width/2, y = r.top + r.height/2;
            const opts = {bubbles:true, cancelable:true, view:window,
                          button:0, buttons:1, clientX:x, clientY:y};
            for (const t of ['mousedown','mouseup','click']) {
                el.dispatchEvent(new MouseEvent(t, {...opts, detail:1}));
            }
            for (const t of ['mousedown','mouseup','click']) {
                el.dispatchEvent(new MouseEvent(t, {...opts, detail:2}));
            }
            el.dispatchEvent(new MouseEvent('dblclick', {...opts, detail:2}));
        """, cell3)
        log.debug(f"  JS event chain dispatched")
    except Exception as e:
        log.debug(f"  strat3 err: {e}")
    if _card_opened(driver, timeout=4):
        log.info("[open] УСПЕХ: strat3 JS event chain")
        return True
    log.info("[open] strat3 не сработал, пробую strat4")

    # ── Стратегия 4: клик по ссылке/anchor если есть
    log.info("[open] strat4: click по <a>/anchor внутри строки")
    try:
        fresh4 = _refresh_first_row(driver) or fresh
        links = fresh4.find_elements(By.XPATH,
            ".//a | .//*[contains(@class,'gwt-Anchor')] | .//*[contains(@class,'cellClickable')]")
        log.debug(f"  найдено anchor/link={len(links)}")
        for idx, lnk in enumerate(links):
            if lnk.is_displayed():
                log.debug(f"  кликаю anchor[{idx}] tag={lnk.tag_name}")
                try:
                    driver.execute_script("arguments[0].click();", lnk)
                except Exception:
                    try:
                        lnk.click()
                    except Exception:
                        continue
                if _card_opened(driver, timeout=4):
                    log.info(f"[open] УСПЕХ: strat4 link-click anchor[{idx}]")
                    return True
    except Exception as e:
        log.debug(f"  strat4 err: {e}")

    log.warning("[open] ПРОВАЛ: все 4 стратегии не сработали")
    return False


# ============================================================
# UI: диалог "Корневая резолюция"
# ============================================================

def _wait_data_value(driver, container, target_value="true", timeout=5):
    """Ждёт пока у контейнера-тоггла data-value станет target_value."""
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        try:
            if container.get_attribute('data-value') == target_value:
                return True
        except Exception:
            pass
        time.sleep(0.2)
    return False


# Маппинг русских лейблов на стабильные id контейнеров тогглов.
# По HTML-дампу: каждый тоггл в форме резолюции лежит в div'е с id
# 'asudik-form-need_report' (или need_control), внутри — два видимых
# div[data-value][class*=switcherContainer], текущее значение = data-value.
# Раньше поиск шёл по xpath-text лейбла, но АСУД пишет «Требуется отчет»
# БЕЗ ё, а в коде искалось С ё — тоггл не находился.
_TOGGLE_CONTAINERS = {
    "Требуется отчёт":       "asudik-form-need_report",
    "Требуется отчет":       "asudik-form-need_report",
    "Контрольная резолюция": "asudik-form-need_control",
}


def toggle_switch(driver, label_text, target_value="true"):
    """Переключает тоггл в нужное состояние.

    Сначала пробует найти контейнер по стабильному id (для известных лейблов),
    fallback — поиск по xpath-text (для лейблов которых нет в маппинге).
    """
    container = None
    container_id = _TOGGLE_CONTAINERS.get(label_text)
    if container_id:
        try:
            wrapper = driver.find_element(By.ID, container_id)
            container = wrapper.find_element(By.CSS_SELECTOR,
                "div[data-value][class*='switcherContainer']")
        except Exception as e:
            log.debug(f"Тоггл '{label_text}' по id {container_id!r} не найден: {e}")

    if container is None:
        # Fallback на старую логику: поиск по тексту лейбла
        try:
            label = driver.find_element(By.XPATH,
                f"//*[normalize-space(text())='{label_text}']")
            container = label.find_element(By.XPATH,
                "./following::*[contains(@class,'switcherContainer')][1]")
        except Exception as e:
            log.warning(f"Тоггл '{label_text}' не найден (ни по id, ни по тексту): {e}")
            return False

    cur = container.get_attribute('data-value')
    if cur == target_value:
        log.info(f"Тоггл '{label_text}' уже = {target_value}")
        return True

    click(driver, container, f"тоггл {label_text} → {target_value}")
    if _wait_data_value(driver, container, target_value, timeout=3):
        log.info(f"Тоггл '{label_text}' = {target_value}")
        return True
    log.warning(f"Тоггл '{label_text}' не переключился (data-value={container.get_attribute('data-value')})")
    return False


def _xpath_lit(s):
    """Безопасно квотит строку для XPath (учитывая возможные апострофы/кавычки)."""
    if "'" not in s:
        return f"'{s}'"
    if '"' not in s:
        return f'"{s}"'
    parts = s.split("'")
    return "concat(" + ", \"'\", ".join(f"'{p}'" for p in parts) + ")"


def select_content_template(driver, template_text):
    """Выбирает в поле «Содержание» пункт из выпадашки шаблонов.

    По HTML-дампу каждый пункт раскрытой выпадашки — это
    <div data-marker="<template_text>">. Стабильнее text-match'а — не
    путается с другими элементами с тем же текстом в DOM (например,
    шаблон уже стоит в открытой карточке).
    """
    try:
        # 1. Клик в поле «Содержание» чтобы раскрыть выпадашку
        inp = None
        candidates = driver.find_elements(By.CSS_SELECTOR,
            "input[placeholder='Общие формулировки']")
        for c in candidates:
            if c.is_displayed():
                inp = c
                break
        if not inp:
            inp = find_input_near_label(driver, "Содержание")
        if not inp:
            log.error("Поле 'Содержание' не найдено")
            return False
        click(driver, inp, "Содержание")

        # 2. Ждём появления элемента выпадашки с нужным data-marker'ом
        target_xpath = f"//div[@data-marker={_xpath_lit(template_text)}]"
        try:
            WebDriverWait(driver, 5).until(
                lambda d: any(e.is_displayed() for e in d.find_elements(By.XPATH, target_xpath)))
        except Exception:
            log.debug(f"Не нашёл по data-marker '{template_text}', пробую по тексту")
            target_xpath = f"//*[normalize-space(text())={_xpath_lit(template_text)}]"

        items = driver.find_elements(By.XPATH, target_xpath)
        target = None
        for it in items:
            try:
                if it.is_displayed():
                    target = it
                    break
            except Exception:
                continue
        if not target:
            log.error(f"Пункт '{template_text}' в выпадашке 'Содержание' не найден")
            return False
        click(driver, target, f"Содержание: {template_text}")
        time.sleep(0.5)
        return True
    except Exception as e:
        log.error(f"Ошибка выбора шаблона содержания: {e}")
        return False


def add_business_days(start, days):
    cur = start
    added = 0
    while added < days:
        cur += timedelta(days=1)
        if cur.weekday() < 5:  # 0-4 = пн-пт
            added += 1
    return cur


def compute_control_date(planned_date_str, n_workdays, receipt_date_str=None):
    """Возвращает строку DD.MM.YYYY для поля «Контрольный этап».

    Приоритет (C+D):
      1) planned_date_str (xlsx «Планируемая дата» = срок, вычисленный P1
         как min(получ+17раб, срок ГИС)). Если распарсилось:
            • дата в будущем → используем ЕЁ
            • дата сегодня/в прошлом (ПРОСРОЧКА, фикс C) → НЕ затираем на
              комфортный today+N, ставим ближайший рабочий день (today+1раб)
              + WARNING. Резолюция получает сигнал срочности, не теряем срок.
      2) Если «Планируемая дата» пуста — пересчитываем из receipt_date_str
         (xlsx «Дата получения») той же формулой compute_deadline (фикс D).
      3) Совсем нет данных → старый fallback settings.stage_date_mode/days
         (today + N) от даты запуска.
    """
    today = date.today()

    # 1) Готовая планируемая дата из реестра
    d = parse_ddmmyyyy(planned_date_str)
    # 2) Пусто → пересчитать из даты получения (D)
    if not d and receipt_date_str:
        recomputed = compute_deadline(receipt_date_str, None, n_workdays)
        d = parse_ddmmyyyy(recomputed)
        if d:
            log.info(f"Срок пересчитан из даты получения {receipt_date_str}: {recomputed}")

    if d:
        if d > today:
            return d.strftime("%d.%m.%Y")
        # Просрочка / сегодня (C): минимальный будущий срок, не today+N
        nxt = add_working_days(today, 1)
        log.warning(f"Срок {d.strftime('%d.%m.%Y')} просрочен/сегодня — "
                    f"ставлю ближайший рабочий день {nxt.strftime('%d.%m.%Y')}")
        return nxt.strftime("%d.%m.%Y")

    # 3) Нет ни планируемой, ни даты получения → старый fallback today+N
    mode = settings.get("stage_date_mode", cfg.DEFAULTS["stage_date_mode"])
    n_days = settings.get("stage_date_days") or n_workdays
    if mode == "calendar":
        return (today + timedelta(days=n_days)).strftime("%d.%m.%Y")
    return add_business_days(today, n_days).strftime("%d.%m.%Y")


def set_stage_date(driver, n_workdays, planned_date=None, receipt_date=None):
    """Заполняет дату в поле 'Контрольный этап'.

    planned_date (DD.MM.YYYY) — приоритетный источник (срок из реестра).
    receipt_date (DD.MM.YYYY) — дата получения, для пересчёта если planned пуст.
    Логика выбора — в compute_control_date (C+D).
    """
    deadline = compute_control_date(planned_date, n_workdays, receipt_date)
    if planned_date or receipt_date:
        log.info(f"Срок: planned={planned_date!r} receipt={receipt_date!r} "
                 f"→ используется {deadline}")
    # Приоритет — внутри #asudik-form-control_stage (стабильный id из дампа).
    inp = None
    try:
        container = driver.find_element(By.ID, "asudik-form-control_stage")
        cands = container.find_elements(By.CSS_SELECTOR, "input[type='text']")
        for c in cands:
            if c.is_displayed():
                inp = c
                break
    except Exception:
        pass
    if not inp:
        try:
            inp = driver.find_element(By.CSS_SELECTOR,
                "input[id*='stage_control_date']")
        except Exception:
            pass
    if not inp:
        log.warning("Поле даты этапа не найдено (ни в asudik-form-control_stage, "
                    "ни по id*='stage_control_date')")
        return False
    try:
        js_set_value(driver, inp, deadline)
        log.info(f"Контрольный этап: {deadline}")
        return True
    except Exception as e:
        log.warning(f"Ошибка установки даты: {e}")
        return False


def fill_executor(driver, fio):
    """Вбивает ФИО в поле 'Исполнитель' (combobox), выбирает из выпадашки.

    Приоритет — стабильный id #select_combobox-input (по HTML-дампу).
    Fallback — find_input_near_label("Исполнитель").
    """
    try:
        inp = None
        try:
            inp = driver.find_element(By.ID, "select_combobox-input")
            if not inp.is_displayed():
                inp = None
        except Exception:
            pass
        if not inp:
            inp = find_input_near_label(driver, "Исполнитель")
        if not inp:
            log.error("Поле 'Исполнитель' не найдено (ни по id select_combobox-input, ни по label)")
            return False

        surname = fio.split()[0]
        inp.click()
        inp.clear()
        for ch in surname:
            inp.send_keys(ch)
        # Debounce: GXT-combobox дёргает серверный autocomplete по событию
        # ключевой паузы. Без задержки send_keys возвращается мгновенно,
        # запрос ещё не ушёл, и мы стартуем поиск кандидатов в пустом DOM.
        time.sleep(0.3)
        log.info(f"Введена фамилия: {surname}")

        # Кандидаты в выпадашке. ВАЖНО:
        # 1) НЕ ищем по `contains(text(),...)` без фильтра — surname может
        #    встречаться в лейблах/заголовках уже открытой карточки и
        #    WebDriverWait моментально пройдёт по ложному кандидату.
        # 2) Фильтр: текст должен быть похож на ФИО (≥2 кириллических
        #    слова с большой буквы) — это отсекает лейблы вроде
        #    «Халецкая (на резолюцию)» из шапки.
        # 3) Таймаут 15с — у АСУД серверный autocomplete иногда 6-10с после
        #    переключения учётки/холодного старта сессии.
        # 4) НЕТ Enter-fallback — Enter в открытой выпадашке закрывает её
        #    и фиксирует пустое значение, после чего любой повторный
        #    ожидающий поиск «зависает» в принципе (ровно тот баг что
        #    юзер видит).
        import re as _re
        # Слово ФИО: либо полное «Иванова», либо инициал «И.» или «И»
        _fio_word = _re.compile(r'^[А-ЯЁ](?:[а-яё]+|\.?)$')
        def _looks_like_fio(text):
            txt = (text or '').strip()
            if not txt or len(txt) > 120:
                return False
            # Делим по пробелам и точкам (чтобы «Ю.В.» распался на «Ю.» и «В.»)
            words = [w for w in _re.split(r'\s+|(?<=\.)', txt) if w.strip(' .')]
            named = sum(1 for w in words if _fio_word.match(w.rstrip('.') + ('.' if w.endswith('.') else '')))
            return named >= 2

        def _candidates():
            results = driver.find_elements(By.XPATH,
                f"//*[contains(text(),'{surname}')]")
            out = []
            for r in results:
                if r == inp or r.tag_name.lower() == 'input':
                    continue
                try:
                    if not r.is_displayed():
                        continue
                    txt = (r.text or '').strip()
                    if not _looks_like_fio(txt):
                        continue
                    out.append(r)
                except Exception:
                    continue
            return out

        candidates = []
        try:
            WebDriverWait(driver, 15).until(lambda d: len(_candidates()) > 0)
            candidates = _candidates()
        except Exception:
            log.error(f"Выпадашка 'Исполнитель' не появилась за 15с после ввода '{surname}'")
            return False

        log.info(f"Кандидатов: {len(candidates)}")
        target = None
        for idx, r in enumerate(candidates, 1):
            try:
                txt = (r.text or '').strip()
                if len(txt) > 200:
                    continue
                ok = match_correspondent(txt, fio)
                preview = txt.replace('\n', ' ')[:80]
                log.info(f"  [{idx}] {'OK' if ok else '--'} | {preview!r}")
                if ok and target is None:
                    target = r
            except Exception:
                continue

        if not target:
            log.error(f"Исполнитель '{fio}' не найден в выпадашке")
            return False

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
        ActionChains(driver).move_to_element(target).pause(0.2).click().perform()
        log.info(f"Исполнитель выбран: {fio}")
        return True
    except Exception as e:
        log.error(f"Ошибка ввода исполнителя: {e}")
        return False


def _wait_button_enabled(driver, btn_id, timeout=15):
    """Ждёт пока кнопка с data-disabled='1' не станет активной."""
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        try:
            btn = driver.find_element(By.ID, btn_id)
            if btn.get_attribute('data-disabled') != '1' and btn.is_displayed():
                return btn
        except Exception:
            pass
        time.sleep(0.3)
    return None


def _click_confirm_yes(driver, timeout=10):
    """Ждёт и кликает 'Да' в confirm-диалоге АСУД (с fallback'ами).

    Используется после 'Сохранить и отправить' (подтверждение
    отправки адресатам) — аналогично confirm после 'На резолюцию'
    в основных скриптах.
    """
    yes_btn = None
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        # 1) Точный id
        try:
            btn = driver.find_element(By.ID, "confirm_dialog_btn_yes")
            if btn.is_displayed():
                yes_btn = btn
                break
        except Exception:
            pass
        # 2) Substring id (GWT может префиксовать)
        try:
            btn = driver.find_element(By.CSS_SELECTOR,
                "[id*='confirm_dialog_btn_yes'], [id*='confirm'][id*='yes']")
            if btn.is_displayed():
                yes_btn = btn
                break
        except Exception:
            pass
        # 3) По тексту "Да"
        try:
            for b in driver.find_elements(By.XPATH,
                    "//*[normalize-space(text())='Да']"):
                if b.is_displayed():
                    yes_btn = b
                    break
        except Exception:
            pass
        if yes_btn:
            break
        time.sleep(0.5)

    if not yes_btn:
        log.warning("Confirm-диалог 'Да' не появился")
        return False

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", yes_btn)
        time.sleep(0.3)
    except Exception:
        pass
    clicked = False
    try:
        ActionChains(driver).move_to_element(yes_btn).pause(0.3).click().perform()
        log.info(f"Клик 'Да' (ActionChains): id={yes_btn.get_attribute('id')}")
        clicked = True
    except Exception:
        pass
    if not clicked:
        try:
            driver.execute_script("arguments[0].click();", yes_btn)
            log.info("Клик 'Да' (JS)")
            clicked = True
        except Exception:
            pass
    if not clicked:
        try:
            yes_btn.click()
            log.info("Клик 'Да' (native)")
            clicked = True
        except Exception:
            pass
    if clicked:
        try:
            ActionChains(driver).send_keys(Keys.ENTER).perform()
        except Exception:
            pass
        # Ждём пока confirm-кнопка исчезнет (диалог закрылся)
        try:
            WebDriverWait(driver, 5).until_not(EC.visibility_of(yes_btn))
        except Exception:
            pass
    return clicked


def submit_resolution(driver):
    """Финальный шаг: 'Сохранить и отправить' → confirm 'Да' → закрыть карточку."""
    log.debug("[submit] жду активации 'Сохранить и отправить' (id=save_and_send_btn)")
    btn = _wait_button_enabled(driver, "save_and_send_btn", timeout=15)
    if not btn:
        log.error("Кнопка 'Сохранить и отправить' не активировалась")
        return False
    log.info("[submit] клик 'Сохранить и отправить'")
    click(driver, btn, "Сохранить и отправить")

    # Подтверждение отправки адресатам — _click_confirm_yes сам поллит появление
    log.info("[submit] жду confirm-диалог 'Да'")
    confirmed = _click_confirm_yes(driver, timeout=10)
    log.info(f"[submit] confirm 'Да': {'OK' if confirmed else 'не появился'}")

    # Ждём пока модалка "Корневая резолюция" уйдёт из DOM (или станет невидимой)
    try:
        WebDriverWait(driver, 5).until_not(
            lambda d: any(t.is_displayed() for t in d.find_elements(
                By.XPATH, "//*[contains(text(),'Корневая резолюция')]")))
        log.debug("[submit] модалка 'Корневая резолюция' закрылась штатно")
    except Exception:
        log.warning("[submit] модалка 'Корневая резолюция' ещё открыта — крестик")
        close_open_modals(driver)
    return True


def click_complete_button(driver, timeout=20, max_attempts=3):
    """Клик по кнопке «Завершить» в открытой карточке.

    HTML-дамп показывает структуру:
        <div id="header-action-btn-finish_task">                    ← outer
          <div class="...Css3ButtonStyle-button" tabindex="0">      ← INNER (реально кликабельный)
            <div class="Css3ButtonStyle-buttonInner">Завершить</div>
          </div>
        </div>

    Клик ПО OUTER часто проваливается: GXT привязывает обработчик к INNER
    div'у с tabindex, плюс если АСУД ещё пересчитывает stage_grid после
    submit_resolution — обработчик может быть ещё не привязан. Юзер видел
    в логе «нажата», а на самом деле — нет, и карточка не закрывалась.

    Новая стратегия:
    1) Находим OUTER в DOM
    2) Drill-down на INNER с tabindex
    3) Кликаем INNER, ждём 2с
    4) Верифицируем: outer должен исчезнуть/стать невидимым → успех
    5) Иначе ретраим до max_attempts (АСУД мог ещё дочитываться)
    """
    log.info(f"Ищу кнопку 'Завершить' (timeout={timeout}с, до {max_attempts} попыток)")
    end = time.monotonic() + timeout
    outer = None
    while time.monotonic() < end:
        try:
            outer = driver.find_element(By.ID, "header-action-btn-finish_task")
            break
        except Exception:
            time.sleep(0.5)
    if outer is None:
        log.warning(f"Кнопка 'Завершить' не появилась за {timeout}с — "
                    f"возможно документ уже завершён")
        return False

    # Ждём пока кнопка settle (data-disabled=0, displayed). Макс 8с.
    settle_end = min(time.monotonic() + 8, end)
    while time.monotonic() < settle_end:
        try:
            if outer.is_displayed() and outer.get_attribute("data-disabled") != "1":
                break
        except Exception:
            break
        time.sleep(0.3)

    for attempt in range(1, max_attempts + 1):
        # Drill-down на внутренний кликабельный div
        try:
            inner = outer.find_element(By.CSS_SELECTOR, "div[tabindex]")
        except Exception:
            inner = outer
        log.debug(f"Завершить: попытка {attempt}/{max_attempts}, "
                  f"target={'inner' if inner is not outer else 'outer'}")
        try:
            click(driver, inner, f"Завершить (попытка {attempt})")
        except Exception as e:
            log.warning(f"Завершить click err: {e}")
        # Верификация: должна исчезнуть/стать невидимой
        time.sleep(2)
        try:
            still = driver.find_element(By.ID, "header-action-btn-finish_task")
            visible = still.is_displayed()
        except Exception:
            visible = False
        if not visible:
            log.info(f"Завершить успешно нажата (после {attempt} попыток, кнопка ушла из DOM)")
            return True
        log.warning(f"Завершить: кнопка ВСЁ ЕЩЁ видна после клика {attempt}/{max_attempts} — "
                    f"АСУД ещё догружает? повторяю")
        # Перечитаем outer (мог стать stale)
        try:
            outer = driver.find_element(By.ID, "header-action-btn-finish_task")
        except Exception:
            log.info("Завершить: outer исчез между попытками — считаем что нажалось")
            return True
        time.sleep(1.5)  # дать АСУД договрузить

    log.error(f"Завершить не сработал за {max_attempts} попыток — fallback на close")
    return False


def close_card_after_complete(driver):
    """После клика по 'Завершить' карточка должна закрыться сама.
    Safety net: если осталась открыта — кликаем header-close-btn.
    """
    # Сначала ждём что главная вернулась сама. Timeout увеличен до 12с —
    # АСУД после «Завершить» иногда пересчитывает stage_grid 7-10 секунд
    # прежде чем закрыть карточку. По логу resolutions_20260611_080147_1
    # из 26 успешных «Завершить» только 9 закрывались за 5с, остальные
    # 25 валились в fallback. Большинство из этих 25 — просто долго.
    try:
        WebDriverWait(driver, 12).until(
            EC.element_to_be_clickable((By.ID, "mainscreen-create-button")))
        log.info("[close] карточка закрылась сама после «Завершить»")
        return
    except Exception:
        pass

    # Резерв — клик по крестику. Срабатывает в реально редких случаях
    # когда АСУД завис или Завершить попал в edge-state.
    log.info("[close] карточка ещё открыта после «Завершить» (>12с), fallback на header-close-btn")
    close_card_after_resolution(driver)


def row_has_resolution_to(row_element, executor_fio):
    """True если строка в списке («На резолюцию»/«Исполнение») уже имеет
    в колонке «Направлено» ФИО executor_fio.

    Колонка «Направлено» в гриде АСУД: header id="AF_HH_dss_child_performer".
    В ячейке строки этой колонки текст рендерится как <nobr>Халецкая Ю.В.</nobr>
    (см. HTML-дампы из main grid view).

    БЫСТРАЯ проверка — не требует открытия карточки. Если фамилия там есть,
    значит резолюция уже выдана, можно skip + mark_status + следующая строка.

    Возвращает True если фамилия исполнителя обнаружена в строке.
    """
    if row_element is None or not executor_fio:
        return False
    surname = executor_fio.split()[0]
    if not surname:
        return False
    try:
        # nobr с фамилией внутри row → 99% случай (колонка «Направлено»)
        nobrs = row_element.find_elements(
            By.XPATH, f".//nobr[contains(text(), '{surname}')]")
        for n in nobrs:
            try:
                if n.is_displayed():
                    return True
            except Exception:
                continue
        # Fallback — любой элемент с текстом-фамилией в ячейках грида
        cells = row_element.find_elements(
            By.XPATH, f".//*[contains(text(), '{surname}')]")
        for c in cells:
            try:
                if c.is_displayed() and c.tag_name.lower() in ('nobr', 'span', 'div'):
                    return True
            except Exception:
                continue
    except Exception as e:
        log.debug(f"row_has_resolution_to({executor_fio!r}): {e}")
    return False


def card_has_any_resolution(driver, timeout=3):
    """True если в открытой карточке есть ХОТЬ КАКАЯ-ТО выданная резолюция
    (кому угодно — не обязательно нашему executor'у).

    Признак — наличие хотя бы одного <td> с div.mainExecutorMark («отв.исп.»)
    в панели «Этапы исполнения». Дополнительно проверяется ячейка даты
    выдачи (lfce-stage-grid-column_dsdt_init_date с DD.MM.YYYY).

    Используется в sync-режиме как fallback после row_has_resolution_to:
    если в Направлено колонке списка нашего executor'а не видно, но в
    карточке всё-таки есть какая-то резолюция (выданная вручную человеком
    другому исполнителю, или нашим daemon'ом до восстановления xlsx) —
    отмечаем done в xlsx, чтобы daemon не плодил дубли.
    """
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        try:
            # Любой td с mainExecutorMark = резолюция выдана
            cells = driver.find_elements(
                By.XPATH, "//td[.//div[contains(@class, 'mainExecutorMark')]]")
            for c in cells:
                try:
                    if c.is_displayed():
                        return True
                except Exception:
                    continue
            # Fallback: ячейка даты в stage_grid
            dates = driver.find_elements(
                By.CSS_SELECTOR, "td.lfce-stage-grid-column_dsdt_init_date")
            for d in dates:
                try:
                    txt = (d.text or '').strip()
                    if txt and re.match(r'\d{2}\.\d{2}\.\d{4}', txt) and d.is_displayed():
                        return True
                except Exception:
                    continue
        except Exception as e:
            log.debug(f"card_has_any_resolution: {e}")
        time.sleep(0.5)
    return False


def has_existing_resolution_to(driver, executor_fio, timeout=3):
    """True если в открытой карточке уже есть резолюция на executor_fio.

    По HTML-дампу АСУД: в панели «Этапы исполнения» каждая выданная
    резолюция отрисовывается как <td>, содержащий и фамилию исполнителя,
    и <div class="mainExecutorMark">отв.исп.</div>.

    Поллинг до timeout секунд: панель резолюций иногда подгружается
    асинхронно после остального содержимого карточки (~1-2с после
    open_doc_card). Без поллинга мы проверяли ДО рендера и получали
    ложный False — см. resolutions_20260611_080147.log (ШАГ 2 OK 10.2с,
    XPath сразу пустой, и Шаг 3 уже жмёт «Создать резолюцию»).

    Используется как защита от дублирования: если daemon потерял отметку
    «Отписано Халецкой» (mark_status упал, юзер открыл xlsx в Excel) —
    на следующей итерации обнаружит резолюцию в АСУД → пропустит без
    создания дубля + синхронизирует xlsx.
    """
    if not executor_fio:
        return False
    surname = executor_fio.split()[0] if executor_fio else ''
    if not surname:
        return False
    xpath = (f"//td[contains(., '{surname}') and "
             f".//div[contains(@class, 'mainExecutorMark')]]")
    end = time.monotonic() + timeout
    attempts = 0
    while time.monotonic() < end:
        attempts += 1
        try:
            cells = driver.find_elements(By.XPATH, xpath)
            for c in cells:
                try:
                    if c.is_displayed():
                        log.info(f"  ✓ В карточке найдена резолюция на «{surname}» "
                                 f"(после {attempts} проверок) — пропуск дубля")
                        return True
                except Exception:
                    continue
        except Exception as e:
            log.debug(f"has_existing_resolution_to({executor_fio!r}): {e}")
        time.sleep(0.5)
    log.info(f"  Резолюции на «{surname}» в карточке нет "
             f"(проверял {timeout}с, {attempts} попыток) — продолжаю создание")
    return False


def close_card_after_resolution(driver):
    """После выдачи резолюции возвращаемся в список через #header-close-btn."""
    # Ждём пока header-close-btn появится (карточка всё ещё открыта) —
    # без фикс. sleep(2). Если не появится — карточка уже закрыта.
    close_btn = None
    try:
        close_btn = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "header-close-btn")))
    except Exception:
        log.info("[close] карточка уже закрыта (header-close-btn не появился)")
        return

    if not close_btn.is_displayed():
        log.info("[close] header-close-btn не видим — карточка уже закрыта")
        return

    log.debug("[close] header-close-btn найден")
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", close_btn)
    except Exception:
        pass
    closed = False
    try:
        ActionChains(driver).move_to_element(close_btn).pause(0.2).click().perform()
        closed = True
        log.info("[close] карточка закрыта (ActionChains)")
    except Exception as e:
        log.debug(f"[close] ActionChains err: {e}")
    if not closed:
        try:
            driver.execute_script("arguments[0].click();", close_btn)
            closed = True
            log.info("[close] карточка закрыта (JS click)")
        except Exception as e:
            log.debug(f"[close] JS click err: {e}")

    # Ждём возврата в главную (mainscreen-create-button становится кликабельной)
    if closed:
        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "mainscreen-create-button")))
        except Exception:
            log.debug("[close] главная не загрузилась за 10s")


# ============================================================
# DOCUMENT FLOW (один документ)
# ============================================================

def process_one(driver, doc, index, total):
    """Обработка одной строки реестра: найти, открыть, выдать резолюцию, закрыть."""
    t_start = time.monotonic()
    log.info("=" * 50)
    log.info(f"ДОКУМЕНТ {index}/{total}: link={doc.get('link')!r}, "
             f"appeal_no={doc.get('appeal_no')}, "
             f"исполнитель={doc.get('executor')}")
    log.debug(f"  доп: asud_id={doc.get('asud_id')!r}, ao={doc.get('ao')!r}, "
             f"row_idx={doc.get('row_idx')}")

    if not doc.get('executor'):
        log.warning(f"Row {doc['row_idx']}: исполнитель не определён → пропускаю")
        return False

    # 1. Найти строку в списке
    log.info(f"--- ШАГ 1/12: поиск документа в списке ---")
    row = find_doc_row(driver, doc, timeout=10)
    if not row:
        log.warning(f"Row {doc['row_idx']}: документ в списке не найден → пропускаю")
        return False
    log.debug(f"  Шаг 1 OK ({time.monotonic()-t_start:.1f}s)")

    # 1.5. Быстрая проверка в самой строке: если в колонке «Направлено»
    # (id заголовка AF_HH_dss_child_performer) уже стоит ФИО исполнителя —
    # резолюция выдана. Skip + sync xlsx БЕЗ открытия карточки.
    if row_has_resolution_to(row, doc.get('executor')):
        surname = doc['executor'].split()[0] if doc.get('executor') else '?'
        log.info(f"Row {doc['row_idx']}: в строке списка уже видно «{surname}» "
                 f"в «Направлено» — пропуск + синхрон xlsx (без открытия карточки)")
        xlsx_path = doc.get('_xlsx_path') or settings.get("_xlsx_path")
        status_col = settings.get("_status_column")
        if xlsx_path and status_col and doc.get('asud_id'):
            if mark_status(xlsx_path, doc['asud_id'], status_col):
                log.debug(f"  → xlsx: {doc['asud_id']} помечен «{status_col}» (синхрон по списку)")
        return True

    # 2. Открыть карточку
    log.info(f"--- ШАГ 2/12: открыть карточку ---")
    if not open_doc_card(driver, row):
        return False
    log.debug(f"  Шаг 2 OK ({time.monotonic()-t_start:.1f}s)")

    # 2.5. Защита от дублирования: если в карточке уже есть резолюция на
    # того же исполнителя — не создаём вторую. Возможные причины расхождения
    # xlsx vs АСУД: mark_status упал (PermissionError), кто-то выдал резолюцию
    # вручную в АСУД мимо нашего daemon'а, рестарт без чтения статуса.
    if has_existing_resolution_to(driver, doc.get('executor')):
        surname = doc['executor'].split()[0] if doc.get('executor') else '?'
        log.info(f"Row {doc['row_idx']}: документ уже имеет резолюцию на "
                 f"{surname} — пропускаю + синхронизирую xlsx")
        # Помечаем в xlsx чтобы daemon в след. раз не пытался открыть снова
        xlsx_path = doc.get('_xlsx_path') or settings.get("_xlsx_path")
        status_col = settings.get("_status_column")
        if xlsx_path and status_col and doc.get('asud_id'):
            if mark_status(xlsx_path, doc['asud_id'], status_col):
                log.debug(f"  → xlsx: {doc['asud_id']} помечен «{status_col}» (синхрон)")
        # Закрываем карточку, возвращаемся в список
        close_card_after_resolution(driver)
        return True  # засчитываем как «обработано» (резолюция и так есть)

    # 3. Создать резолюцию
    log.info(f"--- ШАГ 3/12: кнопка 'Создать резолюцию' ---")
    try:
        btn = driver.find_element(By.ID, "header-action-btn-add_resolution")
        click(driver, btn, "Создать резолюцию")
        # Ждём появления поля "Содержание" — признак что модалка открылась
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "input[placeholder='Общие формулировки']")))
        except Exception:
            log.debug("Поле 'Содержание' не появилось за 10s")
        log.debug(f"  Шаг 3 OK ({time.monotonic()-t_start:.1f}s)")
    except Exception as e:
        log.error(f"Кнопка 'Создать резолюцию' не найдена: {e}")
        return False

    # 4. Содержание
    log.info(f"--- ШАГ 4/12: выбор шаблона 'Содержание' ---")
    select_content_template(driver,
        settings.get("resolution_content", cfg.DEFAULTS["resolution_content"]))
    log.debug(f"  Шаг 4 OK ({time.monotonic()-t_start:.1f}s)")

    # 5. Тоггл "Требуется отчёт"
    if settings.get("require_report", True):
        log.info(f"--- ШАГ 5/12: тоггл 'Требуется отчёт' ---")
        toggle_switch(driver, "Требуется отчёт", "true")
        log.debug(f"  Шаг 5 OK ({time.monotonic()-t_start:.1f}s)")

    # 6. Тоггл "Контрольная резолюция"
    if settings.get("control_resolution", True):
        log.info(f"--- ШАГ 6/12: тоггл 'Контрольная резолюция' ---")
        toggle_switch(driver, "Контрольная резолюция", "true")
        time.sleep(0.5)
        # 7. Дата контрольного этапа: приоритетно — «Планируемая дата» из xlsx
        # (срок ГИС ЖКХ = min(получ+17раб, ГИС)). Если пуста — пересчёт из
        # «Дата получения» (D). Просрочка не затирается на today+N (C).
        log.info(f"--- ШАГ 7/12: дата контрольного этапа ---")
        set_stage_date(driver,
            settings.get("workdays", cfg.DEFAULTS["workdays"]),
            planned_date=doc.get('planned_date'),
            receipt_date=doc.get('receipt_date'))
        log.debug(f"  Шаги 6-7 OK ({time.monotonic()-t_start:.1f}s)")

    # 8. Исполнитель
    log.info(f"--- ШАГ 8/12: исполнитель = {doc['executor']} ---")
    if not fill_executor(driver, doc['executor']):
        log.error(f"Row {doc['row_idx']}: не получилось ввести исполнителя")
        # Закроем диалог чтобы продолжить — но НЕ сохраним
        close_open_modals(driver)
        return False
    log.debug(f"  Шаг 8 OK ({time.monotonic()-t_start:.1f}s)")

    # 9. Клик "Добавить"
    log.info(f"--- ШАГ 9/12: кнопка 'Добавить' ---")
    add_btn = _wait_button_enabled(driver, "add_btn", timeout=10)
    if not add_btn:
        log.error("Кнопка 'Добавить' не активировалась")
        close_open_modals(driver)
        return False
    click(driver, add_btn, "Добавить")
    # Ждём активации save_and_send_btn — её ждёт следом submit_resolution
    log.debug(f"  Шаг 9 OK ({time.monotonic()-t_start:.1f}s)")

    # 10. Клик "Сохранить и отправить"
    log.info(f"--- ШАГ 10/12: 'Сохранить и отправить' + confirm ---")
    if not submit_resolution(driver):
        log.error("Не удалось сохранить и отправить")
        return False
    log.debug(f"  Шаг 10 OK ({time.monotonic()-t_start:.1f}s)")

    # 11. «Завершить» — переводит задачу в «Завершённые» (а не просто
    # закрывает карточку как раньше). Иначе документ остаётся «в работе»
    # у Халецкой и продолжает висеть в списке резолюций.
    # Fallback: если кнопка не появилась (документ уже завершён или
    # карточка не в том состоянии) — обычное закрытие через крестик.
    log.info(f"--- ШАГ 11/12: «Завершить» + закрытие карточки ---")
    if click_complete_button(driver):
        close_card_after_complete(driver)
    else:
        close_card_after_resolution(driver)
    log.debug(f"  Шаг 11 OK ({time.monotonic()-t_start:.1f}s)")

    # 12. Очистить фильтр для следующей итерации
    log.info(f"--- ШАГ 12/12: очистка фильтра ---")
    clear_filter(driver)
    log.debug(f"  Шаг 12 OK")

    # Пометка статуса в xlsx (если знаем путь). Колонка определяется по
    # текущему сценарию: ZHKH-преcет → «Отписано Халецкой», иначе → «Отписано в округ».
    # Приоритет: doc['_xlsx_path'] (per-doc — multi-folder режим) → settings (legacy).
    xlsx_path = doc.get('_xlsx_path') or settings.get("_xlsx_path")
    status_col = settings.get("_status_column")
    if xlsx_path and status_col and doc.get('asud_id'):
        if mark_status(xlsx_path, doc['asud_id'], status_col):
            log.debug(f"  → xlsx: {doc['asud_id']} помечен «{status_col}»")

    log.info(f"ДОКУМЕНТ {index}/{total} ОБРАБОТАН за {time.monotonic()-t_start:.1f}s")
    return True


# ============================================================
# MAIN
# ============================================================

# Глобальный флаг unattended-режима — выставляется из main() при наличии
# любого из batch-флагов (--headless / --preset / --xlsx / --yes / --watch).
# Когда True — все паузы «Enter...» становятся no-op чтобы daemon не висел.
_UNATTENDED = False


def _block_if_interactive(prompt):
    """input() только в интерактивном режиме. В headless/batch — no-op."""
    if _UNATTENDED:
        return ""
    try:
        return input(prompt)
    except EOFError:
        return ""


# ============================================================
# Daemon: Ctrl+C handling
# ============================================================

_stop_flag = False


def _on_sigint(signum, frame):
    global _stop_flag
    if _stop_flag:
        log.error("Повторный Ctrl+C — выхожу немедленно")
        sys.exit(1)
    _stop_flag = True
    log.info("Получен Ctrl+C — остановлюсь после текущего документа")


def _interruptible_sleep(seconds):
    """Sleep с проверкой _stop_flag раз в 0.5с — для быстрого выхода по Ctrl+C."""
    end = time.monotonic() + seconds
    while time.monotonic() < end:
        if _stop_flag:
            return
        time.sleep(0.5)


# ============================================================
# Daemon: session heartbeat + re-login
# ============================================================

HEARTBEAT_INTERVAL_SEC = 300  # каждые 5 минут проверяем не разлогинило ли


def _session_alive(driver, expected_user):
    """Проверяет что в шапке всё ещё видна нужная учётка. Если шапки нет,
    скорее всего нас разлогинило (АСУД редиректит на login-страницу)."""
    try:
        return _account_active(driver, expected_user, timeout=3)
    except Exception:
        return False


def _relogin(driver, url, switch_to, sidebar):
    """Полный re-login: открыть АСУД заново → переключить учётку → сайдбар.
    Используется когда _session_alive отдал False (idle/timeout разлогинил).
    """
    log.warning(f"Heartbeat: сессия с '{switch_to}' потерялась — переподключаюсь")
    try:
        driver.get(url)
        _wait_profile_loaded(driver)
    except Exception as e:
        log.error(f"Re-login: driver.get упал: {e}")
        return False
    if not switch_account(driver, switch_to):
        log.error(f"Re-login: switch_account на '{switch_to}' не сработал")
        return False
    if not click_sidebar_section(driver, sidebar):
        log.error(f"Re-login: sidebar '{sidebar}' не открылся")
        return False
    log.info("Heartbeat: сессия восстановлена")
    return True


# ============================================================
# Daemon: driver crash detection + recovery
# ============================================================
#
# Если Edge упадёт (память / KAV убьёт процесс / MSI обновится в фоне) —
# driver останется со старым session_id, каждый WebDriver-вызов будет
# фейлить с InvalidSessionIdException. Без recovery daemon крутится
# в холостую и никто не замечает.
#
# Триггеры на пересоздание:
#   - InvalidSessionIdException        — сессия мертва
#   - NoSuchWindowException            — окно закрыто
#   - WebDriverException c "chrome not reachable"/"target window already closed"/
#     "session deleted because of page crash" в сообщении

_DRIVER_CRASH_MARKERS = (
    "not reachable",
    "session deleted",
    "target window already closed",
    "no such window",
    "disconnected",
    "browser has closed",
    "chrome failed to start",
    "session not created",
)


def _is_driver_crash(exc):
    """True если исключение похоже на смерть драйвера/браузера."""
    if isinstance(exc, (InvalidSessionIdException, NoSuchWindowException)):
        return True
    if isinstance(exc, WebDriverException):
        msg = (str(exc) or '').lower()
        return any(m in msg for m in _DRIVER_CRASH_MARKERS)
    return False


def _driver_alive(driver):
    """Лёгкая проверка — current_url. Если драйвер мёртв — кидает
    WebDriverException, ловим и отдаём False."""
    try:
        _ = driver.current_url
        return True
    except Exception:
        return False


def _recreate_driver(old_driver, base_dir, url, switch_to, sidebar):
    """Убивает старый driver (best-effort), запускает новый, делает login.

    Возвращает новый driver или None если не смогли.
    """
    log.warning("Driver crash recovery: убиваю старый driver, поднимаю новый")
    try:
        old_driver.quit()
    except Exception:
        pass

    # Edge zombie cleanup: после краша msedge.exe/msedgedriver.exe
    # могут остаться висеть. Лучше прибить — иначе следующая сессия
    # подцепится к мёртвому профилю.
    if sys.platform.startswith('win'):
        for name in ("msedgedriver.exe", "msedge.exe"):
            try:
                import subprocess
                subprocess.run(
                    ["taskkill", "/F", "/IM", name],
                    capture_output=True, timeout=5,
                )
            except Exception:
                pass
        time.sleep(1)  # дать ОС закрыть handles

    try:
        new_driver = _start_browser(base_dir)
    except Exception as e:
        log.error(f"Recovery: _start_browser упал: {e}")
        return None

    try:
        _go_to_asud(new_driver, url)
    except Exception as e:
        log.error(f"Recovery: _go_to_asud упал: {e}")
        try:
            new_driver.quit()
        except Exception:
            pass
        return None

    if not switch_account(new_driver, switch_to):
        log.error("Recovery: switch_account не сработал")
        try:
            new_driver.quit()
        except Exception:
            pass
        return None

    if not click_sidebar_section(new_driver, sidebar):
        log.error("Recovery: sidebar не открылся")
        try:
            new_driver.quit()
        except Exception:
            pass
        return None

    log.info("Driver crash recovery: новый driver запущен и залогинен")
    return new_driver


def _pick_preset(presets):
    """Меню пресетов сценариев. Возвращает выбранный preset dict или None."""
    print("\nВыбери сценарий:")
    for i, p in enumerate(presets, 1):
        name = p.get("name", "?")
        acc = p.get("target_account", "?")
        exe = p.get("force_executor", "") or "(из округа)"
        print(f"  {i}. {name}")
        print(f"      учётка: {acc}  |  исполнитель: {exe}")
    print(f"[Enter] = 1")
    choice = input(f"Номер (1-{len(presets)}) или Enter: ").strip()
    if not choice:
        return presets[0]
    try:
        return presets[int(choice) - 1]
    except (ValueError, IndexError):
        log.warning(f"Неверный выбор '{choice}'")
        return None


def _apply_preset_to_settings(preset):
    """Перекрывает поля settings полями пресета (in-place)."""
    for key in ("target_account", "sidebar_section", "force_executor",
                "resolution_content", "stage_date_mode", "stage_date_days",
                "require_report", "control_resolution"):
        if key in preset:
            settings[key] = preset[key]


def _list_watched_xlsx(watch_list, verbose=False):
    """Расхлопывает watch-list (list of {dir, xlsx_pattern}) в плоский список
    абсолютных путей к существующим xlsx-файлам.

    verbose=True (на старте) — пишет в лог почему пропустил каждую запись.
    Без этого юзер видит «watch: ...» в шапке и потом тихое отсутствие файлов,
    непонятно — папка не существует, или файл с другим именем, или шара
    не примонтирована.

    Пропускает временные файлы Excel ($файл.xlsx, ~$файл.xlsx) — Excel
    создаёт их когда юзер открывает реестр.
    """
    import glob
    out = []
    for entry in watch_list or []:
        d = (entry.get("dir") or "").strip()
        pat = (entry.get("xlsx_pattern") or "*.xlsx").strip()
        if not d:
            if verbose:
                log.warning(f"watch-entry без поля 'dir': {entry!r}")
            continue
        if not os.path.isdir(d):
            if verbose:
                hint = "сетевая шара не примонтирована?" if d[:2] in (r"\\", "//") \
                       else "проверь правильность пути и экранирование \\\\ в JSON"
                log.error(f"watch.dir не существует: {d!r} ({hint})")
            else:
                log.debug(f"_list_watched_xlsx: '{d}' — папка не найдена, пропуск")
            continue
        found = 0
        for p in glob.glob(os.path.join(d, pat)):
            name = os.path.basename(p)
            if name.startswith('~$') or name.startswith('$'):
                continue
            out.append(p)
            found += 1
        if verbose and found == 0:
            log.warning(f"watch.dir {d!r}: папка есть, но '{pat}' не нашёл "
                        f"ни одного файла")
    return out


def _choose_xlsx(base_dir):
    files = [f for f in os.listdir(base_dir) if f.lower().endswith('.xlsx')]
    if not files:
        log.error(f"Нет .xlsx в {base_dir}")
        _block_if_interactive("Enter...")
        sys.exit(1)
    # Сортируем — файлы с '_резолюции' в имени идут первыми (наш формат)
    files.sort(key=lambda f: (0 if 'резолюции' in f.lower() else 1, f))
    if len(files) == 1:
        log.info(f"Файл: {files[0]}")
        return os.path.join(base_dir, files[0])
    print(f"\nНайдено {len(files)} xlsx-файлов:")
    for i, f in enumerate(files, 1):
        marker = ' ← рекомендую' if 'резолюции' in f.lower() and i == 1 else ''
        print(f"  {i}. {f}{marker}")
    # В unattended-режиме автоматически берём первый (рекомендуемый) файл
    if _UNATTENDED:
        log.info(f"Unattended: автовыбор '{files[0]}'")
        return os.path.join(base_dir, files[0])
    choice = input("Выбери номер [1]: ").strip() or "1"
    try:
        return os.path.join(base_dir, files[int(choice) - 1])
    except (ValueError, IndexError):
        log.error("Неверный выбор")
        sys.exit(1)


def _start_browser(base_dir):
    driver_path = os.path.join(base_dir, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error(f"msedgedriver.exe не найден в {base_dir}")
        _block_if_interactive("Enter...")
        sys.exit(1)

    service = EdgeService(executable_path=driver_path)
    options = EdgeOptions()
    if os.environ.get('ASUD_HEADLESS') == '1':
        options.add_argument("--headless=new")
        options.add_argument("--window-size=1920,1080")
        log.info("Edge запущен в HEADLESS режиме")
    else:
        options.add_argument("--start-maximized")
    options.add_argument("--auth-server-whitelist=*.interrao.ru")
    options.add_argument("--auth-negotiate-delegate-whitelist=*.interrao.ru")
    options.add_argument("--log-level=3")
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    return webdriver.Edge(service=service, options=options)


def _wait_profile_loaded(driver, max_wait=120):
    """Ждёт готовности АСУД для resolutions: только readyState + кнопка
    'Создать документ' (как индикатор что главная отрисована).
    Таблицу с задачами НЕ ждём — у Есиной inbox может быть пустым."""
    log.info("Жду готовности АСУД...")
    try:
        WebDriverWait(driver, max_wait).until(
            lambda d: d.execute_script("return document.readyState === 'complete'"))
    except Exception:
        log.warning("readyState не complete")

    try:
        WebDriverWait(driver, max_wait).until(
            EC.element_to_be_clickable((By.ID, "mainscreen-create-button")))
        log.info("АСУД готов")
    except Exception:
        log.warning("Кнопка 'Создать документ' не появилась — продолжаю")


def _go_to_asud(driver, url):
    log.info(f"Открываю {url}")
    driver.get(url)
    _wait_profile_loaded(driver)


def _quit_browser(driver):
    try:
        driver.quit()
        log.info("Браузер закрыт")
    except Exception as e:
        log.warning(f"Ошибка закрытия: {e}")


def main():
    global settings, _UNATTENDED

    parser = argparse.ArgumentParser(
        description="АСУД ИК — выдача резолюций. Без флагов — интерактивный режим.")
    parser.add_argument('--headless', action='store_true',
                        help="Edge без GUI (фоновый режим)")
    parser.add_argument('--preset', type=int, metavar='N',
                        help="Номер пресета (1-N) из config.json, без меню")
    parser.add_argument('--xlsx', metavar='PATH',
                        help="Путь к xlsx-реестру (без меню выбора)")
    parser.add_argument('--yes', action='store_true',
                        help="Не спрашивать «Начать?» — сразу запускать")
    parser.add_argument('--watch', action='store_true',
                        help="Непрерывный мониторинг xlsx (daemon-режим). "
                             "После одного прохода не выходит, а ждёт новых "
                             "строк. Ctrl+C — корректная остановка.")
    parser.add_argument('--poll-interval', type=int, default=None, metavar='SEC',
                        help="Интервал поллинга xlsx в --watch режиме. "
                             "По умолчанию — из config.json (poll_interval_sec), "
                             "либо 300с (5 мин). Флаг переопределяет JSON.")
    parser.add_argument('--sync-only', action='store_true',
                        help="Режим сверки: только проверить состояние АСУД "
                             "vs xlsx для каждой строки (открыть АСУД, для "
                             "необходимых строк прочитать колонку «Направлено», "
                             "если фамилия совпадает с force_executor → отметить "
                             "в xlsx статус). НЕ создаёт резолюций. Полезно "
                             "после ручных правок в АСУД или при подозрении "
                             "на расхождение данных.")
    args = parser.parse_args()

    # Любой batch-флаг → unattended-режим: не зависаем на input("Enter...")
    _UNATTENDED = bool(args.headless or args.preset or args.xlsx or args.yes
                       or args.watch or args.sync_only)

    settings = cfg.load()
    cfg.keep_system_awake(True)

    # Резолвим poll-interval: --poll-interval > config.json::poll_interval_sec > DEFAULTS
    if args.poll_interval is None:
        args.poll_interval = int(settings.get("poll_interval_sec",
                                              cfg.DEFAULTS.get("poll_interval_sec", 300)))

    log.info("=" * 50)
    log.info("АСУД ИК — выдача резолюций")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()
    # В --watch — ротация лога ежесуточно (иначе один файл на дни/недели
    # вырастает до гигабайтов). В one-shot — обычный per-run файл.
    _attach_file_logger(base_dir, rotate_daily=args.watch)

    if args.headless:
        os.environ['ASUD_HEADLESS'] = '1'

    # Меню пресетов (если есть в конфиге)
    presets = settings.get("presets") or []
    preset = None
    if presets:
        if args.preset is not None:
            if not (1 <= args.preset <= len(presets)):
                log.error(f"--preset {args.preset} вне диапазона 1..{len(presets)}")
                sys.exit(1)
            preset = presets[args.preset - 1]
            log.info(f"Пресет (из --preset): {preset.get('name', '?')}")
        else:
            preset = _pick_preset(presets)
            if preset is None:
                log.error("Пресет не выбран — выход")
                sys.exit(1)
            log.info(f"Пресет: {preset.get('name', '?')}")
        _apply_preset_to_settings(preset)
        log.info(f"  учётка:       {settings.get('target_account')}")
        log.info(f"  сайдбар:      {settings.get('sidebar_section')}")
        log.info(f"  исполнитель:  {settings.get('force_executor') or '(из округа)'}")
        log.info(f"  содержание:   {settings.get('resolution_content')}")
        log.info(f"  контр. срок:  {settings.get('stage_date_mode')} "
                 f"+{settings.get('stage_date_days') or settings.get('workdays')}")

    # === Определяем список xlsx-реестров ====================================
    # Приоритет:
    #   1) --xlsx PATH (override всего, один файл)
    #   2) preset["watch"] (multi-folder, без --xlsx)
    #   3) интерактивный _choose_xlsx(base_dir) (fallback: один файл рядом с exe)
    watch_list = preset.get("watch") if preset else None
    xlsx_paths = []
    if args.xlsx:
        if not os.path.isfile(args.xlsx):
            log.error(f"--xlsx {args.xlsx}: файл не найден")
            sys.exit(1)
        xlsx_paths = [args.xlsx]
        log.info(f"Реестр (из --xlsx): {args.xlsx}")
    elif watch_list:
        # verbose=True — диагностика на старте (почему именно не нашлось)
        xlsx_paths = _list_watched_xlsx(watch_list, verbose=True)
        if not xlsx_paths:
            if args.watch:
                # В daemon-режиме не выходим — папки могут появиться позже
                # (сетевая шара, отложенное создание реестров).
                log.warning("Из preset.watch не нашлось ни одного xlsx — "
                            "daemon стартует, будет проверять каждые "
                            f"{args.poll_interval}с")
                xlsx_paths = []  # пустой, но daemon будет переопрашивать
            else:
                log.error("Из preset.watch не нашлось ни одного xlsx — проверь пути")
                _block_if_interactive("Enter...")
                sys.exit(1)
        else:
            log.info(f"Multi-folder режим: {len(xlsx_paths)} реестров из preset.watch")
            for p in xlsx_paths:
                log.info(f"  • {p}")
    else:
        xlsx_paths = [_choose_xlsx(base_dir)]
        log.info(f"Реестр: {xlsx_paths[0]}")

    # === Загружаем все реестры в единый docs (с тегом _xlsx_path) ===========
    force_exe = settings.get("force_executor", "").strip()
    status_col = COL_HALETSKAYA if 'халецк' in force_exe.lower() else COL_OKRUG
    settings["_status_column"] = status_col
    # Для legacy кода (process_one fallback): первый путь, если single-file
    settings["_xlsx_path"] = xlsx_paths[0] if len(xlsx_paths) == 1 else None
    log.info(f"Статус в xlsx будет писаться в колонку: «{status_col}»")

    def _load_and_filter(xpath):
        """Грузит один xlsx, применяет force_executor + filter-already-done.
        Каждой строке проставляет _xlsx_path для роутинга mark_status."""
        d_list = load_excel(xpath) or []
        if force_exe:
            for d in d_list:
                d['executor'] = force_exe
        already = get_done_asud_ids(xpath, status_col)
        before = len(d_list)
        d_list = [d for d in d_list if d.get('asud_id') not in already]
        skipped = before - len(d_list)
        if skipped:
            log.info(f"  {os.path.basename(xpath)}: пропуск {skipped} уже отписанных")
        for d in d_list:
            d['_xlsx_path'] = xpath
        return d_list

    docs = []
    for xp in xlsx_paths:
        docs.extend(_load_and_filter(xp))

    if not docs and not args.watch:
        log.info("Все реестры пусты или уже отработаны — нечего делать")
        _block_if_interactive("Enter...")
        return
    if force_exe:
        log.info(f"Исполнитель подменён для всех строк: {force_exe}")

    # Превью
    print(f"\nПервые 5:")
    for i, d in enumerate(docs[:5], 1):
        flag = '✓' if d['executor'] else '!'
        print(f"  {i}. [{flag}] {d.get('appeal_no') or '???':>10} | "
              f"{(d.get('executor') or 'ИСПОЛНИТЕЛЬ?')[:30]:30} | "
              f"{d.get('subject', '')[:40]}")
    no_executor = sum(1 for d in docs if not d['executor'])
    print(f"\nВсего: {len(docs)} (без исполнителя: {no_executor} — будут пропущены)")

    if args.yes or _UNATTENDED:
        log.info("Unattended: пропускаю подтверждение «Начать?»")
    else:
        if input("Начать? (да/нет): ").strip().lower() not in ("да", "д", "y", "yes", ""):
            print("Отменено.")
            sys.exit(0)

    # SIGINT-handler — на случай --watch чтобы Ctrl+C мягко останавливал.
    # В one-shot режиме тоже не повредит (стандартный KeyboardInterrupt дальше
    # пройдёт через try/finally так же чисто).
    signal.signal(signal.SIGINT, _on_sigint)
    try:
        signal.signal(signal.SIGTERM, _on_sigint)
    except (AttributeError, ValueError):
        pass

    driver = _start_browser(base_dir)
    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        _go_to_asud(driver, url)

        target_account = settings.get("target_account", cfg.DEFAULTS["target_account"])
        sidebar = settings.get("sidebar_section", cfg.DEFAULTS["sidebar_section"])

        # Переключение учётки
        if not switch_account(driver, target_account):
            log.error("Не удалось переключиться на учётку. Прерываю.")
            _block_if_interactive("Enter...")
            return

        # Сайдбар → "На резолюцию" (внутри уже ждёт появления грида)
        click_sidebar_section(driver, sidebar)

        done, err, skip = 0, 0, 0
        last_heartbeat = time.monotonic()

        def _list_accessible():
            """Проверка: видим ли фильтр-input на колонке Номер. Если да —
            список доступен, никакая модалка/карточка не блокирует."""
            try:
                inp = driver.find_element(By.ID, NUMBER_FILTER_INPUT_ID)
                return inp.is_displayed()
            except Exception:
                return False

        def _reset_to_list(force_full=False):
            """Сброс к списку документов. Двухуровневая стратегия:

            1) Lightweight: закрыть открытые модалки/карточки через close_open_modals
               + header-close-btn. Если после этого фильтр-input доступен — выходим
               (быстро, без перезагрузки АСУД).
            2) Full reload: driver.get(url) + switch_account + сайдбар. Только
               если lightweight не помог ИЛИ force_full=True.

            Это решает кейс: process_one успешно отправил резолюцию, но
            close_card_after_resolution не закрыла карточку — раньше следующий
            документ падал, теперь lightweight закрывает её за <1s.
            """
            if not force_full:
                # Lightweight
                try:
                    close_open_modals(driver)
                except Exception:
                    pass
                try:
                    btn = driver.find_element(By.ID, "header-close-btn")
                    if btn.is_displayed():
                        click(driver, btn, "header-close-btn (reset)")
                except Exception:
                    pass
                if _list_accessible():
                    log.debug("[reset] lightweight: список доступен")
                    return
                log.info("[reset] lightweight не помог — полная перезагрузка")
            # Full reload
            try:
                driver.get(url)
                _wait_profile_loaded(driver)
                switch_account(driver,
                    settings.get("target_account", cfg.DEFAULTS["target_account"]))
                click_sidebar_section(driver,
                    settings.get("sidebar_section", cfg.DEFAULTS["sidebar_section"]))
            except Exception as e:
                log.warning(f"Reset to list упал: {e}")

        def _ensure_driver():
            """Проверяет что driver жив. Если нет — пересоздаёт +relogin.
            Возвращает True если driver готов, False если recovery провалился."""
            nonlocal driver, last_heartbeat
            if _driver_alive(driver):
                return True
            new = _recreate_driver(driver, base_dir, url, target_account, sidebar)
            if new is None:
                return False
            driver = new
            last_heartbeat = time.monotonic()
            return True

        def _process_batch(batch_docs):
            """Прогоняет список docs через process_one. Возвращает True если
            прошёл нормально, False если driver упал в середине (тогда caller
            должен сам ensure_driver и попробовать снова на след. итерации)."""
            nonlocal done, skip, err, driver
            for i, doc in enumerate(batch_docs, 1):
                if _stop_flag:
                    log.info("Stop-flag установлен — прерываю batch")
                    return True
                try:
                    ok = process_one(driver, doc, i, len(batch_docs))
                    if ok:
                        done += 1
                        if not _list_accessible():
                            log.info("[main] список не доступен после success — lightweight reset")
                            _reset_to_list()
                    else:
                        skip += 1
                        _reset_to_list()
                except Exception as e:
                    if _is_driver_crash(e):
                        log.error(f"ДРАЙВЕР УПАЛ на документе {i}: {e}")
                        # В daemon — caller на след. итерации сделает recovery
                        # и пересоберёт todo. В one-shot — выходим из batch,
                        # finally вверху закроет всё.
                        return False
                    log.error(f"ОШИБКА документ {i}: {e}")
                    err += 1
                    try:
                        _reset_to_list(force_full=True)
                    except Exception as e2:
                        if _is_driver_crash(e2):
                            log.error(f"Drivers упал во время reset: {e2}")
                            return False
                        log.warning(f"Reset тоже не сработал: {e2}")
            return True

        # === SYNC-only: только сверить АСУД с xlsx, не создавать резолюции ====
        if args.sync_only:
            log.info("=" * 60)
            log.info(f"SYNC-режим: проверка {len(docs)} строк vs АСУД "
                     f"(резолюции НЕ создаются)")
            log.info("=" * 60)
            sync_found = 0
            sync_clean = 0
            sync_missing = 0
            for i, doc in enumerate(docs, 1):
                if _stop_flag:
                    log.info("Stop-flag — прерываю sync")
                    break
                if not _ensure_driver():
                    log.error("Driver упал в sync, recovery не удался — выход")
                    break
                asud_id = doc.get('asud_id')
                executor = doc.get('executor')
                if not asud_id or not executor:
                    continue
                log.info(f"[sync {i}/{len(docs)}] {asud_id} → ожидаю «{executor.split()[0]}»")
                try:
                    row = find_doc_row(driver, doc, timeout=10)
                except Exception as e:
                    if _is_driver_crash(e):
                        log.error(f"  driver crash: {e}")
                        continue
                    log.warning(f"  find_doc_row: {e}")
                    continue
                if not row:
                    log.info(f"  {asud_id}: не найден в списке "
                             f"(возможно ушёл в «Завершённые» — считаем отписанным)")
                    sync_missing += 1
                    # Если документа нет в очереди — значит уже завершён,
                    # помечаем в xlsx чтобы daemon не пытался открыть.
                    xpath = doc.get('_xlsx_path') or settings.get("_xlsx_path")
                    if xpath and mark_status(xpath, asud_id, status_col):
                        log.debug(f"  → xlsx mark «{status_col}» (документ не в очереди)")
                    continue
                # ВСЕГДА открываем карточку — это надёжнее чем смотреть в список
                # (где «Направлено» может быть в соседней task-строке, или вообще
                # отсутствовать в текущем view). В карточке точный источник истины:
                # панель «Этапы исполнения» + контр. срок.
                if not open_doc_card(driver, row):
                    log.warning(f"  ⚠ не открылась карточка — пропускаю как clean")
                    sync_clean += 1
                    continue
                if card_has_any_resolution(driver):
                    log.info(f"  ✓ В карточке есть выданная резолюция — синхрон xlsx")
                    xpath = doc.get('_xlsx_path') or settings.get("_xlsx_path")
                    if xpath and mark_status(xpath, asud_id, status_col):
                        sync_found += 1
                        log.debug(f"  → xlsx mark «{status_col}»")
                else:
                    sync_clean += 1
                    log.info(f"  ⚪ в карточке нет резолюции — оставляю для daemon")
                close_card_after_resolution(driver)
            log.info("=" * 60)
            log.info(f"SYNC ГОТОВО: проверено {len(docs)}, "
                     f"синхронизировано {sync_found}, "
                     f"чистых (ждут обработки) {sync_clean}, "
                     f"не в очереди (помечены done) {sync_missing}")
            log.info("=" * 60)
            return  # sync — это всегда one-shot, выходим без daemon loop

        # Первый проход — по уже загруженному списку docs
        batch_ok = _process_batch(docs)
        if not batch_ok and args.watch:
            # Daemon: попробуем восстановить driver — следующая итерация
            # перечитает реестр и докрутит непрошедшие документы.
            if not _ensure_driver():
                log.error("Recovery после первого batch не удался")

        # === DAEMON-режим: продолжаем опрашивать xlsx до Ctrl+C ===========
        if args.watch and not _stop_flag:
            log.info("=" * 60)
            if len(xlsx_paths) > 1:
                log.info(f"WATCH-режим: опрашиваю {len(xlsx_paths)} реестров "
                         f"каждые {args.poll_interval}с. Ctrl+C для остановки.")
            elif xlsx_paths:
                log.info(f"WATCH-режим: опрашиваю {xlsx_paths[0]} каждые "
                         f"{args.poll_interval}с. Ctrl+C для остановки.")
            else:
                log.info(f"WATCH-режим: реестров пока нет, жду появления "
                         f"(проверка каждые {args.poll_interval}с)")
            log.info("=" * 60)

            # mtime-кэш per-file. После первого прохода знаем mtime каждого
            # файла — на след. итерации перечитываем только те где mtime изменился.
            mtime_cache = {}
            for xp in xlsx_paths:
                try:
                    mtime_cache[xp] = os.path.getmtime(xp)
                except OSError:
                    mtime_cache[xp] = 0
            iters = 0
            # Сколько итераций назад пере-резолвили watch_list. Делаем это
            # раз в ~5 минут чтобы подхватить вновь появившиеся файлы
            # (сетевая шара примонтировалась, юзер создал реестр в новой папке).
            last_watch_rescan = time.monotonic()
            WATCH_RESCAN_INTERVAL = 300
            # Round-robin: каждый тик берёт один реестр в порядке xlsx_paths.
            # Поведение: poll_interval ОЭК → poll_interval ТЭС → ...
            watch_round_robin = bool(settings.get("watch_round_robin", False))
            rr_idx = 0
            if watch_round_robin and len(xlsx_paths) > 1:
                log.info(f"Watch mode: ROUND-ROBIN ({len(xlsx_paths)} реестров "
                         f"по очереди, по одному за тик)")
            elif len(xlsx_paths) > 1:
                log.info(f"Watch mode: ALL ({len(xlsx_paths)} реестров на каждом тике)")

            while not _stop_flag:
                _interruptible_sleep(args.poll_interval)
                if _stop_flag:
                    break
                iters += 1

                # Driver crash check — ставим в начало итерации перед
                # любыми операциями с driver. Если упал — пересоздаём.
                if not _ensure_driver():
                    log.error(f"[итер. {iters}] driver crash recovery не удался, "
                              f"повтор через 60с")
                    _interruptible_sleep(60)
                    continue

                # Heartbeat — проверяем что в шапке всё ещё нужная учётка
                if time.monotonic() - last_heartbeat > HEARTBEAT_INTERVAL_SEC:
                    try:
                        alive = _session_alive(driver, target_account)
                    except Exception as e:
                        if _is_driver_crash(e):
                            log.error(f"Heartbeat: driver упал ({e}) — recovery")
                            _ensure_driver()
                            continue
                        alive = False
                    if not alive:
                        if not _relogin(driver, url, target_account, sidebar):
                            log.error("Heartbeat: re-login не удался, повтор через 60с")
                            _interruptible_sleep(60)
                            continue
                    last_heartbeat = time.monotonic()

                # Раз в WATCH_RESCAN_INTERVAL пере-резолвим watch_list:
                # может за это время примонтировалась шара / появился реестр
                # в новой папке. Только в multi-folder режиме (watch_list задан).
                if watch_list and time.monotonic() - last_watch_rescan > WATCH_RESCAN_INTERVAL:
                    new_paths = _list_watched_xlsx(watch_list)
                    added = [p for p in new_paths if p not in mtime_cache]
                    if added:
                        log.info(f"[итер. {iters}] watch-rescan: появились новые "
                                 f"реестры ({len(added)})")
                        for p in added:
                            log.info(f"  + {p}")
                            mtime_cache[p] = 0
                        xlsx_paths = new_paths
                    last_watch_rescan = time.monotonic()

                # Mode selection:
                # 1) round-robin (watch_round_robin=true в config.json) —
                #    на каждом тике берём ОДИН реестр, следующая итерация —
                #    следующий. Удобно когда хочется чёткое расписание:
                #    «5 мин ОЭК → 5 мин ТЭС → 5 мин ГИСЖКХ → ...»
                # 2) all (по умолчанию) — на каждом тике проверяем ВСЕ реестры,
                #    обрабатываем todo со всех у которых mtime изменился.
                if watch_round_robin and xlsx_paths:
                    cur_xpath = xlsx_paths[rr_idx % len(xlsx_paths)]
                    rr_idx += 1
                    log.info(f"[итер. {iters}] round-robin: проверяю "
                             f"{os.path.basename(cur_xpath)}")
                    paths_this_tick = [cur_xpath]
                else:
                    paths_this_tick = xlsx_paths

                changed_files = 0
                busy_files = 0
                todo_all = []
                for xpath in paths_this_tick:
                    # Приоритет регистрации: если на реестре висит свежий
                    # .lock (P1-регистратор сейчас пишет новые строки) —
                    # пропускаем его на этом тике, возвращаемся через
                    # poll_interval. mark_status в process_one всё равно
                    # ждёт обычным lock-acquire, но тут мы экономим время
                    # на чтении большого xlsx в момент когда он точно
                    # будет освобождён через секунду.
                    if is_xlsx_busy(xpath):
                        log.info(f"[итер. {iters}] {os.path.basename(xpath)} занят "
                                 f"(регистратор пишет?) — пропускаю, попробую на след. тике")
                        busy_files += 1
                        continue
                    try:
                        cur_mtime = os.path.getmtime(xpath)
                    except OSError:
                        continue  # файл сейчас недоступен — пропустим этот раунд
                    if cur_mtime == mtime_cache.get(xpath, 0):
                        continue
                    mtime_cache[xpath] = cur_mtime
                    changed_files += 1
                    todo_all.extend(_load_and_filter(xpath))

                if not todo_all:
                    log.debug(f"[итер. {iters}] изменилось файлов: {changed_files}, todo: 0")
                    continue

                if watch_round_robin:
                    log.info(f"[итер. {iters}] todo: {len(todo_all)} документов "
                             f"из {os.path.basename(paths_this_tick[0])}")
                else:
                    log.info(f"[итер. {iters}] todo: {len(todo_all)} документов "
                             f"из {changed_files} изменённых реестров")
                batch_ok = _process_batch(todo_all)
                if not batch_ok:
                    log.warning(f"[итер. {iters}] batch прерван driver-crash'ем "
                                f"— непрошедшие документы попробуем на след. итерации")
                    # Сбрасываем mtime-кэш чтобы на след. итерации перечитать ВСЁ
                    mtime_cache = {p: 0 for p in xlsx_paths}
        # === END daemon =====================================================

        elapsed_seconds = time.monotonic() - start_time
        elapsed = timedelta(seconds=int(elapsed_seconds))
        avg = (timedelta(seconds=int(elapsed_seconds / done))
               if done else None)
        summary = [
            "",
            "=" * 60,
            "ГОТОВО!",
            f"  Обработано:  {done} / {len(docs)}",
            f"  Пропущено:   {skip}",
            f"  Ошибок:      {err}",
            f"  Затрачено:   {elapsed}" + (f"  (в среднем {avg}/док)" if avg else ""),
            "=" * 60,
        ]
        for line in summary:
            log.info(line)
            print(line)
        _block_if_interactive("\nEnter для закрытия...")
    except Exception as e:
        log.error(f"Ошибка: {e}")
        _block_if_interactive("Enter...")
    finally:
        _quit_browser(driver)
        cfg.keep_system_awake(False)


if __name__ == "__main__":
    main()
