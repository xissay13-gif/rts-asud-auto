"""
xlsx_status.py — Хелперы для отметки статусов в реестре резолюций.

Реестры (Registered/*_резолюции.xlsx) содержат колонки:
  "Отписано Халецкой"  — заполняется когда Басманов выдал резолюцию Халецкой
                          (второй проход ГИСЖКХ)
  "Отписано в округ"   — заполняется когда Халецкая выдала резолюцию окружному
                          начальнику (основной сценарий resolutions.py)

Формат значения: дата DD.MM.YYYY (когда отписали).

Используется и для пометки (после успешной обработки), и для skip-already-done
(чтобы recovery после crash/сна не дублировал отписки).
"""

import os
import time
import logging
from datetime import date

import openpyxl

from xlsx_lock import xlsx_lock

log = logging.getLogger("asud")

# Retry-backoff на PermissionError (юзер открыл xlsx в Excel — OS-level
# lock держит Excel, наш кооперативный xlsx_lock этого не покрывает).
# Попытки: сразу, через 5с, через 10с, через 30с. Итого до ~45с ожидания.
_PERMISSION_RETRY_DELAYS_SEC = (0, 5, 10, 30)

COL_HALETSKAYA = "Отписано Халецкой"
COL_OKRUG     = "Отписано в округ"
STATUS_COLUMNS = (COL_HALETSKAYA, COL_OKRUG)


def _find_asud_col(headers_lower):
    """Возвращает 0-based индекс колонки с asud_id (Номер/ОПТС). None если нет."""
    keys = ('номер', 'опт', 'орт', 'асуд', 'asud', 'регистрацион')
    for i, h in enumerate(headers_lower):
        for k in keys:
            if k in h:
                return i
    return None


def get_done_asud_ids(xlsx_path, column_name):
    """Возвращает set asud_id у которых колонка column_name непустая.
    Используется для skip-already-done в начале прохода.
    """
    done = set()
    if not xlsx_path or not os.path.isfile(xlsx_path):
        return done
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return done
        headers = [str(c or '').strip() for c in rows[0]]
        headers_lower = [h.lower() for h in headers]
        asud_col = _find_asud_col(headers_lower)
        try:
            status_col = headers.index(column_name)
        except ValueError:
            wb.close()
            return done
        if asud_col is None:
            wb.close()
            return done
        for r in rows[1:]:
            if len(r) <= max(asud_col, status_col):
                continue
            aid = str(r[asud_col] or '').strip()
            val = str(r[status_col] or '').strip() if r[status_col] is not None else ''
            if aid and val:
                done.add(aid)
        wb.close()
    except Exception as e:
        log.debug(f"get_done_asud_ids({xlsx_path!r}, {column_name!r}): {e}")
    return done


def mark_status(xlsx_path, asud_id, column_name, value=None):
    """Помечает статус для строки с заданным asud_id в xlsx.

    Поиск строки — по колонке Номер/ОПТС.
    Если column_name отсутствует в шапке — добавляет.
    value=None → подставляется сегодняшняя дата DD.MM.YYYY.

    Возвращает True если запись сохранилась.
    """
    if not xlsx_path or not os.path.isfile(xlsx_path):
        return False
    if value is None:
        value = date.today().strftime("%d.%m.%Y")

    # Lock на время read-modify-save: нужен потому что параллельно с
    # P3 (clean-resolutions) тот же реестр трогают P1 (регистратор
    # дописывает строки) и P2 (zhkh_daemon обновляет «Отписано Халецкой»).
    # Без lock — последний пишущий перезаписывает чужие правки.
    #
    # PermissionError-retry: если юзер открыл xlsx в Excel, Windows держит
    # exclusive file-lock на уровне OS — xlsx_lock этого не покрывает.
    # Ретраи с backoff. Если так и не получилось — возвращаем False,
    # caller увидит unmarked row на следующей итерации и попробует снова
    # (но без дублей в АСУД т.к. документ уже не будет в списке).
    last_err = None
    for attempt, delay in enumerate(_PERMISSION_RETRY_DELAYS_SEC):
        if delay:
            log.warning(f"mark_status: PermissionError на {xlsx_path} "
                        f"(Excel открыт?) — ретрай через {delay}с "
                        f"({attempt}/{len(_PERMISSION_RETRY_DELAYS_SEC)-1})")
            time.sleep(delay)
        try:
            with xlsx_lock(xlsx_path, timeout=60):
                wb = openpyxl.load_workbook(xlsx_path)
                ws = wb.active
                headers = [str(c.value or '').strip() for c in next(ws.iter_rows(max_row=1))]
                headers_lower = [h.lower() for h in headers]
                asud_col = _find_asud_col(headers_lower)
                if asud_col is None:
                    log.warning(f"mark_status: в {xlsx_path} нет колонки Номер/ОПТС")
                    wb.close()
                    return False
                try:
                    status_col = headers.index(column_name)
                except ValueError:
                    status_col = len(headers)
                    ws.cell(row=1, column=status_col + 1).value = column_name
                    ws.cell(row=1, column=status_col + 1).font = openpyxl.styles.Font(bold=True)
                    ws.column_dimensions[
                        openpyxl.utils.get_column_letter(status_col + 1)].width = 18
                target_row = None
                for r in ws.iter_rows(min_row=2):
                    if asud_col >= len(r):
                        continue
                    if str(r[asud_col].value or '').strip() == asud_id:
                        target_row = r[0].row
                        break
                if target_row is None:
                    log.debug(f"mark_status: asud_id {asud_id!r} не найден в {xlsx_path}")
                    wb.close()
                    return False
                ws.cell(row=target_row, column=status_col + 1).value = value
                wb.save(xlsx_path)
                wb.close()
                return True
        except PermissionError as e:
            last_err = e
            continue
        except Exception as e:
            log.warning(f"mark_status({asud_id}, {column_name}): {e}")
            return False

    log.error(f"mark_status({asud_id}, {column_name}): PermissionError "
              f"после {len(_PERMISSION_RETRY_DELAYS_SEC)} попыток ({last_err}). "
              f"Строка останется без отметки — обработается на следующей итерации.")
    return False
