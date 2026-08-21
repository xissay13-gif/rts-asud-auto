"""
flows/email.py — Email-direct: создание Входящих документов прямо из .msg файлов
в указанной папке (без xlsx-реестра).

Два режима:
  • main()        — однопроходный: пробежать все .msg в корне папки и выйти
  • daemon_main() — непрерывный мониторинг: опрос папки раз в N сек,
                     обработка новых .msg, автоперемещение по результату
                     (Завершено / Ошибки / Черновики), Ctrl+C для остановки

Запуск через app.py с --mode=email (одноразовый) или --mode=email --watch (daemon).
"""

import os
import json
import re
import signal
import sys
import time
import logging
from datetime import date, datetime, timedelta

import openpyxl
import extract_msg
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService

from shared import config as cfg
from shared.ui import wait_asud_loaded, set_driver_timeout
from shared.correspondent import extract_fio_from_text
from shared.okrug_parser import okrug_from_textbody
from shared.deadline import (
    parse_ddmmyyyy as _parse_ddmmyyyy,
    add_working_days as _add_working_days,
    compute_deadline as _compute_zhkh_deadline,
)
from shared.attachments import move_to_done, move_to_errors, move_to_drafts
from shared.colors import green, yellow, red, status_colored
from shared.classifier import classify_doc_type
from shared.zhkh_parser import parse_zhkh_body
from shared.zhkh_routing import match_excluded_topic
from shared.feedback_parser import parse_feedback_body
from shared.xlsx_lock import xlsx_lock
from shared.xlsx_format import format_registry_before_save

# Переиспользуем создание/регистрацию/output из mix
from flows import mix as mix_flow

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger("asud")

# extract_msg/olefile очень шумные на INFO — для каждого отсутствующего
# необязательного стрима пишут «Stream ... was requested but could not be
# found». Это НЕ ошибки (у писем просто нет части опциональных свойств —
# SMTP отправителя, message-id и т.п.). Глушим до WARNING.
for _noisy in ("extract_msg", "olefile"):
    logging.getLogger(_noisy).setLevel(logging.WARNING)

start_time = time.monotonic()

settings = {}


class ExclusionMarkerError(RuntimeError):
    """Excluded MSG could not be marked terminal; ASUD processing must stop."""


# ================= ENCODING-FIX =================

def _fix_mojibake(s):
    """Эвристика: если строка похожа на CP1251 байты, прочитанные как Latin-1
    (типичная ошибка extract_msg на некоторых .msg) — реkodируем обратно.

    Triggers (все вместе):
      - 3+ символов из диапазона U+00C0..U+00FF (Latin-1 «cyrillic-looking»)
      - реальной кириллицы (U+0400..U+04FF) меньше чем «бракованных»
      - после re-encode latin-1 → cp1251 ≥80% бракованных символов
        стали кириллицей (раньше сравнивали с len(candidate)//3 — ломалось
        на строках с большим количеством ASCII: «55-2026-33646 24-я Северная,
        д. 168, к. 1 до 14.06» — 13 кириллических из 50 символов, порог 16,
        не проходило, кракозябры оставались)
    Иначе возвращаем строку как есть.
    """
    if not s or not isinstance(s, str):
        return s
    bad = sum(1 for c in s if 0x00C0 <= ord(c) <= 0x00FF)
    if bad < 3:
        return s
    cyrillic = sum(1 for c in s if 0x0400 <= ord(c) <= 0x04FF)
    if cyrillic > bad:
        return s
    try:
        candidate = s.encode('latin-1').decode('cp1251')
        new_cyr = sum(1 for c in candidate if 0x0400 <= ord(c) <= 0x04FF)
        # 0xC0-0xFF в CP1251 — вся кириллица. True mojibake → каждый bad
        # становится кириллической буквой. Допуск 80% — на случай если
        # пара байт оказалась знаком/латиницей.
        if new_cyr >= bad * 0.8:
            return candidate
    except Exception:
        pass
    return s


# ================= FOLDER CHECK WITH RETRY =================

def _normalize_unc(folder):
    """Чинит частый косяк в JSON-конфиге: путь к UNC-шаре указан с одним
    бэкслешем (\\\\server в JSON → \\server в строке) вместо двух
    (\\\\\\\\server в JSON → \\\\server в строке).

    Симптом: log.error выводит repr типа '\\\\interrao.ru\\\\oms10\\\\...' —
    repr показывает каждый реальный \\ как \\\\, и видно что в начале только
    один настоящий \\. На os.path.isdir такая «полу-UNC» отдаёт False.

    Эвристика: путь начинается с \\ + хост.домен.* (т.е. с точкой в первой
    после \\ части) и os.path.isdir(folder) == False — добавляем ведущий \\.
    """
    if not folder or os.path.isdir(folder):
        return folder
    if folder.startswith('\\') and not folder.startswith('\\\\'):
        # Берём первую часть после ведущего \\ — если в ней есть точка,
        # похоже на FQDN (interrao.ru, server.local, etc.)
        head = folder[1:].split('\\', 1)[0]
        if '.' in head:
            fixed = '\\' + folder  # дописываем второй ведущий \\
            if os.path.isdir(fixed):
                log.warning(f"Путь {folder!r} был не-UNC (один \\ в начале) — "
                            f"автоисправлено на {fixed!r}. ИСПРАВЬ settings.json: "
                            f"в начале UNC-пути должно быть \\\\\\\\ (четыре \\\\), "
                            f"они декодируются JSON'ом в два.")
                return fixed
    return folder


def _wait_for_folder(folder, retries=4, delays=(0, 2, 5, 10)):
    """os.path.isdir с retry-backoff. Нужно для UNC-путей (\\\\server\\share)
    у которых первый доступ может уйти в timeout (DNS resolve + Kerberos
    handshake + проверка прав занимают несколько секунд после холодного
    старта процесса).

    Дополнительно: если путь похож на UNC но с одним \\ в начале вместо
    двух (распространённый косяк в settings.json) — авто-исправляет.

    Возвращает (True, normalized_path) если папка доступна, (False, folder)
    после всех попыток.
    """
    import time as _time
    # Сразу пробуем нормализовать (быстрая проверка без retry)
    folder = _normalize_unc(folder)
    for attempt, delay in enumerate(delays[:retries]):
        if delay:
            log.info(f"Папка {folder!r}: попытка {attempt+1}/{retries} "
                     f"через {delay}с (UNC холодный старт?)")
            _time.sleep(delay)
        if os.path.isdir(folder):
            if attempt > 0:
                log.info(f"Папка появилась после {attempt} ретраев")
            return True, folder
    return False, folder


# ================= EMAIL → DOC_DATA =================

# Имя .msg-файла начинается с даты-времени: '2026-05-06 10-58-43.msg'
_FILENAME_DATE_RE = re.compile(r'^(\d{4}-\d{2}-\d{2})')


def _msg_date_prefix(file_path):
    """Извлекает 'YYYY-MM-DD' из имени .msg. None если не распарсилось."""
    m = _FILENAME_DATE_RE.match(os.path.basename(file_path))
    return m.group(1) if m else None


def _msg_link(msg, file_path):
    """Генерирует Link для doc_data.

    Приоритет:
      1) Дата письма из .msg (msg.date)
      2) Имя файла без расширения
    Формат — как в mix-flow реестре: 'DD.MM.YYYY HH-MM-SS'.
    """
    try:
        if msg is not None and msg.date:
            return msg.date.strftime("%d.%m.%Y %H-%M-%S")
    except Exception:
        pass
    return os.path.splitext(os.path.basename(file_path))[0]


def _raw_msg_text(msg_path, prop_id):
    """Читает строковый стрим .msg НАПРЯМУЮ через olefile, минуя codepage-
    логику extract_msg.

    Зачем: у некоторых писем в свойствах врёт кодировка (напр. заявлена
    gb2312/китайская, а тело — кириллица cp1251). extract_msg доверяет тегу
    и падает с UnicodeDecodeError на первом кириллическом байте, теряя всё
    письмо. Мы читаем стрим сами: сначала UTF-16 (суффикс 001F — он всегда
    корректный, если есть), потом cp1251 (001E).

    prop_id — '0037' (тема) / '1000' (тело). Возвращает строку или ''.
    """
    try:
        import olefile
    except Exception:
        return ''
    try:
        ole = olefile.OleFileIO(msg_path)
    except Exception:
        return ''
    try:
        for suffix, codec in (('001F', 'utf-16-le'), ('001E', 'cp1251')):
            stream = f'__substg1.0_{prop_id}{suffix}'
            try:
                if ole.exists(stream):
                    raw = ole.openstream(stream).read()
                    return raw.decode(codec, errors='replace')
            except Exception:
                continue
        return ''
    finally:
        try:
            ole.close()
        except Exception:
            pass


def _safe_field(msg, attr, msg_path, prop_id):
    """Достаёт msg.subject / msg.body, при ошибке декодирования (кривой
    codepage) — fallback на прямое чтение OLE-стрима через _raw_msg_text."""
    try:
        val = getattr(msg, attr) if msg is not None else None
        if val:
            return val
    except Exception as e:
        log.info(f"{os.path.basename(msg_path)}: extract_msg не смог "
                 f"декодировать «{attr}» ({e}) — читаю стрим напрямую")
        return _raw_msg_text(msg_path, prop_id)
    # val пустой/None — на всякий случай тоже пробуем сырьё
    return _raw_msg_text(msg_path, prop_id) if msg is None else (val or "")


# ================= PER-DATE XLSX REGISTRY =================

# Колонки реестра (под per-date).
# Колонки 1-5 — данные документа, 6 — статус/дата обработки, 7-8 —
# заполняются позже («Отписано Халецкой» — после второго прохода ГИСЖКХ,
# «Отписано в округ» — после прогона clean-resolutions).
_REGISTRY_HEADERS = ["Номер", "Link", "Округ", "Subject", "Body",
                     "Дата получения", "Планируемая дата", "Статус",
                     "Отписано Халецкой", "Отписано в округ",
                     "GIS Номер обращения", "External GUID"]
_REGISTRY_WIDTHS = {1: 18, 2: 22, 3: 8, 4: 50, 5: 60, 6: 16, 7: 16, 8: 28,
                    9: 18, 10: 18, 11: 24, 12: 38}

_API_REGISTRY_HEADERS = ("GIS Номер обращения", "External GUID")


def _xlsx_path(base_dir, suffix=None, target_folder=None):
    """Путь к накопительному реестру (один файл, все записи append'ятся).

    Раньше был per-date (<YYYY-MM-DD>_<suffix>_резолюции.xlsx) — на каждый день
    свой файл. Теперь один файл-накопитель: <suffix>_резолюции.xlsx. Так удобнее:
      • daemon мониторит один файл, не glob
      • вся история в одном месте, фильтр в Excel по дате через колонку «Статус»
      • меньше мусора в папке

    Куда кладём:
      • Если target_folder задан И существует — прямо туда (рядом с .msg).
      • Иначе — fallback на <base_dir>/Registered/ (легаси).

    Без suffix: <where>/резолюции.xlsx
    С suffix:   <where>/<suffix>_резолюции.xlsx
    """
    name = f"{suffix}_резолюции.xlsx" if suffix else "резолюции.xlsx"

    if target_folder and os.path.isdir(target_folder):
        return os.path.join(target_folder, name)

    registered_dir = os.path.join(base_dir, "Registered")
    os.makedirs(registered_dir, exist_ok=True)
    return os.path.join(registered_dir, name)


# Backward-compat alias — старое имя ещё может встречаться в каком-то коде.
def _dated_xlsx_path(base_dir, date_prefix, suffix=None, target_folder=None):
    """DEPRECATED: используй _xlsx_path. date_prefix теперь игнорируется."""
    return _xlsx_path(base_dir, suffix=suffix, target_folder=target_folder)


def _ensure_dated_xlsx(path):
    """Создаёт per-date xlsx с шапкой если не существует."""
    if os.path.isfile(path):
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Резолюции"
        ws.append(_REGISTRY_HEADERS)
        for c in range(1, len(_REGISTRY_HEADERS) + 1):
            ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
        for col, w in _REGISTRY_WIDTHS.items():
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = w
        ws.freeze_panes = "A2"
        format_registry_before_save(ws, path, changed_row=1)
        wb.save(path)
    except Exception as e:
        log.warning(f"Не удалось создать {path}: {e}")


def _append_dated_row(path, doc, asud_id, status="OK"):
    """Дописывает строку в накопительный xlsx.

    Адаптивен к схеме файла: читает шапку, выбирает значения по именам
    колонок. Если в xlsx нет какой-то колонки (например, старый файл
    без «Планируемая дата») — её просто пропускаем. Если в xlsx новая
    колонка которой нет в этой функции — пишем пусто. Обратная
    совместимость гарантирована.

    status — internal код результата обработки. В колонку «Статус» пишем:
      OK        → «Зарегистрирован DD.MM.YYYY HH:MM»
      DRAFT     → «Черновик DD.MM.YYYY HH:MM»
      DUPLICATE → «Дубликат DD.MM.YYYY HH:MM»
      REGISTERED_ONLY → «Зарегистрирован без резолюции ...»
      SUBMISSION_UNKNOWN → «Результат регистрации не определён ...»
    """
    ts = datetime.now().strftime("%d.%m.%Y %H:%M")
    status_text = {
        "OK":        f"Зарегистрирован {ts}",
        "DRAFT":     f"Черновик {ts}",
        "DUPLICATE": f"Дубликат {ts}",
        "REGISTERED_ONLY": f"Зарегистрирован без резолюции {ts}",
        "SUBMISSION_UNKNOWN": f"Результат регистрации не определён {ts}",
    }.get(status, f"{status} {ts}")

    # Знакомые колонки → значения. ЛЮБОЕ имя колонки в шапке xlsx, которое
    # есть в этом словаре, будет заполнено. Остальные останутся пустыми.
    # «Дата получения» — чистая DD.MM.YYYY из тела ГИС ЖКХ (без времени).
    # Нужна P3 (clean-resolutions) чтобы пересчитать срок если «Планируемая
    # дата» в реестре пустая. Для не-ГИС писем — пусто.
    _recv = _parse_ddmmyyyy(doc.get("дата_обращения"))
    recv_str = _recv.strftime("%d.%m.%Y") if _recv else ""

    values = {
        "Номер":             asud_id or "",
        "Link":              doc.get("link") or "",
        "Округ":             doc.get("округ_прогноз") or "",
        "Subject":           doc.get("тема") or "",
        "Body":              doc.get("содержание") or "",  # уже _clean_body
        "Дата получения":    recv_str,
        "Планируемая дата":  doc.get("планируемая_дата") or "",
        "Статус":            status_text,
    }
    try:
        with xlsx_lock(path, timeout=60):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            # Шапка существующего файла
            headers = [str(c.value or '').strip() for c in next(ws.iter_rows(max_row=1))]
            row = [values.get(h, "") for h in headers]
            ws.append(row)
            format_registry_before_save(ws, path, changed_row=ws.max_row)
            wb.save(path)
    except Exception as e:
        log.warning(f"Не удалось записать строку в {path}: {e}")


def _api_registry_values(doc, asud_id, external_guid):
    """Builds the strict registry row used only by the GIS API backend."""
    gis_number = str(doc.get("номер_обращения") or "").strip()
    external_guid = str(external_guid or "").strip()
    asud_id = str(asud_id or "").strip()
    if not gis_number:
        raise ValueError("у ГИС ЖКХ документа отсутствует номер обращения")
    if not external_guid:
        raise ValueError("API не вернул External GUID")
    if not asud_id:
        raise ValueError("API не вернул регистрационный номер АСУД")

    received = _parse_ddmmyyyy(doc.get("дата_обращения"))
    received_text = received.strftime("%d.%m.%Y") if received else ""
    timestamp = datetime.now().strftime("%d.%m.%Y %H:%M")
    return {
        "Номер": asud_id,
        "Link": doc.get("link") or "",
        "Округ": doc.get("округ_прогноз") or "",
        "Subject": doc.get("тема") or "",
        "Body": doc.get("содержание") or "",
        "Дата получения": received_text,
        "Планируемая дата": doc.get("планируемая_дата") or "",
        "Статус": f"Зарегистрирован {timestamp}",
        "GIS Номер обращения": gis_number,
        "External GUID": external_guid,
    }


def _upsert_api_dated_row(path, doc, asud_id, external_guid):
    """Idempotently writes a successfully registered GIS API document.

    The normal Selenium path deliberately continues to use append.  The API
    path has a durable external GUID and therefore can safely resume local
    finalisation after a crash without creating a second daemon work item.
    Existing resolution columns are never overwritten.
    """
    values = _api_registry_values(doc, asud_id, external_guid)
    temp_path = path + ".asud_api.tmp.xlsx"

    with xlsx_lock(path, timeout=60):
        if os.path.isfile(path):
            workbook = openpyxl.load_workbook(path)
            worksheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Резолюции"
            worksheet.append(_REGISTRY_HEADERS)
            for column, width in _REGISTRY_WIDTHS.items():
                worksheet.cell(row=1, column=column).font = (
                    openpyxl.styles.Font(bold=True)
                )
                worksheet.column_dimensions[
                    openpyxl.utils.get_column_letter(column)
                ].width = width
            worksheet.freeze_panes = "A2"

        headers = [str(cell.value or "").strip()
                   for cell in worksheet[1]]
        nonempty_headers = [header for header in headers if header]
        if len(nonempty_headers) != len(set(nonempty_headers)):
            workbook.close()
            raise ValueError("в реестре есть повторяющиеся названия колонок")

        # Older registries are extended in place.  Their column order and all
        # historical data stay untouched.
        for header in _API_REGISTRY_HEADERS:
            if header not in headers:
                column = len(headers) + 1
                cell = worksheet.cell(row=1, column=column, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                headers.append(header)
                width = _REGISTRY_WIDTHS.get(
                    _REGISTRY_HEADERS.index(header) + 1, 24)
                worksheet.column_dimensions[
                    openpyxl.utils.get_column_letter(column)].width = width

        header_columns = {header: index + 1
                          for index, header in enumerate(headers) if header}
        # Old registries did not always contain every descriptive column.
        # Preserve that adaptive legacy schema, but require all identity
        # columns needed for a strict idempotent upsert.
        required = {
            "Номер", "Link", "GIS Номер обращения", "External GUID",
        }
        missing = sorted(required - set(header_columns))
        if missing:
            workbook.close()
            raise ValueError(
                "в реестре отсутствуют обязательные колонки: "
                + ", ".join(missing)
            )

        guid_column = header_columns["External GUID"]
        gis_column = header_columns["GIS Номер обращения"]
        link_column = header_columns.get("Link")
        number_column = header_columns.get("Номер")
        wanted_guid = values["External GUID"]
        wanted_gis = values["GIS Номер обращения"]
        wanted_link = str(values.get("Link") or "").strip()
        wanted_number = str(values["Номер"] or "").strip()

        primary_matches = []
        legacy_matches = []
        for row_index in range(2, worksheet.max_row + 1):
            row_guid = str(worksheet.cell(
                row=row_index, column=guid_column).value or "").strip()
            row_gis = str(worksheet.cell(
                row=row_index, column=gis_column).value or "").strip()
            guid_match = bool(row_guid and row_guid == wanted_guid)
            gis_match = bool(row_gis and row_gis == wanted_gis)
            if guid_match or gis_match:
                if row_guid and row_guid != wanted_guid:
                    workbook.close()
                    raise ValueError(
                        f"конфликт External GUID в строке {row_index}")
                if row_gis and row_gis != wanted_gis:
                    workbook.close()
                    raise ValueError(
                        f"конфликт номера ГИС ЖКХ в строке {row_index}")
                primary_matches.append(row_index)
                continue

            # Backward compatibility for a row written before API columns
            # existed.  Link alone is insufficient; the ASUD number must also
            # agree so a timestamp collision cannot merge different records.
            if (not row_guid and not row_gis and wanted_link
                    and link_column and number_column):
                row_link = str(worksheet.cell(
                    row=row_index, column=link_column).value or "").strip()
                row_number = str(worksheet.cell(
                    row=row_index, column=number_column).value or "").strip()
                if row_link == wanted_link and row_number == wanted_number:
                    legacy_matches.append(row_index)

        # A primary API row and a matching legacy row are still two records
        # for the same document.  Do not silently prefer one: the operator
        # must reconcile the duplicate before local finalisation can proceed.
        matches = primary_matches + legacy_matches
        if len(matches) > 1:
            workbook.close()
            raise ValueError(
                "в реестре найдено несколько строк для одного GIS документа"
            )

        if matches:
            changed_row = matches[0]
            existing_number = str(worksheet.cell(
                row=changed_row, column=number_column).value or ""
            ).strip()
            if existing_number and existing_number != wanted_number:
                workbook.close()
                raise ValueError(
                    "конфликт регистрационного номера АСУД в строке "
                    f"{changed_row}: {existing_number!r} != "
                    f"{wanted_number!r}"
                )
        else:
            changed_row = worksheet.max_row + 1

        for header, value in values.items():
            if header not in header_columns:
                continue
            worksheet.cell(
                row=changed_row,
                column=header_columns[header],
                value=value,
            )

        format_registry_before_save(
            worksheet, path, changed_row=changed_row)
        try:
            workbook.save(temp_path)
            workbook.close()
            os.replace(temp_path, path)
        except Exception:
            workbook.close()
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except OSError:
                pass
            raise

    return changed_row


def _list_root_msgs(folder_path):
    """Возвращает sorted list абсолютных путей к .msg в корне folder_path
    (без рекурсии). Пустой list если папки нет / I/O ошибка."""
    try:
        out = []
        recover_api_state = _api_backend_selected()
        for f in os.listdir(folder_path):
            full = os.path.join(folder_path, f)
            if not (os.path.isfile(full) and f.lower().endswith('.msg')):
                continue
            terminal = os.path.isfile(full + ".asud_terminal.json")
            claim = os.path.isfile(full + ".asud_api.claim")
            state = os.path.isfile(full + ".asud_api_state.json")
            if terminal:
                continue
            if claim or state:
                # Only a complete write-ahead pair is safe to replay through
                # the API recovery path.  It performs no network mutation and
                # only finishes local XLSX/terminal state.  A lone sidecar is
                # kept fail-closed for manual inspection in every backend.
                if recover_api_state and claim and state:
                    out.append(full)
                continue
            out.append(full)
        return sorted(out)
    except OSError as e:
        log.error(f"Не могу прочитать папку {folder_path}: {e}")
        return []


def _write_terminal_marker(msg_path, status, reason, **metadata):
    """Атомарно исключает MSG из повторного сканирования.

    Возвращает путь к маркеру либо ``None``. Сам MSG не изменяет.
    """
    if not msg_path:
        return None
    marker = str(msg_path) + ".asud_terminal.json"
    tmp = marker + ".tmp"
    payload = {
        "version": 1,
        "status": str(status),
        "reason": str(reason),
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    payload.update({key: value for key, value in metadata.items()
                    if value is not None})
    try:
        with open(tmp, "w", encoding="utf-8") as stream:
            json.dump(payload, stream, ensure_ascii=False, indent=2)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(tmp, marker)
        return marker
    except Exception as exc:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        log.error(f"Не удалось записать terminal-маркер для {msg_path}: {exc}")
        return None


def _email_registration_backend():
    backend = str(os.environ.get(
        "ASUD_EMAIL_REGISTRATION_BACKEND", "selenium")
    ).strip().casefold()
    if backend not in {"selenium", "asud_api"}:
        raise ValueError(
            "ASUD_EMAIL_REGISTRATION_BACKEND must be selenium or asud_api"
        )
    return backend


def _api_backend_selected():
    return _email_registration_backend() == "asud_api"


def _api_external_guid(outcome, doc):
    """Returns the public GUID exposed by core without inspecting its state."""
    value = getattr(outcome, "external_guid", None)
    if not value:
        value = doc.get("asud_api_external_guid")
    return str(value or "").strip()


def _mark_api_terminal(msg_path, status, reason, outcome=None, **metadata):
    """Durably stops automatic API/Selenium processing while keeping MSG.

    Keeping the original MSG in the source root is intentional: the external
    GIS downloader uses it for deduplication.  The adjacent terminal marker is
    what hides it from this program's scanner.
    """
    outcome_metadata = {}
    if outcome is not None:
        outcome_metadata = {
            "object_id": getattr(outcome, "object_id", None),
            "registration_number": getattr(
                outcome, "registration_number", None),
            "state_path": str(getattr(outcome, "state_path", None) or "")
                          or None,
        }
    outcome_metadata.update(metadata)
    marker = _write_terminal_marker(
        msg_path,
        status,
        reason,
        **outcome_metadata,
    )
    if not marker:
        raise ExclusionMarkerError(
            "не удалось записать terminal-маркер API; "
            "автоматическая обработка остановлена"
        )
    return marker


def _process_doc_via_gis_api(doc, base_dir, folder, output_suffix):
    """Runs the isolated GIS API backend and maps its durable outcome.

    This function never calls Selenium and never falls back to it.  All
    uncertain post-submission statuses get a durable terminal marker before
    control returns to the generic one-shot summary.
    """
    msg_path = doc.get("файл")
    doc["_processed_via_asud_api"] = True

    from flows.gis_api import process_gis_document, should_use_gis_api

    if not should_use_gis_api(doc, settings):
        # The backend was explicitly selected, so passing a non-GIS message to
        # Selenium would be a dangerous implicit fallback.
        reason = (
            "API backend выбран, но письмо не распознано структурным "
            "парсером ГИС ЖКХ; Selenium fallback запрещён"
        )
        log.error(reason)
        return "API_FAILED", None, None

    try:
        outcome = process_gis_document(
            doc,
            settings,
            logger=log,
        )
    except Exception as exc:
        # Core is expected to classify known pre/post mutation failures.  An
        # uncaught exception has unknown delivery semantics, therefore it is
        # terminal rather than retryable.
        reason = f"Непредвиденная ошибка GIS API: {exc}"
        log.exception(reason)
        _mark_api_terminal(
            msg_path,
            "MANUAL_REVIEW",
            reason,
        )
        return "MANUAL_REVIEW", None, None

    raw_status = getattr(outcome, "status", "")
    status = str(getattr(raw_status, "value", raw_status) or "").strip().upper()
    object_id = str(getattr(outcome, "object_id", None) or "").strip()
    registration_number = str(
        getattr(outcome, "registration_number", None) or ""
    ).strip()
    message = str(getattr(outcome, "message", None) or status).strip()

    if status == "OK":
        external_guid = _api_external_guid(outcome, doc)
        written_xlsx = _xlsx_path(
            base_dir, output_suffix, target_folder=folder)
        try:
            _upsert_api_dated_row(
                written_xlsx,
                doc,
                registration_number,
                external_guid,
            )
        except Exception as exc:
            reason = (
                "API завершил регистрацию, но строгая финализация XLSX "
                f"не выполнена: {exc}"
            )
            log.error(reason)
            _mark_api_terminal(
                msg_path,
                "MANUAL_REVIEW",
                reason,
                outcome=outcome,
                external_guid=external_guid or None,
                gis_number=doc.get("номер_обращения"),
            )
            return (
                "MANUAL_REVIEW",
                registration_number or object_id or None,
                None,
            )

        _mark_api_terminal(
            msg_path,
            "OK",
            "GIS API: документ зарегистрирован и отправлен на резолюцию",
            outcome=outcome,
            external_guid=external_guid,
            gis_number=doc.get("номер_обращения"),
            registry_path=written_xlsx,
        )
        log.info(
            f"GIS API: {registration_number} зарегистрирован; "
            "MSG оставлен в корне с terminal-маркером"
        )
        return "OK", registration_number, written_xlsx

    if status in {
            "MANUAL_REVIEW", "SUBMISSION_UNKNOWN", "REGISTERED_ONLY"}:
        _mark_api_terminal(
            msg_path,
            status,
            message,
            outcome=outcome,
            external_guid=_api_external_guid(outcome, doc) or None,
            gis_number=doc.get("номер_обращения"),
        )
        log.error(
            f"GIS API: terminal status {status}; MSG оставлен в корне, "
            "повтор и Selenium fallback запрещены"
        )
        return status, registration_number or object_id or None, None

    if status in {"DRY_RUN", "PROBE"}:
        log.info(
            f"GIS API: {status}; сеть/мутации ограничены режимом, "
            "terminal-маркер не создаётся"
        )
        return status, registration_number or object_id or None, None

    if status == "FAILED":
        # Core guarantees FAILED is pre-mutation and safe to repeat after the
        # operator fixes configuration/connectivity.  Keep the MSG in place and
        # use a distinct status so the legacy loop cannot move it to Ошибки/.
        log.error(
            f"GIS API: pre-mutation failure: {message}; "
            "MSG оставлен в корне, повтор безопасен после исправления"
        )
        return "API_FAILED", None, None

    reason = f"GIS API вернул неизвестный статус {status!r}: {message}"
    log.error(reason)
    _mark_api_terminal(
        msg_path,
        "MANUAL_REVIEW",
        reason,
        outcome=outcome,
        external_guid=_api_external_guid(outcome, doc) or None,
        gis_number=doc.get("номер_обращения"),
    )
    return "MANUAL_REVIEW", registration_number or object_id or None, None


def _exclude_msg_from_asud(msg_path, doc):
    """Оставляет GIS MSG в корне, но навсегда исключает его из ASUD-очереди.

    GIS downloader дедуплицирует обращения по наличию исходного ``.msg`` в
    корне. Поэтому файл нельзя удалять или переносить: иначе он скачается снова.
    """
    topic_code = doc.get("zhkh_topic_code")
    topic_title = doc.get("тема_обращения")
    reason = (
        f"Тема ГИС ЖКХ {topic_code or ''} исключена из регистрации в АСУД"
    ).strip()
    marker = _write_terminal_marker(
        msg_path,
        "EXCLUDED",
        reason,
        topic_code=topic_code,
        topic_title=topic_title,
    )
    if not marker:
        return False
    log.warning(
        f"{os.path.basename(str(msg_path))}: НЕ регистрируется в АСУД; "
        f"MSG оставлен для дедупликации выгрузчика, marker={marker}"
    )
    return True


def _quarantine_terminal_msg(msg_path, folder, status, reason):
    """Не даёт письму с неопределённой регистрацией попасть в очередь снова.

    Маркер пишется атомарно *до* перемещения. Если UNC/антивирус держит MSG и
    ``Ошибки/`` временно недоступна, корневой сканер всё равно пропустит файл.
    При успешном перемещении маркер больше не нужен и удаляется.
    """
    marker = _write_terminal_marker(msg_path, status, reason)
    marker_written = bool(marker)

    moved = bool(move_to_errors(msg_path, folder, reason))
    if moved:
        try:
            if marker and os.path.exists(marker):
                os.remove(marker)
        except Exception as exc:
            # MSG уже вне корня, оставшийся marker безвреден.
            log.debug(f"Terminal-маркер не удалён: {exc}")
        return True
    if marker_written:
        log.critical(
            f"MSG не перемещён, но исключён из очереди marker-файлом: {marker}"
        )
        return True
    return False


def _parse_one_msg(msg_path, process_mode="mix"):
    """Парсит один .msg в doc_data dict. None если не получилось/пустое/брак.

    process_mode:
      'mix'   — классификатор НЕ вызывается, тип всегда из settings
                ('default_type_idx', по умолчанию 8). Корреспондент из ФИО в теле.
      'smart' — вызывается classify_doc_type (для определения тип_индекс).
                cat -1 → пропуск. cat 0 → тип 8 + force_draft.
                Корреспондент всё равно будет затёрт в _process_doc на «Неизвестный...».

    Использует module-global settings.
    """
    unknown = settings.get("unknown_correspondent",
                            cfg.DEFAULTS["unknown_correspondent"])

    msg = None
    try:
        msg = extract_msg.openMsg(msg_path)
    except Exception as e:
        # extract_msg вообще не открыл файл — попробуем сырое чтение стримов
        log.warning(f"{os.path.basename(msg_path)}: extract_msg.openMsg упал "
                    f"({e}) — fallback на прямое чтение OLE")
    # Тему и тело достаём защищённо: при кривом codepage extract_msg кидает
    # UnicodeDecodeError на .subject/.body — тогда читаем стрим напрямую.
    subject = _fix_mojibake(_safe_field(msg, 'subject', msg_path, '0037'))
    body = _fix_mojibake(_safe_field(msg, 'body', msg_path, '1000'))
    link = _msg_link(msg, msg_path)
    if msg is not None:
        try:
            msg.close()
        except Exception:
            pass

    if not body and not subject:
        log.warning(f"Пустое письмо {os.path.basename(msg_path)} — пропускаю")
        return None

    # ГИС ЖКХ — спец-парсер табличного body. Запускаем ДО классификатора —
    # для ZHKH-писем тип всегда 8 (письма граждан), классификатор не трогаем.
    zhkh = parse_zhkh_body(body)
    # Feedback с сайта Омск РТС («Новый вопрос с сайта Омск РТС») — структурный
    # парсер. Старый шаблон содержит ФИО, новый может содержать только адрес.
    # Без ФИО используем адрес как корреспондента: весь адрес идёт в поле
    # «Фамилия», а обязательные Имя/Отчество получают прочерки.
    # Парсим ТОЛЬКО если не ZHKH (zhkh > feedback по приоритету).
    feedback = None if zhkh else parse_feedback_body(body, subject=subject)

    skip_asud_registration = False
    zhkh_topic_code = None
    if zhkh:
        try:
            zhkh_topic_code = match_excluded_topic(
                zhkh.get("тема_обращения"),
                settings.get(
                    "zhkh_excluded_topics",
                    cfg.DEFAULTS["zhkh_excluded_topics"],
                ),
            )
        except ValueError as exc:
            log.error(
                f"{os.path.basename(msg_path)}: некорректная политика "
                f"исключений ГИС ЖКХ — {exc}; письмо не будет передано в АСУД"
            )
            return None
        skip_asud_registration = bool(zhkh_topic_code)
        if skip_asud_registration:
            log.warning(
                f"{os.path.basename(msg_path)}: тема ГИС ЖКХ "
                f"{zhkh_topic_code} исключена из регистрации в АСУД"
            )

    force_draft = False
    if zhkh:
        # ZHKH-письмо: тип 8 железно, классификатор скипаем
        type_idx = 8
        log.info(f"{os.path.basename(msg_path)}: ГИС ЖКХ → {zhkh.get('фио') or '(аноним)'}, "
                 f"№{zhkh.get('номер_обращения')} от {zhkh.get('дата_обращения')}")
    elif feedback:
        # Feedback-письмо: тип 8 (письма граждан). Корреспондентом будет ФИО
        # из старого шаблона либо адрес из нового.
        type_idx = 8
        log.info(f"{os.path.basename(msg_path)}: feedback Омск РТС → "
                 f"{feedback.get('фио') or '(без ФИО)'}, "
                 f"адрес {feedback.get('адрес') or '—'}, "
                 f"ЛС {feedback.get('лицевой_счет') or '—'}")
    elif process_mode == "smart":
        # Не ZHKH: для smart-режима пускаем общий классификатор
        type_idx = classify_doc_type(body)
        if type_idx == -1:
            log.warning(f"{os.path.basename(msg_path)}: помечено классификатором "
                        f"как 'случайно отправил' — пропускаю")
            return None
        if type_idx == 0:
            type_idx = 8
            force_draft = True
            log.info(f"{os.path.basename(msg_path)}: тип не определился → 8, "
                     f"оставляю в черновике для ручной проверки")
    else:
        # mix-режим без ZHKH: дефолтный тип из settings
        type_idx = settings.get("default_type_idx", 8)
    type_name = cfg.DOC_TYPE_MAP.get(
        type_idx, "Письма, заявления и жалобы граждан, акционеров")

    clean_subject = re.sub(r'^(FW:|RE:|Fwd:)\s*', '',
                            str(subject).strip(), flags=re.IGNORECASE)
    # ГИС ЖКХ: если тема приехала без префикса (начинается сразу с номера
    # обращения вроде «55-2026-36415 ул. ...») — допишем «ГИС ЖКХ» вперёд
    # для единообразия в реестре и в карточке АСУД.
    if zhkh and not re.match(r'^\s*ГИС\s*ЖКХ', clean_subject, re.IGNORECASE):
        clean_subject = f"ГИС ЖКХ {clean_subject}"
    body_clean = mix_flow._clean_body(body) if body else clean_subject

    correspondent_kind = "person"
    if zhkh:
        # ФИО может быть None для анонимных ГИС ЖКХ-писем (нет поля «Заявитель»)
        # — тогда корреспондент = «Неизвестный», но номер/тема/адрес из zhkh
        # всё равно используются.
        correspondent = zhkh.get('фио') or unknown
        corr_found = bool(zhkh.get('фио'))
        fio_src = "zhkh"
    elif feedback:
        correspondent = feedback.get('корреспондент') or unknown
        corr_found = bool(feedback.get('корреспондент'))
        correspondent_kind = feedback.get('корреспондент_тип') or "person"
        fio_src = "feedback"
    elif force_draft:
        # Smart + cat-0: принудительно черновик с фикс. корреспондентом
        # (чтобы письмо точно ушло в DRAFT-ветку mix.create_one_document).
        corr_found = False
        correspondent = unknown
        fio_src = ""
    else:
        fio, fio_src = extract_fio_from_text(body_clean)
        correspondent = fio if fio else unknown
        corr_found = bool(fio)

    try:
        okrug = okrug_from_textbody(body_clean, base_dir_fn=cfg.get_base_dir)
    except Exception as e:
        log.warning(f"okrug_parser упал для {os.path.basename(msg_path)}: {e}")
        okrug = None

    # Для ZHKH и feedback-писем краткое содержание в АСУД = тема (subject),
    # а не полный body (там длинный шаблонный текст с подписью).
    # Для feedback дополнительно приклеиваем адрес из тела: «<subject> — <адрес>»
    # (subject обычно «Новый вопрос с сайта Омск РТС» — оставляем как есть,
    # юзеру удобнее видеть и его, и адрес).
    # Срок отработки ГИС ЖКХ: РТС даёт 17 РАБОЧИХ дней с даты получения,
    # но если срок по ГИС ЖКХ раньше — берём его (min из двух).
    zhkh_deadline = None
    if zhkh:
        zhkh_deadline = _compute_zhkh_deadline(
            zhkh.get('дата_обращения'), zhkh.get('планируемая_дата'))

    if feedback:
        addr = (feedback.get('адрес') or '').strip()
        soderzhanie = f"{clean_subject} — {addr}" if addr else clean_subject
    elif zhkh:
        # Типовое краткое содержание ГИС ЖКХ:
        #   «ГИС ЖКХ <номер обращения> <адрес> срок до <дата>»
        # Все три поля — из структурного парсера. Пустые сегменты пропускаем.
        zparts = ["ГИС ЖКХ"]
        znum = (zhkh.get('номер_обращения') or '').strip()
        zaddr = (zhkh.get('адрес') or '').strip()
        if znum:
            zparts.append(znum)
        if zaddr:
            zparts.append(zaddr)
        if zhkh_deadline:
            zparts.append(f"срок до {zhkh_deadline}")
        soderzhanie = " ".join(zparts)
        # Подстраховка: если ни номера, ни адреса не распарсилось —
        # откатываемся на тему письма, чтобы содержание не было голым «ГИС ЖКХ».
        if not znum and not zaddr:
            soderzhanie = clean_subject
    else:
        soderzhanie = body_clean

    return {
        "row_idx": 1,
        "содержание": soderzhanie,
        "корреспондент": correspondent,
        "корреспондент_тип": correspondent_kind,
        "корр_найден": corr_found,
        "корр_источник": fio_src,
        "тема": clean_subject,
        "тип_индекс": type_idx,
        "тип_название": type_name,
        "link": link,
        "файл": msg_path,
        "округ_прогноз": okrug,
        "msg_date_prefix": _msg_date_prefix(msg_path),
        "force_draft": force_draft,
        # ZHKH-специфичные поля (None если письмо не из ГИС ЖКХ).
        # Дальше передадутся в mix.create_one_document для заполнения
        # доп. полей АСУД-карточки.
        "номер_обращения":  zhkh.get('номер_обращения')  if zhkh else None,
        "дата_обращения":   zhkh.get('дата_обращения')   if zhkh else None,
        # Планируемая дата = вычисленный срок отработки (min(получ.+17 раб.дн,
        # срок ГИС)). Этот же срок уходит в контрольную дату резолюции.
        # Fallback на сырую дату ГИС если вычисление не удалось.
        "планируемая_дата": (zhkh_deadline or zhkh.get('планируемая_дата')) if zhkh else None,
        "тема_обращения":   zhkh.get('тема_обращения') if zhkh else None,
        "zhkh_topic_code":  zhkh_topic_code,
        # Явная бизнес-политика: карточку документа не открывать и ничего не
        # писать в реестр резолюций. Caller оставит MSG в корне с terminal-marker.
        "skip_asud_registration": skip_asud_registration,
    }


def load_emails(folder_path, process_mode="mix"):
    """Парсит все .msg в корне folder_path. Возвращает list of doc dicts.
    process_mode пробрасывается в _parse_one_msg (см. его docstring)."""
    msg_files = _list_root_msgs(folder_path)
    log.info(f"Найдено .msg файлов: {len(msg_files)}")

    rows, skipped = [], 0
    for idx, msg_path in enumerate(msg_files, 1):
        doc = _parse_one_msg(msg_path, process_mode)
        if doc is None:
            skipped += 1
            continue
        doc["row_idx"] = idx
        rows.append(doc)

    log.info(f"Загружено: {len(rows)} писем, пропущено: {skipped}")
    return rows


# ================= PROCESSING ONE DOC =================

def _print_doc_line(index, total, status, info=""):
    """Одна строка результата по документу — с цветным статусом в консоль.
    OK — зелёный, DRAFT — жёлтый, FAILED — красный, DUPLICATE — без цвета."""
    label = status_colored(status)
    suffix = f" — {info}" if info else ""
    print(f"  [{index}/{total}] {label}{suffix}")


def _process_doc(driver, doc, base_dir, folder, index, total, in_daemon,
                 process_mode="mix", output_suffix=None):
    """Обрабатывает один doc: create_one_document → ветвление по статусу
    → запись в xlsx + перенос .msg.

    process_mode:
      'mix'   — текущая логика: ФИО найдено → register, нет → DRAFT
      'smart' — всегда черновик: forсируем корр_найден=False + фикс. корреспондент;
                 DRAFT считается УСПЕХОМ → пишем в реестр и переносим в Завершено/

    output_suffix — суффикс в имени per-date xlsx (для разделения реестров
    при параллельных запусках двух пресетов).

    Возвращает финальный статус: 'OK' | 'DUPLICATE' | 'DRAFT' | 'EXCLUDED' |
    'REGISTERED_ONLY' | 'SUBMISSION_UNKNOWN' | 'FAILED' либо API-only статусы
    'DRY_RUN' | 'PROBE' | 'MANUAL_REVIEW' | 'API_FAILED'.
    in_daemon=True (mix-режим): DRAFT → перенос в Черновики/.
    """
    msg_path = doc.get("файл")
    written_xlsx = None  # путь к xlsx куда записали (для второго прохода)

    # Business exclusion is terminal and happens before any ASUD/UI call.
    # The MSG stays in the root so the GIS downloader does not fetch it again;
    # the adjacent marker makes _list_root_msgs ignore it forever.
    if doc.get("skip_asud_registration"):
        if not _exclude_msg_from_asud(msg_path, doc):
            raise ExclusionMarkerError(
                "не удалось записать marker исключения; вход в АСУД запрещён"
            )
        return "EXCLUDED", None, None

    # Explicit API backend is isolated from Selenium.  Structurally invalid
    # messages are rejected inside the API adapter; they never fall through to
    # mix.create_one_document().
    if _api_backend_selected():
        if in_daemon:
            raise RuntimeError(
                "GIS API backend поддерживает только однопроходный режим"
            )
        return _process_doc_via_gis_api(
            doc,
            base_dir,
            folder,
            output_suffix,
        )

    if process_mode == "smart" and doc.get("корр_источник") not in ("zhkh", "feedback"):
        # Smart-пресет: каждый .msg создаётся как черновик с фикс. корреспондентом.
        # Исключение — структурно распознанные письма (ZHKH / feedback): у них
        # корреспондент надёжно получен из отдельного поля ФИО либо адреса.
        doc["корр_найден"] = False
        doc["корреспондент"] = settings.get("unknown_correspondent",
                                             cfg.DEFAULTS["unknown_correspondent"])

    asud_id = mix_flow.create_one_document(driver, doc, index, total)
    status = mix_flow._last_result.get("status", "FAILED")

    if status == "OK":
        written_xlsx = _xlsx_path(base_dir, output_suffix, target_folder=folder)
        _ensure_dated_xlsx(written_xlsx)
        _append_dated_row(written_xlsx, doc, asud_id, status="OK")
        move_to_done(msg_path, folder)
    elif status == "DUPLICATE":
        log.info(f"Документ {index}: уже зарегистрирован — .msg в Завершено/")
        move_to_done(msg_path, folder)
    elif status == "DRAFT":
        if process_mode == "smart":
            # Smart: черновик — это нормальный исход. Пишем в реестр (без АСУД-ID)
            # и переносим .msg в Завершено/.
            written_xlsx = _xlsx_path(base_dir, output_suffix, target_folder=folder)
            _ensure_dated_xlsx(written_xlsx)
            _append_dated_row(written_xlsx, doc, asud_id or "", status="DRAFT")
            move_to_done(msg_path, folder)
            log.info(f"Документ {index}: создан как черновик (smart) — .msg в Завершено/")
        elif in_daemon:
            log.info(f"Документ {index}: ФИО не найдено — .msg в Черновики/")
            move_to_drafts(msg_path, folder)
        # one-shot mix: оставляем в корне как и было
    elif status in {"REGISTERED_ONLY", "SUBMISSION_UNKNOWN"}:
        # После отправленного события регистрацию нельзя безопасно повторять:
        # документ либо уже существует, либо результат ответа АСУД неизвестен.
        written_xlsx = _xlsx_path(base_dir, output_suffix, target_folder=folder)
        _ensure_dated_xlsx(written_xlsx)
        _append_dated_row(
            written_xlsx, doc, asud_id or "", status=status
        )
        log.error(
            f"Документ {index}: {status} — требуется ручная проверка; "
            "автоматический повтор запрещён"
        )
    else:  # FAILED — caller сам решает что делать (retry / move-to-errors)
        pass

    return status, asud_id, written_xlsx


# ================= MAIN =================

def main():
    global settings
    settings = cfg.load()
    cfg.setup_file_logger("email")
    cfg.keep_system_awake(True)

    # Читаем логику обработки сразу — нужно для корректного превью
    process_mode = os.environ.get('ASUD_EMAIL_PROCESS_MODE', 'mix')
    api_backend = _api_backend_selected()

    log.info("=" * 50)
    log.info("АСУД ИК — Email-direct (создание из .msg-писем)")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()

    # Запрос пути к папке с .msg — спрашиваем всегда, даже при пресете.
    # Дефолтом подставляется ASUD_EMAIL_FOLDER (от пресета) или email_folder из конфига.
    default = os.environ.get('ASUD_EMAIL_FOLDER') \
              or settings.get("email_folder", "")
    print(f"\nПапка с .msg-письмами (только из корня папки, подпапки игнорируются).")
    if default:
        print(f"Enter — использовать: {default}")
    user_dir = input("Путь: ").strip().strip('"').strip("'")
    folder = user_dir or default

    if folder:
        ok, folder = _wait_for_folder(folder)
    else:
        ok = False
    if not ok:
        log.error(f"Папка не найдена (после ретраев): {folder!r}")
        input("Enter...")
        sys.exit(1)

    log.info(f"Папка писем: {folder}")

    # Парсим письма (process_mode определяет, использовать ли классификатор)
    docs = load_emails(folder, process_mode)
    if not docs:
        log.error("Нет .msg файлов или все пропущены")
        input("Enter...")
        sys.exit(1)

    if api_backend:
        non_gis = [doc for doc in docs
                   if doc.get("корр_источник") != "zhkh"]
        api_eligible = [doc for doc in docs
                        if (doc.get("корр_источник") == "zhkh"
                            and not doc.get("skip_asud_registration"))]
        if non_gis or len(api_eligible) > 1:
            if non_gis:
                log.critical(
                    "GIS API preflight: в очереди есть не-ГИС письма "
                    f"({len(non_gis)}). Вся партия отклонена; "
                    "Selenium fallback запрещён."
                )
            if len(api_eligible) > 1:
                log.critical(
                    "GIS API preflight: найдено "
                    f"{len(api_eligible)} обращений для API, а live-one "
                    "разрешает максимум одно. Вся партия отклонена до "
                    "первого API-вызова."
                )
            cfg.keep_system_awake(False)
            input("Enter для закрытия...")
            return

    # Превью — зависит от backend/process_mode
    if api_backend:
        structured = sum(
            1 for doc in docs if doc.get("корр_источник") == "zhkh")
        print(f"\nПервые 5 (тестовый backend GIS API, без Edge):")
        for i, doc in enumerate(docs[:5], 1):
            flag = "GIS" if doc.get("корр_источник") == "zhkh" else "НЕ GIS"
            print(f"  {i}. [{flag}] {doc['тема'][:70]}")
        print(f"\nВсего: {len(docs)}  (структурно ГИС ЖКХ: {structured})")
        print("режим: GIS API ONE-SHOT — Selenium fallback запрещён; "
              "MSG остаются в корне")
    elif process_mode == "smart":
        print(f"\nПервые 5 (все будут созданы как ЧЕРНОВИКИ):")
        for i, d in enumerate(docs[:5], 1):
            print(f"  {i}. [тип {d['тип_индекс']}] {d['тема'][:60]}")
        print(f"\nВсего: {len(docs)}  (тип определён классификатором, "
              f"корреспондент будет «Неизвестный...»)")
        print("режим: EMAIL/SMART  —  создание ТОЛЬКО как черновик + .msg "
              "(без регистрации)")
    else:  # mix
        known = sum(1 for d in docs if d["корр_найден"])
        unknown_n = len(docs) - known
        print(f"\nПервые 5:")
        for i, d in enumerate(docs[:5], 1):
            flag = 'OK' if d["корр_найден"] else '!!'
            print(f"  {i}. [тип {d['тип_индекс']}] {flag} "
                  f"{d['корреспондент'][:30]} | {d['тема'][:50]}")
        print(f"\nВсего: {len(docs)}  (ФИО найдено: {known}, без ФИО: {unknown_n})")
        print("режим: EMAIL/MIX  —  создание + регистрация + На резолюцию + .msg")

    if input("Начать? (да/нет): ").strip().lower() not in ("да", "д", "y", "yes", ""):
        sys.exit(0)

    # === Запуск обработки. API one-shot intentionally does not require Edge.
    driver = None
    if not api_backend:
        driver_path = os.path.join(base_dir, "msedgedriver.exe")
        if not os.path.exists(driver_path):
            log.error(f"msedgedriver.exe не найден в {base_dir}")
            input("Enter...")
            sys.exit(1)

        options = cfg.build_edge_options()
        service = EdgeService(executable_path=driver_path)
        driver = webdriver.Edge(service=service, options=options)
        set_driver_timeout(
            driver,
            settings.get(
                "asud_load_timeout_sec",
                cfg.DEFAULTS["asud_load_timeout_sec"],
            ),
        )

    # Настраиваем mix_flow.settings — legacy Selenium path uses this global.
    mix_flow.settings = settings

    output_suffix = os.environ.get('ASUD_OUTPUT_SUFFIX') or None
    log.info(f"Логика обработки: {process_mode}"
             + (", backend: asud_api (one-shot)" if api_backend else "")
             + (f", суффикс реестра: {output_suffix}" if output_suffix else ""))

    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        if driver is not None:
            log.info(f"Открываю {url}")
            driver.get(url)
            wait_asud_loaded(driver)
        else:
            log.info("GIS API backend: Edge/msedgedriver не запускаются")

        # Per-date реестры: Registered/YYYY-MM-DD[_<suffix>]_резолюции.xlsx.
        # Каждый doc пишется в xlsx своей даты (из имени .msg).
        log.info(f"Per-date реестры в: {os.path.join(base_dir, 'Registered')}")

        done_count = dup_count = draft_count = excluded_count = err_count = 0
        dry_run_count = probe_count = manual_review_count = api_failed_count = 0
        # Просто счётчик зарегистрированных ГИСЖКХ — для итогового лога.
        # Сама отписка делается отдельным процессом
        # (АСУД_ЖКХ_резолюции_Халецкой.bat).
        registered_docs = []
        for i, doc in enumerate(docs, 1):
            msg_path = doc.get("файл")
            try:
                status, asud_id, written_xlsx = _process_doc(
                                       driver, doc, base_dir, folder,
                                       i, len(docs), in_daemon=False,
                                       process_mode=process_mode,
                                       output_suffix=output_suffix)
                if status == "OK":
                    done_count += 1
                    if asud_id:
                        registered_docs.append(asud_id)
                elif status == "DUPLICATE":
                    dup_count += 1
                elif status == "DRAFT":
                    draft_count += 1
                elif status == "EXCLUDED":
                    excluded_count += 1
                elif status == "DRY_RUN":
                    dry_run_count += 1
                elif status == "PROBE":
                    probe_count += 1
                elif status == "API_FAILED":
                    api_failed_count += 1
                    err_count += 1
                    log.error(
                        "GIS API не выполнял мутаций; MSG оставлен в корне. "
                        "Исправьте конфигурацию/доступ и повторите безопасно."
                    )
                elif (status in {
                        "MANUAL_REVIEW", "REGISTERED_ONLY",
                        "SUBMISSION_UNKNOWN"}
                      and doc.get("_processed_via_asud_api")):
                    # `_process_doc_via_gis_api` already persisted a terminal
                    # marker.  Never move the MSG or append an uncertain row.
                    manual_review_count += 1
                    err_count += 1
                elif status in {"REGISTERED_ONLY", "SUBMISSION_UNKNOWN"}:
                    reason = (
                        "Документ зарегистрирован, но резолюция не подтверждена"
                        if status == "REGISTERED_ONLY" else
                        "Результат регистрации в АСУД не определён"
                    ) + "; требуется ручная проверка"
                    quarantined = _quarantine_terminal_msg(
                        msg_path, folder, status, reason
                    )
                    err_count += 1
                    if not quarantined:
                        log.critical(
                            "Не удалось ни переместить MSG, ни записать "
                            "terminal-маркер — останавливаю обработку во "
                            "избежание повторной регистрации"
                        )
                        break
                else:  # FAILED
                    move_to_errors(msg_path, folder,
                                   f"Регистрация не удалась (status={status})")
                    err_count += 1
                _print_doc_line(i, len(docs), status,
                                 doc.get("тема", "")[:60])
            except ExclusionMarkerError as e:
                log.critical(f"ОСТАНОВКА на документе {i}: {e}")
                err_count += 1
                _print_doc_line(i, len(docs), "FAILED", str(e))
                break
            except Exception as e:
                log.error(f"ОШИБКА документ {i}: {e}")
                err_count += 1
                _print_doc_line(i, len(docs), "FAILED", str(e)[:80])
                if doc.get("_processed_via_asud_api"):
                    # Unknown API delivery must never become a Selenium retry
                    # or move out of the downloader's deduplication root.
                    try:
                        _mark_api_terminal(
                            msg_path,
                            "MANUAL_REVIEW",
                            f"Непредвиденная ошибка интеграции после API: {e}",
                        )
                    except ExclusionMarkerError:
                        log.critical(
                            "API terminal-маркер не записан — "
                            "останавливаю обработку"
                        )
                    break
                move_to_errors(msg_path, folder, f"Exception: {e}")
                if driver is not None:
                    try:
                        driver.get(url)
                        wait_asud_loaded(driver)
                    except Exception:
                        pass
                continue

        elapsed_seconds = time.monotonic() - start_time
        elapsed = timedelta(seconds=int(elapsed_seconds))
        avg = (timedelta(seconds=int(elapsed_seconds / done_count))
               if done_count else None)
        # Лог-файл — плоский (без ANSI), консоль — цветной
        if api_backend:
            plain = [
                "",
                "=" * 60,
                "ГОТОВО — GIS API ONE-SHOT",
                f"  API OK:       {done_count} / {len(docs)}  "
                "(MSG + terminal-marker в корне)",
                f"  Dry-run:      {dry_run_count}  (без marker, можно повторить)",
                f"  Probe:        {probe_count}  (без marker, можно повторить)",
                f"  Ручная проверка: {manual_review_count}  "
                "(terminal-marker, автоповтор запрещён)",
                f"  API до мутации: {api_failed_count} ошибок  "
                "(без marker, повтор безопасен после исправления)",
                f"  Исключено:    {excluded_count}  (в АСУД не передавались)",
                f"  Всего ошибок: {err_count}  (MSG не перемещались)",
                f"  Затрачено:    {elapsed}"
                + (f"  (в среднем {avg}/док)" if avg else ""),
                "=" * 60,
            ]
        else:
            plain = [
                "",
                "=" * 60,
                "ГОТОВО!",
                f"  Обработано:   {done_count} / {len(docs)}  (→ Завершено/)",
                f"  Дубликаты:    {dup_count}  (уже были в АСУД, → Завершено/)",
                f"  В черновиках: {draft_count}  (ФИО не найдено, .msg остался в корне)",
                f"  Исключено:    {excluded_count}  (в АСУД не передавались)",
                f"  Ошибок:       {err_count}  (→ Ошибки/)",
                f"  Затрачено:    {elapsed}"
                + (f"  (в среднем {avg}/док)" if avg else ""),
                "=" * 60,
            ]
        for line in plain:
            log.info(line)
        print("")
        print("=" * 60)
        if api_backend:
            print("ГОТОВО — GIS API ONE-SHOT")
            print(f"  API OK:       {green(str(done_count))} / {len(docs)}  "
                  "(MSG + marker в корне)")
            print(f"  Dry-run:      {yellow(str(dry_run_count))}  (без marker)")
            print(f"  Probe:        {yellow(str(probe_count))}  (без marker)")
            print(f"  Ручная проверка: {red(str(manual_review_count))}  "
                  "(автоповтор запрещён)")
            print(f"  API до мутации: {red(str(api_failed_count))} ошибок  "
                  "(повтор безопасен после исправления)")
            print(f"  Исключено:    {yellow(str(excluded_count))}")
            print(f"  Всего ошибок: {red(str(err_count))}  (MSG не перемещались)")
        else:
            print("ГОТОВО!")
            print(f"  Обработано:   {green(str(done_count))} / {len(docs)}  (→ Завершено/)")
            print(f"  Дубликаты:    {dup_count}  (уже были в АСУД, → Завершено/)")
            print(f"  В черновиках: {yellow(str(draft_count))}  (ФИО не найдено, .msg в корне)")
            print(f"  Исключено:    {yellow(str(excluded_count))}  (в АСУД не передавались)")
            print(f"  Ошибок:       {red(str(err_count))}  (→ Ошибки/)")
        print(f"  Затрачено:    {elapsed}"
              + (f"  (в среднем {avg}/док)" if avg else ""))
        print("=" * 60)

        # Раньше тут запускался ZHKH-второй проход (Басманов → резолюции
        # Халецкой). Теперь это отдельный процесс — zhkh_daemon, который
        # читает xlsx-реестр и отписывает в фоне.
        # Запускается отдельно: АСУД_ЖКХ_резолюции_Халецкой.bat
        if output_suffix == "ГИСЖКХ" and registered_docs:
            log.info(f"Зарегистрировано {len(registered_docs)} ГИСЖКХ-документов. "
                     f"Для отписки запусти АСУД_ЖКХ_резолюции_Халецкой.bat "
                     f"(или он уже работает фоном).")

        input("\nEnter для закрытия...")
    except Exception as e:
        log.error(f"Ошибка: {e}")
        input("Enter...")
    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception:
                pass
        cfg.keep_system_awake(False)


# ================= DAEMON MODE =================

# Sigint handler — устанавливает флаг, текущий документ доделывается, потом выход.
_stop_flag = False


def _on_sigint(signum, frame):
    global _stop_flag
    if _stop_flag:
        log.warning("Повторный Ctrl+C — выход немедленно")
        sys.exit(130)
    _stop_flag = True
    log.info("Ctrl+C получен — закончу текущий документ и выйду из мониторинга")


def _interruptible_sleep(seconds):
    """Sleep с пробуждением по _stop_flag (проверка раз в секунду)."""
    for _ in range(int(seconds)):
        if _stop_flag:
            return
        time.sleep(1)


def daemon_main():
    """Непрерывный мониторинг папки: опрос раз в N сек, обработка новых .msg.
    Ctrl+C для остановки (graceful — после текущего документа)."""
    global settings
    settings = cfg.load()
    cfg.setup_file_logger("email_daemon")
    cfg.keep_system_awake(True)

    log.info("=" * 50)
    log.info("АСУД ИК — Email-DAEMON (непрерывный мониторинг)")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()
    interval = int(settings.get("email_watch_interval_sec",
                                cfg.DEFAULTS["email_watch_interval_sec"]))
    max_retries = int(settings.get("email_max_retries",
                                    cfg.DEFAULTS["email_max_retries"]))
    process_mode = os.environ.get('ASUD_EMAIL_PROCESS_MODE', 'mix')
    output_suffix = os.environ.get('ASUD_OUTPUT_SUFFIX') or None

    if _api_backend_selected():
        log.critical(
            "GIS API backend поддерживает только однопроходный запуск. "
            "WATCH/daemon остановлен до запуска Edge и до обработки MSG."
        )
        cfg.keep_system_awake(False)
        raise SystemExit(2)

    # Multi-folder режим — пресет с "folders" списком вместо "folder" строки.
    # Если ASUD_EMAIL_FOLDERS_JSON задан, обрабатываем все эти папки
    # (по очереди если round_robin=true, либо все за тик если false).
    import json as _json
    folders_list = []
    round_robin_mode = False
    raw_json = os.environ.get('ASUD_EMAIL_FOLDERS_JSON')
    if raw_json:
        try:
            raw_list = _json.loads(raw_json)
            # Нормализуем элементы: string → {dir, output_suffix=basename},
            # dict → как есть
            for e in raw_list:
                if isinstance(e, dict):
                    if e.get("dir"):
                        folders_list.append({
                            "dir": e["dir"],
                            "output_suffix": e.get("output_suffix") or os.path.basename(e["dir"]),
                        })
                elif isinstance(e, str):
                    folders_list.append({
                        "dir": e,
                        "output_suffix": os.path.basename(e),
                    })
            round_robin_mode = os.environ.get('ASUD_EMAIL_ROUND_ROBIN') == '1'
        except Exception as e:
            log.error(f"ASUD_EMAIL_FOLDERS_JSON не распарсился: {e}")
            folders_list = []

    log.info(f"Логика обработки: {process_mode}"
             + (f", суффикс реестра: {output_suffix}" if output_suffix and not folders_list else ""))

    if folders_list:
        # Multi-folder: показываем список, валидируем (но не падаем если что-то
        # не нашлось — daemon будет переопрашивать)
        log.info(f"Multi-folder режим: {len(folders_list)} папок"
                 + (" (round-robin)" if round_robin_mode else " (все за тик)"))
        for e in folders_list:
            ok, new_dir = _wait_for_folder(e["dir"])
            e["dir"] = new_dir  # сохраним возможный UNC-fix
            marker = "✓" if ok else "⚠"
            log.info(f"  {marker} {e['dir']} (suffix={e['output_suffix']})")
        # folder для legacy-вызовов (move_to_errors и т.п. перед стартом loop'а)
        # просто берём первую папку. Дальше в loop'е current_folder per-tick.
        folder = folders_list[0]["dir"]
    else:
        # Single-folder (legacy): спрашиваем папку у юзера
        default = os.environ.get('ASUD_EMAIL_FOLDER') \
                  or settings.get("email_folder", "")
        print(f"\nПапка с .msg-письмами для непрерывного мониторинга.")
        if default:
            print(f"Enter — использовать: {default}")
        user_dir = input("Путь: ").strip().strip('"').strip("'")
        folder = user_dir or default
        if folder:
            ok, folder = _wait_for_folder(folder)
        else:
            ok = False
        if not ok:
            log.error(f"Папка не найдена (после ретраев): {folder!r}")
            input("Enter...")
            sys.exit(1)
        log.info(f"Папка: {folder}")

    log.info(f"Опрос: каждые {interval} сек, макс retry: {max_retries}")
    print(f"\nМониторинг включён. Ctrl+C для остановки.")

    # Browser (API daemon has already failed closed above).
    driver = None
    driver_path = os.path.join(base_dir, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error(f"msedgedriver.exe не найден в {base_dir}")
        input("Enter...")
        sys.exit(1)

    options = cfg.build_edge_options()
    service = EdgeService(executable_path=driver_path)
    driver = webdriver.Edge(service=service, options=options)
    set_driver_timeout(driver, settings.get("asud_load_timeout_sec",
                                              cfg.DEFAULTS["asud_load_timeout_sec"]))
    mix_flow.settings = settings

    signal.signal(signal.SIGINT, _on_sigint)

    url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
    log.info(f"Открываю {url}")
    driver.get(url)
    wait_asud_loaded(driver)

    # Счётчики и retry-state
    retry_count = {}  # абсолютный путь MSG → int (фейлов подряд)
    totals = {
        "OK": 0,
        "DUPLICATE": 0,
        "DRAFT": 0,
        "EXCLUDED": 0,
        "FAILED": 0,
        "ITER": 0,
    }
    rr_idx = 0  # для round-robin между папками

    def _process_folder(current_folder, current_suffix):
        """Обрабатывает все .msg из current_folder. Логика как раньше,
        просто вынесена в функцию чтобы вызываться для каждой папки в
        multi-folder режиме. Использует closure: driver, base_dir и т.д."""
        global _stop_flag
        queue = _list_root_msgs(current_folder)
        if not queue:
            return 0  # ничего не было
        log.info(f"[итер. {totals['ITER']}] {os.path.basename(current_folder)}: "
                 f"в очереди {len(queue)}")
        for idx, msg_path in enumerate(queue, 1):
            if _stop_flag:
                return idx
            name = os.path.basename(msg_path)
            retry_key = os.path.normcase(os.path.abspath(msg_path))

            doc = _parse_one_msg(msg_path, process_mode)
            if doc is None:
                move_to_errors(msg_path, current_folder,
                               "Не удалось распарсить или пустое")
                totals["FAILED"] += 1
                retry_count.pop(retry_key, None)
                _print_doc_line(idx, len(queue), "FAILED",
                                 "не распарсилось / пустое")
                continue

            try:
                status, _asud_id, _xlsx = _process_doc(
                                       driver, doc, base_dir, current_folder,
                                       idx, len(queue), in_daemon=True,
                                       process_mode=process_mode,
                                       output_suffix=current_suffix)
                if status in {"REGISTERED_ONLY", "SUBMISSION_UNKNOWN"}:
                    # Повторять такой MSG нельзя: документ уже зарегистрирован,
                    # иначе следующая итерация создаст дубль вместо доведения
                    # существующей карточки до резолюции.
                    reason = (
                        "Документ зарегистрирован, но резолюция не подтверждена"
                        if status == "REGISTERED_ONLY" else
                        "Результат регистрации в АСУД не определён"
                    ) + "; требуется ручная проверка"
                    quarantined = _quarantine_terminal_msg(
                        msg_path, current_folder, status, reason
                    )
                    retry_count.pop(retry_key, None)
                    totals["FAILED"] += 1
                    _print_doc_line(
                        idx, len(queue), "FAILED",
                        ("зарегистрирован без подтверждённой резолюции"
                         if status == "REGISTERED_ONLY" else
                         "результат регистрации не определён") + " → Ошибки/",
                    )
                    if not quarantined:
                        log.critical(
                            "Не удалось ни переместить MSG, ни записать "
                            "terminal-маркер — останавливаю daemon во "
                            "избежание повторной регистрации"
                        )
                        _stop_flag = True
                        return idx
                    try:
                        driver.get(url)
                        wait_asud_loaded(driver)
                    except Exception:
                        pass
                elif status == "FAILED":
                    retry_count[retry_key] = retry_count.get(retry_key, 0) + 1
                    if retry_count[retry_key] >= max_retries:
                        move_to_errors(msg_path, current_folder,
                            f"Регистрация не удалась за {max_retries} попыток")
                        retry_count.pop(retry_key, None)
                        totals["FAILED"] += 1
                        _print_doc_line(idx, len(queue), "FAILED",
                                         f"max_retries ({max_retries}) → Ошибки/")
                    else:
                        log.warning(f"{name}: фейл {retry_count[retry_key]}/{max_retries} "
                                    f"— оставляю в корне на следующую итерацию")
                        _print_doc_line(idx, len(queue), "FAILED",
                                         f"retry {retry_count[retry_key]}/{max_retries}")
                    try:
                        driver.get(url)
                        wait_asud_loaded(driver)
                    except Exception:
                        pass
                else:
                    totals[status] = totals.get(status, 0) + 1
                    retry_count.pop(retry_key, None)
                    _print_doc_line(idx, len(queue), status,
                                     doc.get("тема", "")[:60])
            except ExclusionMarkerError as e:
                log.critical(f"ОСТАНОВКА на {name}: {e}")
                totals["FAILED"] += 1
                _print_doc_line(idx, len(queue), "FAILED", str(e))
                _stop_flag = True
                return idx
            except Exception as e:
                log.error(f"Exception на {name}: {e}")
                retry_count[retry_key] = retry_count.get(retry_key, 0) + 1
                if retry_count[retry_key] >= max_retries:
                    move_to_errors(msg_path, current_folder, f"Exception: {e}")
                    retry_count.pop(retry_key, None)
                    totals["FAILED"] += 1
                _print_doc_line(idx, len(queue), "FAILED", str(e)[:80])
                try:
                    driver.get(url)
                    wait_asud_loaded(driver)
                except Exception:
                    pass
        return len(queue)

    try:
        while not _stop_flag:
            totals["ITER"] += 1

            # Какие папки обрабатываем на этом тике
            if folders_list and round_robin_mode:
                entry = folders_list[rr_idx % len(folders_list)]
                rr_idx += 1
                log.info(f"[итер. {totals['ITER']}] round-robin: "
                         f"{os.path.basename(entry['dir'])}")
                tick_folders = [entry]
            elif folders_list:
                tick_folders = folders_list
            else:
                tick_folders = [{"dir": folder, "output_suffix": output_suffix}]

            processed = 0
            for entry in tick_folders:
                if _stop_flag:
                    break
                processed += _process_folder(entry["dir"], entry.get("output_suffix"))

            if processed == 0:
                log.info(f"[итер. {totals['ITER']}] очереди пусты — sleep {interval}s")

            if _stop_flag:
                break
            log.info(f"  итог итер. {totals['ITER']}: "
                     f"OK={totals['OK']} DUP={totals['DUPLICATE']} "
                     f"DRAFT={totals['DRAFT']} EXCLUDED={totals['EXCLUDED']} "
                     f"FAIL={totals['FAILED']}")
            print(f"  итог итер. {totals['ITER']}: "
                  f"OK={green(str(totals['OK']))} DUP={totals['DUPLICATE']} "
                  f"DRAFT={yellow(str(totals['DRAFT']))} "
                  f"EXCLUDED={yellow(str(totals['EXCLUDED']))} "
                  f"FAIL={red(str(totals['FAILED']))}")
            _interruptible_sleep(interval)

        log.info("=" * 60)
        log.info("МОНИТОРИНГ ОСТАНОВЛЕН")
        log.info(f"  Итераций:   {totals['ITER']}")
        log.info(f"  Обработано: {totals['OK']}")
        log.info(f"  Дубликаты:  {totals['DUPLICATE']}")
        log.info(f"  Черновики:  {totals['DRAFT']}")
        log.info(f"  Исключено:  {totals['EXCLUDED']}")
        log.info(f"  Ошибки:     {totals['FAILED']}")
        log.info("=" * 60)
        print("=" * 60)
        print("МОНИТОРИНГ ОСТАНОВЛЕН")
        print(f"  Итераций:   {totals['ITER']}")
        print(f"  Обработано: {green(str(totals['OK']))}")
        print(f"  Дубликаты:  {totals['DUPLICATE']}")
        print(f"  Черновики:  {yellow(str(totals['DRAFT']))}")
        print(f"  Исключено:  {yellow(str(totals['EXCLUDED']))}")
        print(f"  Ошибки:     {red(str(totals['FAILED']))}")
        print("=" * 60)

    finally:
        if driver is not None:
            try:
                driver.quit()
            except Exception:
                pass
        cfg.keep_system_awake(False)


if __name__ == "__main__":
    if os.environ.get("ASUD_WATCH") == "1":
        daemon_main()
    else:
        main()
