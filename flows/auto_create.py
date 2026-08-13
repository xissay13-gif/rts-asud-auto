"""
auto_create_correspondent.py — Пакетное создание Входящих документов в АСУД ИК.

Читает Excel (B=Содержание, C=Корреспондент), автоматически создаёт
корреспондента если не найден. Регистрирует + На резолюцию.

Модули:
  config.py       — настройки (+ config.json)
  ui.py           — Selenium UI-хелперы
  correspondent.py — создание корреспондентов
  attachments.py  — прикрепление файлов (пустышка)
"""

import json
import hashlib
import math
import os
import re
import sys
import time
import logging
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from shared import config as cfg
from shared.ui import (click, wait_and_click, find_input_near_label,
                wait_asud_loaded, wait_modal_closed, close_open_modals,
                js_set_value, js_type_combobox, find_dropdown_options,
                wait_pointer_events_auto, is_duplicate_warning,
                set_driver_timeout)
from shared.correspondent import fill_correspondent_field, match_correspondent
from shared.attachments import get_dummy_msg, attach_content
from shared.xlsx_format import format_registry_before_save
from shared.registration import run_registration


# ================= LOGGING =================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
)
log = logging.getLogger("asud")
start_time = time.monotonic()


# ================= REGISTRATION STATE =================

STATE_SUFFIX = ".asud_auto_create_state.json"
TERMINAL_REGISTRATION_STATUSES = frozenset({
    "OK",
    "DUPLICATE",
    "REGISTERED_ONLY",
    "SUBMISSION_UNKNOWN",
    # Write-ahead marker. If the process is killed after the click, there is no
    # safe way to distinguish an accepted request from a no-op on restart.
    "SUBMISSION_PENDING",
})
REGISTRATION_STATUSES = TERMINAL_REGISTRATION_STATUSES | {"RETRYABLE"}


def _registration_state_path(excel_path):
    """Sidecar рядом с исходным Excel, привязанный к его имени."""
    source = Path(excel_path).resolve()
    return source.with_name(source.stem + STATE_SUFFIX)


def load_registration_state(excel_path):
    """Читает sidecar; повреждённый файл блокирует небезопасный повтор."""
    path = _registration_state_path(excel_path)
    if not path.is_file():
        return {"version": 1, "documents": {}}
    try:
        state = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(state, dict) or not isinstance(
                state.get("documents"), dict):
            raise ValueError("поле documents отсутствует или имеет неверный тип")
        state.setdefault("version", 1)
        return state
    except Exception as exc:
        raise RuntimeError(
            f"Не удалось прочитать state {path}: {exc}. "
            "Автоповтор заблокирован во избежание дубля."
        ) from exc


def save_registration_state(excel_path, state):
    """Атомарно сохраняет sidecar через temp-файл в той же папке."""
    path = _registration_state_path(excel_path)
    tmp = path.with_name(path.name + ".tmp")
    try:
        with open(tmp, "w", encoding="utf-8") as stream:
            json.dump(state, stream, ensure_ascii=False, indent=2)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(tmp, path)
    except Exception:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        raise


def _document_state(state, doc_data):
    source_key = doc_data.get("_source_key")
    if not source_key:
        return None
    return state.get("documents", {}).get(source_key)


def _document_is_terminal(state, doc_data):
    entry = _document_state(state, doc_data) or {}
    return entry.get("status") in TERMINAL_REGISTRATION_STATUSES


def _record_document_state(excel_path, state, doc_data, status, *,
                           asud_id=None, error=""):
    if status not in REGISTRATION_STATUSES:
        raise ValueError(f"Неизвестный registration status: {status}")
    source_key = doc_data.get("_source_key")
    source_row = doc_data.get("_source_row")
    if not source_key or not isinstance(source_row, int):
        raise ValueError("У документа отсутствуют _source_key/_source_row")
    documents = state.setdefault("documents", {})
    documents[source_key] = {
        "source_key": source_key,
        "source_row": source_row,
        "status": status,
        "asud_id": str(asud_id or ""),
        "error": str(error or "")[:500],
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    save_registration_state(excel_path, state)


def _registration_status(outcome):
    """Переводит outcome в безопасный persisted status.

    ``RETRYABLE`` означает, что state machine доказал отсутствие отправленного
    клика. Во всех неопределённых случаях write-ahead marker остаётся
    терминальным, чтобы перезапуск не создал дубль.
    """
    if outcome.registered and outcome.resolved:
        return "OK"
    if outcome.registered:
        return "REGISTERED_ONLY"
    if getattr(outcome, "submission_uncertain", False):
        return "SUBMISSION_UNKNOWN"
    return "RETRYABLE"


def _normalised_raw_row(row):
    """Каноническое представление полной Excel-строки для source key."""
    cells = []
    for value in row:
        if value is None:
            cells.append(["null", None])
        elif isinstance(value, bool):
            cells.append(["bool", value])
        elif isinstance(value, datetime):
            cells.append(["datetime", value.isoformat(timespec="microseconds")])
        elif isinstance(value, date):
            cells.append(["date", value.isoformat()])
        elif isinstance(value, int):
            cells.append(["number", str(value)])
        elif isinstance(value, float):
            numeric = format(value, ".15g") if math.isfinite(value) else repr(value)
            cells.append(["number", numeric])
        else:
            text = unicodedata.normalize("NFC", str(value))
            text = text.replace("\r\n", "\n").replace("\r", "\n").strip()
            cells.append(["text", text])
    return json.dumps(cells, ensure_ascii=False, separators=(",", ":"))


def _source_key_for_raw_row(row, identical_counts):
    canonical = _normalised_raw_row(row)
    ordinal = identical_counts.get(canonical, 0) + 1
    identical_counts[canonical] = ordinal
    digest = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    return f"row-sha256:{digest}:{ordinal}"


# ================= EXCEL =================

def load_excel(file_path):
    """Читает реестр обращений.

    Поддерживаемые форматы (определяются по заголовкам):
      • НОВЫЙ: лист 'результат', колонки № | LS | фио | обращение | ao | fio | тема
        — фио (cyr) = корреспондент, fio (lat) = исполнитель, тема = индекс из DOC_TYPE_MAP
      • СТАРЫЙ: B = содержание, C = корреспондент (первый лист)
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    # Лист 'результат' (новый) или active (старый)
    ws = wb['результат'] if 'результат' in wb.sheetnames else wb.active
    log.info(f"Лист: '{ws.title}'")

    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(max_row=1))]
    log.info(f"Заголовки: {headers}")

    def find_idx(*names):
        for name in names:
            for i, h in enumerate(headers):
                if h.lower() == name.lower():
                    return i
        return None

    # Новый формат: 'фио' (cyr) дважды — первое корреспондент, второе исполнитель
    fio_indices = [i for i, h in enumerate(headers) if h.lower() == 'фио']
    idx_content = find_idx('обращение', 'содержание', 'textbody')
    idx_corr = fio_indices[0] if fio_indices else find_idx('корреспондент')
    idx_executor = find_idx('fio')  # латиница
    idx_ao = find_idx('ao', 'округ')
    idx_type = find_idx('тема', 'тип')

    # Старый формат fallback
    if idx_content is None:
        idx_content = 1  # B
    if idx_corr is None:
        idx_corr = 2  # C

    log.info(f"Колонки: содержание={idx_content}, корреспондент={idx_corr}, "
             f"ao={idx_ao}, исполнитель={idx_executor}, тема={idx_type}")

    rows = []
    identical_counts = {}
    for source_row, row in enumerate(
            ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
        source_key = _source_key_for_raw_row(row, identical_counts)
        if not row:
            continue
        content = row[idx_content] if len(row) > idx_content else None
        corr = row[idx_corr] if len(row) > idx_corr else None
        if not (content and corr):
            continue
        item = {
            "содержание": str(content).strip(),
            "корреспондент": str(corr).strip(),
            "_source_row": source_row,
            "_source_key": source_key,
        }
        if idx_ao is not None and len(row) > idx_ao and row[idx_ao]:
            item["ao"] = str(row[idx_ao]).strip()
        if idx_executor is not None and len(row) > idx_executor and row[idx_executor]:
            item["исполнитель"] = str(row[idx_executor]).strip()
        if idx_type is not None and len(row) > idx_type and row[idx_type] is not None:
            try:
                item["тема_индекс"] = int(row[idx_type])
            except Exception:
                pass
        rows.append(item)
    wb.close()
    log.info(f"Загружено: {len(rows)} документов")
    return rows


# ================= FORM FILLING =================

def fill_text(driver, text):
    """Заполняет краткое содержание (textarea) — JS-set, plain-поле."""
    try:
        areas = driver.find_elements(By.TAG_NAME, "textarea")
        visible = [a for a in areas if a.is_displayed()]
        if visible:
            js_set_value(driver, visible[0], text)
            log.info("Краткое содержание заполнено (JS)")
        else:
            log.warning("Textarea не найдена")
    except Exception as e:
        log.error(f"Ошибка содержания: {e}")


def fill_corr_number(driver, index):
    """Заполняет 'Номер у корреспондента' = 'б/н (N)' — JS-set."""
    value = f"б/н ({index})"
    inp = find_input_near_label(driver, "Номер у корреспондента")
    if not inp:
        log.warning("Поле 'Номер у корреспондента' не найдено")
        return
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", inp)
        js_set_value(driver, inp, value)
        log.info(f"Номер (JS): {value}")
    except Exception as e:
        log.warning(f"Номер: ошибка {e}")


def fill_corr_date(driver):
    """Заполняет 'Дата у корреспондента' = сегодня."""
    from datetime import date
    today = date.today().strftime("%d.%m.%Y")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Дата у корреспондента']")
    inp = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 6):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                visible = [i for i in inputs
                           if i.is_displayed() and i.get_attribute("readonly") is None]
                if visible:
                    inp = visible[0]
                    break
            if inp:
                break
        except Exception:
            continue
    if not inp:
        log.warning("Поле 'Дата у корреспондента' не найдено")
        return
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", inp)
        js_set_value(driver, inp, today)
        log.info(f"Дата (JS): {today}")
    except Exception as e:
        log.warning(f"Дата: ошибка {e}")


def fill_delivery_method(driver):
    """Выбирает 'Электронная почта'."""
    target = settings.get("delivery_method", "Электронная почта")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Способ получения']")
    trigger = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 8):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                for sel in ["input[type='text']", "div[class*='trigger']", "img[class*='trigger']"]:
                    try:
                        el = parent.find_element(By.CSS_SELECTOR, sel)
                        if el.is_displayed():
                            trigger = el
                            break
                    except Exception:
                        continue
                if trigger:
                    break
            if trigger:
                break
        except Exception:
            continue
    if not trigger:
        log.warning("Поле 'Способ получения' не найдено")
        return
    click(driver, trigger, "Способ получения")

    # Ждём появления нужного пункта в дропдауне (вместо sleep(1.5)+цикла)
    def _option_visible(d):
        for c in d.find_elements(By.XPATH, f"//*[contains(text(),'{target}')]"):
            try:
                if c.is_displayed() and c.tag_name.lower() != 'input':
                    return c
            except Exception:
                continue
        return False
    try:
        opt = WebDriverWait(driver, 5).until(_option_visible)
        click(driver, opt, target)
        log.info(f"Способ получения: {target}")
    except Exception:
        log.warning(f"'{target}' не найдена в выпадашке")


def add_addressee(driver, person_name):
    """Добавляет адресата через combobox (JS-typing + WebDriverWait)."""
    inp = find_input_near_label(driver, "Адресаты")
    if not inp:
        log.warning("Поле адресата не найдено")
        return
    surname = person_name.split()[0]
    inp.click()
    js_type_combobox(driver, inp, surname)

    all_r = []
    try:
        WebDriverWait(driver, 5).until(
            lambda d: len(find_dropdown_options(d, surname, inp)) > 0)
        all_r = find_dropdown_options(driver, surname, inp)
    except Exception:
        try:
            inp.send_keys(Keys.ENTER)
            WebDriverWait(driver, 3).until(
                lambda d: len(find_dropdown_options(d, surname, inp)) > 0)
            all_r = find_dropdown_options(driver, surname, inp)
        except Exception:
            pass

    target = None
    for r in all_r:
        try:
            if match_correspondent(r.text, person_name):
                target = r
                break
        except Exception:
            continue
    if not target and all_r:
        target = all_r[0]
    if target:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
        ActionChains(driver).move_to_element(target).pause(0.2).click().perform()
        log.info(f"Адресат: {person_name}")
    else:
        log.warning(f"Адресат не найден: {person_name}")


# ================= REGISTRATION =================

_CAPTURE_ASUD_ID_JS = r"""
// Один проход по DOM на стороне браузера. Возвращает регистрационный
// номер или null.
// Реальная структура АСУД (из лога diag): номер лежит в body innerText
// в виде "№ ОРТС/8/20890 от 05.05.2026". В ScreenHeader1 — только панель
// действий, никаких <b> с номером.
// JS \b НЕ работает с кириллицей даже под флагом /u — поэтому используем
// явный контекст "№ ... от" для главного матча.
const RE_NUM_OT = /№\s+([А-Я]{2,5}(?:\/[А-Я0-9.\-]+){2,})\s+от/u;
const RE_LOOSE = /([А-Я]{2,5}(?:\/[А-Я0-9.\-]+){2,})/u;

function looksLike(t) {
    if (!t) return false;
    t = t.trim();
    if (!t.includes('/') || t.length < 6) return false;
    if (/^\d{2}\.\d{2}\.\d{4}/.test(t)) return false;
    if (!/\d/.test(t)) return false;
    return true;
}

// Главный путь: явный паттерн "№ ... от" в body.innerText
const body = document.body.innerText || '';
let m = body.match(RE_NUM_OT);
if (m && looksLike(m[1])) return m[1];

// Fallback 1: ScreenHeader1 → <b> (вдруг АСУД когда-нибудь начнёт класть сюда)
const header = document.querySelector("[data-marker='ScreenHeader1']");
if (header) {
    for (const b of header.querySelectorAll('b')) {
        const t = (b.textContent || '').trim();
        if (looksLike(t)) return t;
    }
}

// Fallback 2: общий regex по body без \b (захватит любую конструкцию вида ХХ/.../...)
m = body.match(RE_LOOSE);
if (m && looksLike(m[1])) return m[1];

return null;
"""


_DIAG_ASUD_ID_JS = r"""
// Диагностика: что в DOM, когда capture не сработал.
const out = {header_present: false, header_text: null, header_bolds: [], body_snippet: null};
const header = document.querySelector("[data-marker='ScreenHeader1']");
if (header) {
    out.header_present = true;
    out.header_text = (header.textContent || '').trim().slice(0, 300);
    for (const b of header.querySelectorAll('b')) {
        out.header_bolds.push((b.textContent || '').trim().slice(0, 100));
    }
}
out.body_snippet = (document.body.innerText || '').slice(0, 800).replace(/\s+/g, ' ');
return out;
"""


def capture_asud_id(driver, timeout=15):
    """Читает регистрационный номер документа после регистрации.
    Один JS-вызов на итерацию (быстрый поллинг 100ms).
    Строгий поиск ТОЛЬКО в ScreenHeader1, чтобы не зацепить тип служебки
    или другой id-подобный текст на странице."""
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        try:
            asud_id = driver.execute_script(_CAPTURE_ASUD_ID_JS)
            if asud_id:
                log.info(f"  asud_id: {asud_id!r}")
                return asud_id
        except Exception:
            pass
        time.sleep(0.1)
    # Диагностика: что было в DOM на момент таймаута
    try:
        diag = driver.execute_script(_DIAG_ASUD_ID_JS) or {}
        log.warning(f"  capture diag: header_present={diag.get('header_present')}")
        if diag.get('header_text'):
            log.warning(f"  header_text: {diag['header_text']!r}")
        for i, b in enumerate(diag.get('header_bolds', [])):
            log.warning(f"  header_bold[{i}]: {b!r}")
        if diag.get('body_snippet'):
            log.warning(f"  body[0:800]: {diag['body_snippet']!r}")
    except Exception:
        pass
    log.warning("Регистрационный номер не захватили — пуст в output")
    return None


def _post_register_check(driver, timeout=5):
    """После клика 'Зарегистрировать' проверяет что регистрация СОСТОЯЛАСЬ:
    'На резолюцию' должна стать видимой И НЕ disabled. Если нет — клик ушёл вхолостую.
    Возвращает элемент res_btn если успех, None если регистрация не сработала."""
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        try:
            res = driver.find_element(By.ID, "header-action-btn-send_on_resolution")
            if res.is_displayed():
                cls = (res.get_attribute('class') or '').lower()
                data_dis = res.get_attribute('data-disabled')
                aria_dis = res.get_attribute('aria-disabled')
                if data_dis != '1' and aria_dis != 'true' and 'disabled' not in cls:
                    return res
        except Exception:
            pass
        time.sleep(0.1)
    return None


def _legacy_register_and_resolve(driver, index, total):
    """Регистрирует + На резолюцию + Да; возвращает (registered, asud_id)."""
    log.info("Регистрирую...")
    registered = False
    asud_id = None
    res_btn = None
    try:
        btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                "#header-action-btn-register, [id*='header-action-btn-register']")))
        # GXT в transition-анимации после Save может ставить
        # pointer-events:none на кнопку — клик пройдёт впустую.
        # Ждём пока pointer-events станет 'auto' (макс 2s).
        wait_pointer_events_auto(driver, btn, timeout=2)
        click(driver, btn, "Зарегистрировать")

        # Verify: 'На резолюцию' стала enabled? Если нет — повтор клика
        res_btn = _post_register_check(driver, timeout=5)
        if not res_btn:
            log.warning(f"Документ {index}/{total}: клик 'Зарегистрировать' не дал эффекта — повторяю")
            close_open_modals(driver)
            try:
                btn = driver.find_element(By.CSS_SELECTOR,
                    "#header-action-btn-register, [id*='header-action-btn-register']")
                click(driver, btn, "Зарегистрировать (retry)")
                res_btn = _post_register_check(driver, timeout=5)
            except Exception as e:
                log.error(f"Retry клик упал: {e}")
            if not res_btn:
                log.error(f"Документ {index}/{total}: НЕ зарегистрирован — пропускаю 'На резолюцию'")
                return False, None
        registered = True
    except Exception:
        try:
            btn = driver.find_element(By.XPATH, "//div[contains(text(),'Зарегистрировать')]")
            click(driver, btn, "Зарегистрировать (fallback)")
            res_btn = _post_register_check(driver, timeout=5)
            if res_btn:
                registered = True
        except Exception as e:
            log.error(f"'Зарегистрировать' не найдена: {e}")

    if not registered:
        return False, None

    # res_btn уже получен из _post_register_check выше — не ждём заново.
    # Захват номера до 1.5s — если за это время не появился, идём дальше
    # (регистрация прошла, в xlsx запишем пусто но flow не ломается).
    asud_id = capture_asud_id(driver, timeout=1.5)
    if asud_id:
        log.info(f"Документ {index}/{total} ЗАРЕГИСТРИРОВАН: {asud_id}")
    else:
        log.warning(f"Документ {index}/{total} ЗАРЕГИСТРИРОВАН (номер не захватили)")

    if res_btn:
        click(driver, res_btn, "На резолюцию")

        # Ждём confirm-кнопку "Да" одним WebDriverWait — три селектора в lambda,
        # поллинг 500ms (vs прежний цикл с sleep(1))
        def _find_yes(d):
            for sel in [(By.ID, "confirm_dialog_btn_yes"),
                        (By.CSS_SELECTOR, "[id*='confirm_dialog_btn_yes'], [id*='confirm'][id*='yes']")]:
                try:
                    b = d.find_element(*sel)
                    if b.is_displayed():
                        return b
                except Exception:
                    continue
            for b in d.find_elements(By.XPATH, "//*[normalize-space(text())='Да']"):
                try:
                    if b.is_displayed():
                        return b
                except Exception:
                    continue
            return False
        try:
            yes_btn = WebDriverWait(driver, 10).until(_find_yes)
        except Exception:
            yes_btn = None

        if yes_btn:
            # Helper: re-find yes_btn freshly чтобы обойти stale element
            # после перерисовки GXT-диалога между нашими click-стратегиями.
            def _refind_yes():
                try:
                    b = driver.find_element(By.ID, "confirm_dialog_btn_yes")
                    if b.is_displayed():
                        return b
                except Exception:
                    pass
                try:
                    b = driver.find_element(By.CSS_SELECTOR,
                        "[id*='confirm_dialog_btn_yes'], [id*='confirm'][id*='yes']")
                    if b.is_displayed():
                        return b
                except Exception:
                    pass
                try:
                    for b in driver.find_elements(By.XPATH, "//*[normalize-space(text())='Да']"):
                        if b.is_displayed():
                            return b
                except Exception:
                    pass
                return None

            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", yes_btn)
                time.sleep(0.3)
            except Exception:
                pass
            clicked = False
            # 1) ActionChains
            try:
                fresh = _refind_yes() or yes_btn
                ActionChains(driver).move_to_element(fresh).pause(0.3).click().perform()
                log.info(f"Клик 'Да' (ActionChains): id={fresh.get_attribute('id')}")
                clicked = True
            except Exception as e:
                log.debug(f"ActionChains 'Да' не сработал: {type(e).__name__}")
            # 2) JS
            if not clicked:
                try:
                    fresh = _refind_yes() or yes_btn
                    driver.execute_script("arguments[0].click();", fresh)
                    log.info("Клик 'Да' (JS)")
                    clicked = True
                except Exception as e:
                    log.debug(f"JS 'Да' не сработал: {type(e).__name__}")
            # 3) Нативный
            if not clicked:
                try:
                    fresh = _refind_yes() or yes_btn
                    fresh.click()
                    log.info("Клик 'Да' (native)")
                    clicked = True
                except Exception as e:
                    log.debug(f"Native 'Да' не сработал: {type(e).__name__}")
            # 4) Enter — GWT-диалоги часто принимают его (отправляем
            # независимо от clicked: если кнопка stale всё равно закрывает диалог)
            try:
                ActionChains(driver).send_keys(Keys.ENTER).perform()
            except Exception:
                pass
            # Ждём пока модалка закроется (refind вернёт None когда уйдёт)
            try:
                WebDriverWait(driver, 5).until_not(
                    lambda d: bool(_refind_yes()))
            except Exception:
                pass
            log.info(f"Документ {index}/{total} НА РЕЗОЛЮЦИИ")
        else:
            log.warning("Диалог 'Да' не появился за 10 сек")
    else:
        log.warning("'На резолюцию' не появилась")
    return True, asud_id


def register_and_resolve(driver, index, total, *, before_register=None):
    """Общая проверяемая регистрация; возвращает подробный outcome.

    ``before_register`` — write-ahead barrier: если sidecar не удалось
    сохранить, вызов ``run_registration`` (и возможный клик) не выполняется.
    """
    if before_register is not None:
        before_register()
    outcome = run_registration(
        driver,
        timeout=max(20, cfg.DEFAULTS["timeout"]),
        retry_grace=2.5,
        capture_id=lambda current_driver: capture_asud_id(
            current_driver, timeout=1.5),
        logger=log,
    )
    if not outcome.registered:
        log.error(
            f"Документ {index}/{total}: регистрация не подтверждена: "
            f"{outcome.reason}"
        )
    elif not outcome.resolved:
        log.error(
            f"Документ {index}/{total}: зарегистрирован, но отправка "
            f"на резолюцию не подтверждена: {outcome.reason}"
        )
    else:
        log.info(f"Документ {index}/{total} НА РЕЗОЛЮЦИИ")
    return outcome


# ================= DOCUMENT FLOW =================

def create_one_document(driver, doc_data, index, total, *, before_register=None):
    """Создаёт один входящий документ; возвращает ``(status, asud_id)``."""
    log.info(f"{'='*50}")
    log.info(f"ДОКУМЕНТ {index}/{total}")
    log.info(f"Содержание: {doc_data['содержание'][:60]}...")
    log.info(f"Корреспондент: {doc_data['корреспондент']}")

    el = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
        EC.presence_of_element_located((By.ID, "mainscreen-create-button")))
    click(driver, el, "Создать документ")

    wait_and_click(driver, By.XPATH,
        "//div[contains(text(),'Входящий документ')]", "Входящий документ")

    # Тип/вид документа: если в реестре есть 'тема' (индекс) — используем
    # DOC_TYPE_MAP, иначе берём дефолт из настроек
    type_idx = doc_data.get("тема_индекс")
    if type_idx and type_idx in cfg.DOC_TYPE_MAP:
        subtype = cfg.DOC_TYPE_MAP[type_idx]
        log.info(f"Тип из реестра: {type_idx} → {subtype}")
    else:
        subtype = settings.get("doc_subtype", "Письма, заявления и жалобы граждан, акционеров")
    short = subtype[:30]
    wait_and_click(driver, By.XPATH,
        f"//div[contains(text(),'{short}')] | //td[contains(text(),'{short}')]", subtype)

    wait_and_click(driver, By.XPATH,
        "//button[contains(text(),'Создать документ')] | //div[contains(text(),'Создать документ')]",
        "Создать документ")

    # Ждём пока форма документа отрендерится — появится textarea (краткое содержание)
    try:
        WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            lambda d: any(t.is_displayed() for t in d.find_elements(By.TAG_NAME, "textarea")))
    except Exception:
        log.warning("Textarea формы не появилась")

    fill_text(driver, doc_data["содержание"])
    if not fill_correspondent_field(driver, doc_data["корреспондент"]):
        log.error("Корреспондент не выбран и не создан — документ остановлен")
        close_open_modals(driver)
        raise RuntimeError("поле Корреспондент не подтверждено")
    fill_corr_number(driver, index)
    fill_corr_date(driver)

    for person in settings.get("addressees", ["Басманов Александр Владимирович"]):
        add_addressee(driver, person)

    fill_delivery_method(driver)

    try:
        save_btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.element_to_be_clickable((By.ID, "header-save-btn")))
        click(driver, save_btn, "Сохранить")
        # Quick-check: 'уже зарегистрирован' warning от АСУД?
        time.sleep(1.5)
        if is_duplicate_warning(driver):
            log.warning(f"Документ {index}/{total}: АСУД говорит УЖЕ ЗАРЕГИСТРИРОВАН — пропускаю")
            close_open_modals(driver)
            try:
                driver.get(settings.get("asud_url", cfg.DEFAULTS["asud_url"]))
                wait_asud_loaded(driver)
            except Exception:
                pass
            return "DUPLICATE", None
        # Ждём появления "Зарегистрировать" — признак что save прошёл
        WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                "#header-action-btn-register, [id*='header-action-btn-register']")))
        log.info(f"Документ {index}/{total} сохранён")
    except Exception as e:
        log.error(f"Ошибка сохранения: {e}")
        close_open_modals(driver)
        raise RuntimeError("Сохранение документа не подтверждено") from e

    if doc_data.get("файл"):
        attached = attach_content(driver, doc_data["файл"])
        wait_modal_closed(driver)
        if attached:
            log.info(f"Документ {index}/{total}: вложение прикреплено ✓")
        else:
            log.warning(f"Документ {index}/{total}: вложение НЕ прикреплено ✗")

    outcome = register_and_resolve(
        driver, index, total, before_register=before_register)
    status = _registration_status(outcome)
    asud_id = outcome.asud_id
    if status != "OK":
        # Карточку не продолжаем трогать: main сохранит terminal status и
        # вернётся на главный экран. Повторная регистрация запрещена sidecar'ом.
        return status, asud_id

    # Ждём что появится первым: либо главный экран (карточка авто-закрылась),
    # либо close-btn (карточка ещё открыта). До 10s, но обычно <1s.
    def _state(d):
        try:
            main_btn = d.find_element(By.ID, "mainscreen-create-button")
            if main_btn.is_displayed() and main_btn.is_enabled():
                return ('main', main_btn)
        except Exception:
            pass
        try:
            close_btn = d.find_element(By.ID, "header-close-btn")
            if close_btn.is_displayed():
                return ('close', close_btn)
        except Exception:
            pass
        return False

    state = None
    try:
        state = WebDriverWait(driver, 10).until(_state)
    except Exception:
        pass

    if state and state[0] == 'close':
        try:
            ActionChains(driver).move_to_element(state[1]).pause(0.2).click().perform()
        except Exception:
            pass
        # После закрытия ждём главный экран
        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "mainscreen-create-button")))
        except Exception:
            driver.get(settings.get("asud_url", cfg.DEFAULTS["asud_url"]))
            wait_asud_loaded(driver)
    elif not state:
        # Не нашли ни одной из кнопок — перезагружаемся
        driver.get(settings.get("asud_url", cfg.DEFAULTS["asud_url"]))
        wait_asud_loaded(driver)
    return status, asud_id


# ================= OUTPUT XLSX (для clean-resolutions) =================

OKRUG_TO_FIO = {
    "САО": "Гренц Екатерина Александровна",
    "ЦАО": "Емельянова Татьяна Николаевна",
    "ОАО": "Рендюк Юлия Павловна",
    "ЛАО": "Вырва Елена Анатольевна",
    "КАО": "Кравец Татьяна Александровна",
}


def _output_xlsx_path(excel_path):
    """<реестр>_резолюции.xlsx в папке Registered/ рядом с exe."""
    registered_dir = os.path.join(cfg.get_base_dir(), "Registered")
    os.makedirs(registered_dir, exist_ok=True)
    name = os.path.splitext(os.path.basename(excel_path))[0]
    return os.path.join(registered_dir, name + "_резолюции.xlsx")


def _ensure_output_xlsx(path):
    if os.path.isfile(path):
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Резолюции"
        ws.append(["ОПТС", "Округ", "ФИО", "Link"])
        for c in range(1, 5):
            ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
        widths = {1: 30, 2: 8, 3: 35, 4: 22}
        for col, w in widths.items():
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = w
        ws.freeze_panes = "A2"
        format_registry_before_save(ws, path, changed_row=1)
        wb.save(path)
    except Exception as e:
        log.warning(f"Не удалось создать {path}: {e}")


def _append_output_row(path, doc_data, asud_id):
    """Дописывает в output xlsx: ОПТС | Округ | ФИО | Link.

    Округ/ФИО берутся из реестра (новый формат), при отсутствии —
    парсятся из 'Содержание' и/или маппятся через OKRUG_TO_FIO.
    """
    okrug = doc_data.get("ao") or None
    if not okrug:
        try:
            from shared.okrug_parser import okrug_from_textbody
            okrug = okrug_from_textbody(doc_data.get("содержание"),
                                         base_dir_fn=cfg.get_base_dir)
        except Exception as e:
            log.warning(f"okrug_parser упал: {e}")
    fio = doc_data.get("исполнитель") or (OKRUG_TO_FIO.get(okrug) if okrug else None)
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        ws.append([asud_id or "", okrug or "", fio or "", ""])
        format_registry_before_save(ws, path, changed_row=ws.max_row)
        wb.save(path)
        wb.close()
        log.info(f"  → {os.path.basename(path)}: "
                 f"{asud_id or '—'} | {okrug or '—'} | {fio or '—'}")
    except Exception as e:
        log.warning(f"Не удалось записать в {path}: {e}")


# ================= MAIN =================

settings = {}


def main():
    global settings
    settings = cfg.load()
    cfg.setup_file_logger("auto_create")
    cfg.keep_system_awake(True)

    log.info("=" * 50)
    log.info("АСУД ИК — Входящие + автосоздание корреспондентов")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()

    # Если xlsx уже выбран извне (через app.py) — используем его
    excel_path = os.environ.get('ASUD_XLSX')
    if not excel_path:
        xlsx_files = [f for f in os.listdir(base_dir) if f.lower().endswith('.xlsx')]
        if not xlsx_files:
            log.error(f"Нет .xlsx в {base_dir}")
            input("Enter...")
            sys.exit(1)
        elif len(xlsx_files) == 1:
            excel_path = os.path.join(base_dir, xlsx_files[0])
        else:
            print(f"\nНайдено {len(xlsx_files)} xlsx:")
            for i, f in enumerate(xlsx_files, 1):
                print(f"  {i}. {f}")
            choice = input("Номер: ").strip()
            try:
                excel_path = os.path.join(base_dir, xlsx_files[int(choice) - 1])
            except (ValueError, IndexError):
                log.error("Неверный выбор")
                sys.exit(1)

    # Папка с .msg-файлами (для прикрепления) — интерактивный ввод.
    # В auto-create обычно используется пустышка, но пользователь может
    # указать папку с реальными .msg чтобы каждый док получил «свой» файл.
    default_outlook = settings.get("outlook_dir", "") or cfg.DEFAULTS.get("outlook_dir", "")
    print(f"\nПапка с .msg-файлами (Enter — без поиска, прикрепляем пустышку).")
    if default_outlook:
        print(f"Enter — использовать: {default_outlook}")
    user_dir = input("Путь: ").strip().strip('"').strip("'")
    if user_dir:
        settings["outlook_dir"] = user_dir
    outlook_dir = settings.get("outlook_dir") or ""
    if outlook_dir and not os.path.isdir(outlook_dir):
        log.warning(f"Папка '{outlook_dir}' не существует — все вложения как пустышка")
    elif outlook_dir:
        log.info(f"Папка вложений: {outlook_dir}")

    msg_path = get_dummy_msg(base_dir)
    docs = load_excel(excel_path)
    for doc in docs:
        doc["файл"] = msg_path

    try:
        registration_state = load_registration_state(excel_path)
    except RuntimeError as exc:
        log.error(str(exc))
        input("Enter...")
        return

    if not docs:
        log.error("Нет данных!")
        input("Enter...")
        sys.exit(1)

    print(f"\nПервые 5:")
    for i, d in enumerate(docs[:5], 1):
        ao = d.get("ao") or ""
        ti = d.get("тема_индекс") or ""
        print(f"  {i}. [{ao:3} тема={ti}] {d['корреспондент']} | "
              f"{d['содержание'][:50]}...")
    mode_label = os.environ.get('ASUD_MODE', 'auto-create')
    print(f"\nВсего: {len(docs)}  •  режим: {mode_label.upper()}")
    print("auto-create: создание + регистрация + На резолюцию (без .msg)")

    if input("Начать? (да/нет): ").strip().lower() not in ("да", "д", "y", "yes", ""):
        sys.exit(0)

    driver_path = os.path.join(base_dir, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error(f"msedgedriver.exe не найден")
        input("Enter...")
        sys.exit(1)

    options = cfg.build_edge_options()

    driver = webdriver.Edge(service=EdgeService(executable_path=driver_path), options=options)
    set_driver_timeout(driver, settings.get("asud_load_timeout_sec",
                                              cfg.DEFAULTS["asud_load_timeout_sec"]))

    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        driver.get(url)
        wait_asud_loaded(driver)

        # Output xlsx: ОПТС|Округ|ФИО|Link рядом с реестром
        output_path = _output_xlsx_path(excel_path)
        _ensure_output_xlsx(output_path)
        log.info(f"Output xlsx: {output_path}")

        done_count, err_count, skipped_count = 0, 0, 0
        for i, doc in enumerate(docs, 1):
            previous = _document_state(registration_state, doc) or {}
            if _document_is_terminal(registration_state, doc):
                skipped_count += 1
                log.warning(
                    f"Документ {i}/{len(docs)} пропущен по state: "
                    f"{previous.get('status')} "
                    f"(строка {doc.get('_source_row')})"
                )
                continue

            def _write_ahead(current_doc=doc):
                _record_document_state(
                    excel_path,
                    registration_state,
                    current_doc,
                    "SUBMISSION_PENDING",
                )

            try:
                status, asud_id = create_one_document(
                    driver,
                    doc,
                    i,
                    len(docs),
                    before_register=_write_ahead,
                )
                _record_document_state(
                    excel_path,
                    registration_state,
                    doc,
                    status,
                    asud_id=asud_id,
                )
                if status in {"OK", "DUPLICATE"}:
                    _append_output_row(output_path, doc, asud_id)
                    done_count += 1
                elif status == "RETRYABLE":
                    err_count += 1
                    log.warning(
                        f"Документ {i}: клик регистрации точно не отправлен; "
                        "на следующем запуске строку можно повторить"
                    )
                    driver.get(url)
                    wait_asud_loaded(driver)
                else:
                    err_count += 1
                    log.error(
                        f"Документ {i}: terminal status {status}; "
                        "автоматический повтор запрещён, нужна ручная проверка"
                    )
                    driver.get(url)
                    wait_asud_loaded(driver)
            except Exception as e:
                log.error(f"ОШИБКА документ {i}: {e}")
                err_count += 1
                current = _document_state(registration_state, doc) or {}
                if current.get("status") == "SUBMISSION_PENDING":
                    try:
                        _record_document_state(
                            excel_path,
                            registration_state,
                            doc,
                            "SUBMISSION_UNKNOWN",
                            asud_id=current.get("asud_id"),
                            error=e,
                        )
                    except Exception as state_exc:
                        # На диске остаётся предыдущий атомарно записанный
                        # SUBMISSION_PENDING, поэтому перезапуск всё равно safe.
                        log.error(f"Не удалось обновить sidecar: {state_exc}")
                driver.get(url)
                wait_asud_loaded(driver)

        elapsed_seconds = time.monotonic() - start_time
        elapsed = timedelta(seconds=int(elapsed_seconds))
        avg = timedelta(seconds=int(elapsed_seconds / done_count)) if done_count else None

        summary = [
            "",
            "=" * 60,
            f"ГОТОВО!",
            f"  Обработано: {done_count} / {len(docs)}",
            f"  Пропущено:  {skipped_count}",
            f"  Ошибок:     {err_count}",
            f"  Затрачено:  {elapsed}" + (f"  (в среднем {avg}/док)" if avg else ""),
            "=" * 60,
        ]
        for line in summary:
            log.info(line)
            print(line)

        input("\nEnter для закрытия...")
    except Exception as e:
        log.error(f"Ошибка: {e}")
        input("Enter...")
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        cfg.keep_system_awake(False)


if __name__ == "__main__":
    main()
