"""
flows/smart.py — Smart-routing: создание Входящих документов в АСУД ИК
с прикреплением .msg, БЕЗ регистрации и без выдачи на резолюцию.

Документы остаются в черновиках для ручной проверки/регистрации.
Корреспондент — фиксированный «Неизвестный Неизвестный Неизвестный»
(переопределяется через settings['correspondent']).

Запускается через app.py с --mode=smart.
"""

import os
import re
import sys
import socket
import time
import logging
from datetime import date, datetime, timedelta

import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from shared import config as cfg
from shared.ui import (click, wait_and_click, find_input_near_label,
                        wait_asud_loaded, wait_modal_closed, close_open_modals,
                        js_set_value, js_type_combobox, find_dropdown_options,
                        set_driver_timeout)
from shared.correspondent import fill_correspondent_field, match_correspondent
from shared.addressee import add_addressee as _add_addressee_verified
from shared.attachments import find_msg_by_link, get_dummy_msg, attach_content, move_to_done


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger("asud.smart")
start_time = time.monotonic()

settings = {}


# ================= EXCEL =================

def _clean_body(text):
    """Очищает TextBody от служебных строк (warning о внешнем письме, Original Message)."""
    if not text:
        return ""
    t = str(text).replace('_x000D_', '\n')
    lines = []
    for line in t.split('\n'):
        s = line.strip()
        if re.search(r'внимание!?\s*письмо\s+было\s+отправлено\s+внешним', s, re.IGNORECASE):
            continue
        if re.match(r'^-{3,}\s*Original\s*Message\s*-{3,}$', s, re.IGNORECASE):
            continue
        lines.append(line)
    t = '\n'.join(lines)
    return re.sub(r'\n\s*\n\s*\n+', '\n\n', t).strip()


def load_excel(file_path):
    """Читает реестр. Колонки: A=Link, B=Subject, C=TextBody, D=Тип (1-8)."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    rows, skipped = [], 0
    correspondent = settings.get("correspondent",
                                 "Неизвестный Неизвестный Неизвестный")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or len(row) < 4:
            skipped += 1
            continue
        subject, body, type_idx = row[1], row[2], row[3]
        if not subject:
            skipped += 1
            continue
        try:
            type_idx = int(type_idx) if type_idx is not None else 0
        except (ValueError, TypeError):
            type_idx = 0
        if type_idx == 0 or type_idx not in cfg.DOC_TYPE_MAP:
            skipped += 1
            continue

        clean_subject = re.sub(r'^(FW:|RE:|Fwd:)\s*', '',
                               str(subject).strip(), flags=re.IGNORECASE)
        body_clean = _clean_body(body) if body else clean_subject

        rows.append({
            "содержание": body_clean,
            "корреспондент": correspondent,
            "тема": clean_subject,
            "тип_индекс": type_idx,
            "тип_название": cfg.DOC_TYPE_MAP[type_idx],
            "link": row[0],
        })
    wb.close()
    log.info(f"Загружено: {len(rows)}, пропущено: {skipped}")
    return rows


# ================= FORM FILLING =================

def fill_text(driver, text):
    """Заполняет краткое содержание (textarea) — JS-set."""
    try:
        areas = driver.find_elements(By.TAG_NAME, "textarea")
        visible = [a for a in areas if a.is_displayed()]
        if visible:
            js_set_value(driver, visible[0], text)
            log.info("Краткое содержание заполнено (JS)")
        else:
            log.warning("Textarea не найдена")
    except Exception as e:
        log.error(f"Ошибка содержания: {e}")


def fill_corr_number(driver, link=None):
    """Заполняет 'Номер у корреспондента' = 'б/н <link>' — JS-set."""
    if isinstance(link, (datetime, date)):
        link_str = link.strftime("%d.%m.%Y %H-%M-%S")
    elif link:
        link_str = str(link).strip()
    else:
        link_str = ""
    value = f"б/н {link_str}" if link_str else "б/н"

    inp = find_input_near_label(driver, "Номер у корреспондента")
    if not inp:
        log.warning("Поле 'Номер у корреспондента' не найдено")
        return
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        js_set_value(driver, inp, value)
        log.info(f"Номер (JS): {value}")
    except Exception as e:
        log.warning(f"Номер: ошибка {e}")


def fill_corr_date(driver):
    """Заполняет 'Дата у корреспондента' = сегодня — JS-set."""
    today = date.today().strftime("%d.%m.%Y")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Дата у корреспондента']")
    inp = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 6):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                visible = [i for i in inputs
                           if i.is_displayed() and i.get_attribute("readonly") is None]
                if visible:
                    inp = visible[0]
                    break
            if inp:
                break
        except Exception:
            continue
    if not inp:
        log.warning("Поле 'Дата у корреспондента' не найдено")
        return
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        js_set_value(driver, inp, today)
        log.info(f"Дата (JS): {today}")
    except Exception as e:
        log.warning(f"Дата: ошибка {e}")


def fill_delivery_method(driver):
    """Выбирает 'Электронная почта'."""
    target = settings.get("delivery_method", "Электронная почта")
    labels = driver.find_elements(By.XPATH,
        "//*[normalize-space(text())='Способ получения']")
    trigger = None
    for label in labels:
        try:
            if not label.is_displayed():
                continue
            for level in range(1, 8):
                parent = label
                for _ in range(level):
                    parent = parent.find_element(By.XPATH, "..")
                for sel in ["input[type='text']", "div[class*='trigger']", "img[class*='trigger']"]:
                    try:
                        el = parent.find_element(By.CSS_SELECTOR, sel)
                        if el.is_displayed():
                            trigger = el
                            break
                    except Exception:
                        continue
                if trigger:
                    break
            if trigger:
                break
        except Exception:
            continue
    if not trigger:
        log.warning("Поле 'Способ получения' не найдено")
        return
    click(driver, trigger, "Способ получения")

    def _option_visible(d):
        for c in d.find_elements(By.XPATH, f"//*[contains(text(),'{target}')]"):
            try:
                if c.is_displayed() and c.tag_name.lower() != 'input':
                    return c
            except Exception:
                continue
        return False
    try:
        opt = WebDriverWait(driver, 5).until(_option_visible)
        click(driver, opt, target)
        log.info(f"Способ получения: {target}")
    except Exception:
        log.warning(f"'{target}' не найдена")


def _add_addressee_legacy(driver, person_name):
    """Добавляет адресата через combobox (JS-typing + WebDriverWait)."""
    inp = find_input_near_label(driver, "Адресаты")
    if not inp:
        log.warning("Поле адресата не найдено")
        return
    surname = person_name.split()[0]
    inp.click()
    js_type_combobox(driver, inp, surname)

    all_r = []
    try:
        WebDriverWait(driver, 5).until(
            lambda d: len(find_dropdown_options(d, surname, inp)) > 0)
        all_r = find_dropdown_options(driver, surname, inp)
    except Exception:
        try:
            inp.send_keys(Keys.ENTER)
            WebDriverWait(driver, 3).until(
                lambda d: len(find_dropdown_options(d, surname, inp)) > 0)
            all_r = find_dropdown_options(driver, surname, inp)
        except Exception:
            pass

    target = None
    for r in all_r:
        try:
            if match_correspondent(r.text, person_name):
                target = r
                break
        except Exception:
            continue
    if not target and all_r:
        target = all_r[0]
    if target:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
        ActionChains(driver).move_to_element(target).pause(0.2).click().perform()
        log.info(f"Адресат: {person_name}")
    else:
        log.warning(f"Адресат не найден: {person_name}")


def add_addressee(driver, person_name):
    """Совместимый wrapper общей fail-closed реализации."""
    return _add_addressee_verified(driver, person_name, logger=log)


# ================= DOCUMENT FLOW =================

def create_one_document(driver, doc_data, index, total):
    """Создаёт один документ-черновик (без регистрации)."""
    log.info(f"{'='*50}")
    log.info(f"ДОКУМЕНТ {index}/{total}: {doc_data['тема'][:60]}")
    log.info(f"Корреспондент: {doc_data['корреспондент']}")
    log.info(f"Тип: {doc_data['тип_название']}")

    # [1/6] Создать документ → Входящий документ → Вид
    el = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
        EC.presence_of_element_located((By.ID, "mainscreen-create-button")))
    click(driver, el, "Создать документ")

    wait_and_click(driver, By.XPATH,
        "//div[contains(text(),'Входящий документ')]", "Входящий документ")

    subtype = doc_data["тип_название"]
    short = subtype[:30]
    wait_and_click(driver, By.XPATH,
        f"//div[contains(text(),'{short}')] | //td[contains(text(),'{short}')]", subtype)

    wait_and_click(driver, By.XPATH,
        "//button[contains(text(),'Создать документ')] | //div[contains(text(),'Создать документ')]",
        "Создать документ")

    # Ждём textarea как маркер готовности формы
    try:
        WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            lambda d: any(t.is_displayed() for t in d.find_elements(By.TAG_NAME, "textarea")))
    except Exception:
        log.warning("Textarea формы не появилась")

    # [2/6] Заполнение
    fill_text(driver, doc_data["содержание"])
    if not fill_correspondent_field(driver, doc_data["корреспондент"]):
        log.error("Корреспондент не выбран и не создан — документ остановлен")
        close_open_modals(driver)
        raise RuntimeError("поле Корреспондент не подтверждено")
    fill_corr_number(driver, doc_data.get("link"))
    fill_corr_date(driver)
    if not add_addressee(driver, settings.get(
            "addressee", "Басманов Александр Владимирович")):
        log.error("Адресат не подтверждён — документ остановлен до сохранения")
        close_open_modals(driver)
        raise RuntimeError("поле Адресат не подтверждено")
    fill_delivery_method(driver)

    # [3/6] Сохранить (документ становится черновиком)
    try:
        save_btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.element_to_be_clickable((By.ID, "header-save-btn")))
        click(driver, save_btn, "Сохранить")
        # После Save должен появиться 'Зарегистрировать' (форма ушла в режим
        # черновика-готового-к-регистрации) — но мы НЕ регистрируем.
        WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                "#header-action-btn-register, [id*='header-action-btn-register']")))
        log.info(f"Документ {index}/{total} сохранён в черновиках")
    except Exception as e:
        log.error(f"Ошибка сохранения: {e}")
        close_open_modals(driver)
        raise RuntimeError("Сохранение черновика не подтверждено") from e

    # [4/6] Прикрепление .msg по Link (или dummy)
    outlook_dir = settings.get("outlook_dir", cfg.DEFAULTS["outlook_dir"])
    dummy_path = doc_data.get("файл")
    attach_path = find_msg_by_link(doc_data.get("link"), outlook_dir, dummy_path)
    if attach_path:
        log.info(f"Прикрепляю: {os.path.basename(attach_path)}")
        attached = attach_content(driver, attach_path)
        wait_modal_closed(driver)
        if attached:
            log.info(f"Документ {index}/{total}: вложение прикреплено ✓")
        else:
            log.warning(f"Документ {index}/{total}: вложение НЕ прикреплено ✗")
        # Реальный (не dummy) файл — переносим в Завершено/.
        # В smart-routing критерий — успешный attach (документ в черновиках).
        if attach_path != dummy_path:
            move_to_done(attach_path, outlook_dir)
    else:
        log.info("Нет файла — пропускаю")

    # [5/6] Закрыть карточку (без регистрации — остаётся в черновиках)
    log.info(f"Документ {index}/{total} в черновиках")
    closed = False
    # Сначала ждём появления видимой кнопки header-close-btn (АСУД иногда
    # создаёт несколько таких элементов — скрытые модалки + видимая карточка).
    try:
        WebDriverWait(driver, 5).until(
            lambda d: any(b.is_displayed()
                          for b in d.find_elements(By.ID, "header-close-btn")))
    except Exception:
        log.debug("header-close-btn не появился за 5s")

    candidates = driver.find_elements(By.ID, "header-close-btn")
    for btn in candidates:
        try:
            if not btn.is_displayed():
                continue
        except Exception:
            continue
        # Multi-strategy click
        try:
            ActionChains(driver).move_to_element(btn).pause(0.2).click().perform()
            closed = True
            break
        except Exception:
            try:
                driver.execute_script("arguments[0].click();", btn)
                closed = True
                break
            except Exception:
                continue

    if not closed:
        # Fallback — ESC. GXT-карточки часто реагируют.
        try:
            ActionChains(driver).send_keys(Keys.ESCAPE).perform()
            log.debug("ESC отправлен для закрытия карточки")
        except Exception:
            pass

    # [6/6] Дождаться главной
    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "mainscreen-create-button")))
    except Exception:
        log.warning("Главная не появилась — перезагружаю страницу АСУД")
        driver.get(settings.get("asud_url", cfg.DEFAULTS["asud_url"]))
        wait_asud_loaded(driver)


# ================= MAIN =================

def _is_port_open(host, port, timeout=0.5):
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.settimeout(timeout)
    try:
        s.connect((host, port))
        s.close()
        return True
    except Exception:
        return False


def main():
    global settings
    settings = cfg.load()
    cfg.setup_file_logger("smart")
    cfg.keep_system_awake(True)

    log.info("=" * 50)
    log.info("АСУД ИК — Smart-routing (черновики с прикреплением .msg)")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()
    excel_path = os.environ.get('ASUD_XLSX')
    if not excel_path:
        xlsx_files = [f for f in os.listdir(base_dir) if f.lower().endswith('.xlsx')]
        if not xlsx_files:
            log.error(f"Нет .xlsx в {base_dir}")
            input("Enter...")
            sys.exit(1)
        elif len(xlsx_files) == 1:
            excel_path = os.path.join(base_dir, xlsx_files[0])
        else:
            print(f"\nНайдено {len(xlsx_files)} xlsx:")
            for i, f in enumerate(xlsx_files, 1):
                print(f"  {i}. {f}")
            choice = input("Номер: ").strip()
            try:
                excel_path = os.path.join(base_dir, xlsx_files[int(choice) - 1])
            except (ValueError, IndexError):
                log.error("Неверный выбор")
                sys.exit(1)

    # Папка с .msg — приоритет: ASUD_OUTLOOK_DIR env (от пресета) → интерактивно
    env_outlook = os.environ.get('ASUD_OUTLOOK_DIR', '').strip()
    if env_outlook:
        settings["outlook_dir"] = env_outlook
        log.info(f"Папка вложений из пресета: {env_outlook}")
    else:
        default_outlook = settings.get("outlook_dir", "") or cfg.DEFAULTS.get("outlook_dir", "")
        print(f"\nПапка с .msg-файлами (поиск рекурсивно по подпапкам).")
        if default_outlook:
            print(f"Enter — использовать: {default_outlook}")
        else:
            print(f"Enter — без поиска .msg (все документы с пустышкой)")
        user_dir = input("Путь: ").strip().strip('"').strip("'")
        if user_dir:
            settings["outlook_dir"] = user_dir
    outlook_dir = settings.get("outlook_dir") or ""
    if outlook_dir and not os.path.isdir(outlook_dir):
        log.warning(f"Папка '{outlook_dir}' не существует — все вложения как пустышка")
    elif outlook_dir:
        log.info(f"Папка вложений: {outlook_dir}")

    msg_path = get_dummy_msg(base_dir)
    docs = load_excel(excel_path)
    if not docs:
        log.error("Реестр пуст")
        input("Enter...")
        sys.exit(1)

    for d in docs:
        d["файл"] = msg_path

    print(f"\nПервые 5:")
    for i, d in enumerate(docs[:5], 1):
        print(f"  {i}. [{d['тип_индекс']}] {d['тема'][:50]}...")
    mode_label = os.environ.get('ASUD_MODE', 'smart')
    print(f"\nВсего: {len(docs)}  •  режим: {mode_label.upper()}")
    print(f"smart: создание ЧЕРНОВИКА + .msg (без регистрации)")
    print(f"корреспондент = {docs[0]['корреспондент']}")
    if input("Начать? (да/нет): ").strip().lower() not in ("да", "д", "y", "yes", ""):
        sys.exit(0)

    # Edge — общие опции из shared/config (учитывают ASUD_HEADLESS)
    options = cfg.build_edge_options()

    driver_path = os.path.join(base_dir, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error("msedgedriver.exe не найден")
        input("Enter...")
        sys.exit(1)

    driver = webdriver.Edge(service=EdgeService(executable_path=driver_path),
                             options=options)
    set_driver_timeout(driver, settings.get("asud_load_timeout_sec",
                                              cfg.DEFAULTS["asud_load_timeout_sec"]))

    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        driver.get(url)
        wait_asud_loaded(driver)

        done, err = 0, 0
        for i, doc in enumerate(docs, 1):
            try:
                create_one_document(driver, doc, i, len(docs))
                done += 1
            except Exception as e:
                log.error(f"ОШИБКА документ {i}: {e}")
                err += 1
                try:
                    driver.get(url)
                    wait_asud_loaded(driver)
                except Exception:
                    pass

        elapsed = timedelta(seconds=int(time.monotonic() - start_time))
        avg = (timedelta(seconds=int((time.monotonic() - start_time) / done))
               if done else None)
        summary = [
            "",
            "=" * 60,
            "ГОТОВО! (документы в черновиках)",
            f"  Обработано: {done} / {len(docs)}",
            f"  Ошибок:     {err}",
            f"  Затрачено:  {elapsed}" + (f"  (в среднем {avg}/док)" if avg else ""),
            "=" * 60,
        ]
        for line in summary:
            log.info(line)
            print(line)
        input("\nEnter для закрытия...")
    except Exception as e:
        log.error(f"Ошибка: {e}")
        input("Enter...")
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        cfg.keep_system_awake(False)


if __name__ == "__main__":
    main()
