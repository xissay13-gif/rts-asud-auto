"""
mix_routing.py — Пакетное создание Входящих документов (auto-create + smart-routing).

Читает Excel (Лист2), извлекает ФИО корреспондента из TextBody.
Для каждой строки:
  - Определяет тип документа по индексу (колонка D)
  - Создаёт карточку, заполняет поля
  - Прикрепляет .msg из D:\\OutlookSubjects по Link
  - Если ФИО найдено → регистрирует + На резолюцию + Да
  - Если ФИО не найдено → корреспондент="Неизвестный...", оставляет в ЧЕРНОВИКАХ
    + WARNING в логе (чтобы вручную доработать после прогона)

Excel Лист2: A=Link, B=Subject, C=TextBody, D=Тип, E=To (игнорируем — пересыльщик).

Модули:
  config.py        — настройки (+ config.json)
  ui.py            — Selenium UI-хелперы
  correspondent.py — выбор/создание корреспондента + extract_fio_from_text
  attachments.py   — поиск и прикрепление файлов
"""

import os
import re
import sys
import json
import time
import logging
from datetime import date, datetime, timedelta

import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from shared import config as cfg
from shared.ui import (click, wait_and_click, find_input_near_label,
                wait_asud_loaded, wait_modal_closed, close_open_modals,
                js_set_value, js_type_combobox, find_dropdown_options,
                wait_pointer_events_auto, is_duplicate_warning,
                set_driver_timeout)
from shared.correspondent import (fill_correspondent_field, match_strict, fio_to_initials,
                           extract_fio_from_text)
from shared.addressee import add_addressee as _add_addressee_verified
from shared.attachments import find_msg_by_link, get_dummy_msg, attach_content, move_to_done
from shared.xlsx_format import format_registry_before_save
from shared.registration import run_registration


# ================= LOGGING =================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger("asud")
start_time = time.monotonic()


# ================= EXCEL =================

def _clean_body(text):
    """Очищает TextBody от служебных строк (Original Message, ВНИМАНИЕ…)."""
    if not text:
        return ""
    t = str(text).replace('_x000D_', '\n')
    lines = t.split('\n')
    cleaned = []
    for line in lines:
        s = line.strip()
        if re.search(r'внимание!?\s*письмо\s+было\s+отправлено\s+внешним', s, re.IGNORECASE):
            continue
        if re.match(r'^-{3,}\s*Original\s*Message\s*-{3,}$', s, re.IGNORECASE):
            continue
        cleaned.append(line)
    t = '\n'.join(cleaned)
    t = re.sub(r'\n\s*\n\s*\n+', '\n\n', t)
    return t.strip()


# ================= STATE (resume после крэша) =================

def _link_key(link):
    """Стабильный строковый ключ из Link (для state-файла)."""
    if link is None:
        return ""
    if isinstance(link, (datetime, date)):
        try:
            return link.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return str(link)
    return str(link).strip()


def _state_path(xlsx_path):
    """Путь к state-файлу рядом с exe, привязан к имени Excel."""
    name = os.path.splitext(os.path.basename(xlsx_path))[0]
    # Безопасное имя файла
    safe = re.sub(r'[^\w.\-]+', '_', name)
    return os.path.join(cfg.get_base_dir(), f"mix_state_{safe}.json")


def load_state(xlsx_path):
    """Загружает set обработанных Link-ключей."""
    path = _state_path(xlsx_path)
    if not os.path.isfile(path):
        return set()
    try:
        with open(path, encoding='utf-8') as f:
            data = json.load(f)
        return set(data.get('processed', []))
    except Exception as e:
        log.warning(f"Не удалось прочитать state {path}: {e}")
        return set()


def save_state(xlsx_path, processed_set):
    """Атомарно перезаписывает state-файл."""
    path = _state_path(xlsx_path)
    try:
        tmp = path + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump({'processed': sorted(processed_set)}, f,
                      ensure_ascii=False, indent=2)
        os.replace(tmp, path)
    except Exception as e:
        log.warning(f"Не удалось сохранить state {path}: {e}")


# ================= OUTPUT XLSX (для clean-resolutions) =================

# Округ → ФИО начальника абонентского отдела
OKRUG_TO_FIO = {
    "САО": "Гренц Екатерина Александровна",
    "ЦАО": "Емельянова Татьяна Николаевна",
    "ОАО": "Рендюк Юлия Павловна",
    "ЛАО": "Вырва Елена Анатольевна",
    "КАО": "Кравец Татьяна Александровна",
}


def _output_xlsx_path(excel_path):
    """<реестр>_резолюции.xlsx в папке Registered/ рядом с exe."""
    registered_dir = os.path.join(cfg.get_base_dir(), "Registered")
    os.makedirs(registered_dir, exist_ok=True)
    name = os.path.splitext(os.path.basename(excel_path))[0]
    return os.path.join(registered_dir, name + "_резолюции.xlsx")


def _ensure_output_xlsx(path):
    """Создаёт output xlsx с шапкой если не существует.

    Колонки 1-4 — данные документа, 5 — статус/дата обработки,
    6-7 — заполняются позже: «Отписано Халецкой» (Басманов→Халецкая,
    второй проход ГИСЖКХ); «Отписано в округ» (Халецкая→окружные,
    clean-resolutions).
    """
    if os.path.isfile(path):
        return
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Резолюции"
        headers = ["ОПТС", "Округ", "ФИО", "Link", "Статус",
                   "Отписано Халецкой", "Отписано в округ"]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
        widths = {1: 30, 2: 8, 3: 35, 4: 22, 5: 28, 6: 18, 7: 18}
        for col, w in widths.items():
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = w
        ws.freeze_panes = "A2"
        format_registry_before_save(ws, path, changed_row=1)
        wb.save(path)
    except Exception as e:
        log.warning(f"Не удалось создать {path}: {e}")


def _append_output_row(path, doc_data, asud_id, status="OK"):
    """Дописывает строку в output xlsx после обработки документа.
    Колонки: ОПТС | Округ | ФИО | Link | Статус

    status:
      OK        → «Зарегистрирован DD.MM.YYYY HH:MM»
      DRAFT     → «Черновик DD.MM.YYYY HH:MM»
      DUPLICATE → «Дубликат DD.MM.YYYY HH:MM»
      REGISTERED_ONLY → «Зарегистрирован без резолюции ...»
      SUBMISSION_UNKNOWN → «Результат регистрации не определён ...»
    """
    # Округ — пытаемся определить автоматически из TextBody
    okrug = None
    try:
        from shared.okrug_parser import okrug_from_textbody
        okrug = okrug_from_textbody(doc_data.get("содержание"),
                                     base_dir_fn=cfg.get_base_dir)
    except Exception as e:
        log.warning(f"okrug_parser упал: {e}")
    fio = OKRUG_TO_FIO.get(okrug) if okrug else None
    link = doc_data.get("link")
    link_str = ""
    if link:
        if isinstance(link, (datetime, date)):
            link_str = link.strftime("%d.%m.%Y %H-%M-%S")
        else:
            link_str = str(link)
    ts = datetime.now().strftime("%d.%m.%Y %H:%M")
    status_text = {
        "OK":        f"Зарегистрирован {ts}",
        "DRAFT":     f"Черновик {ts}",
        "DUPLICATE": f"Дубликат {ts}",
        "REGISTERED_ONLY": f"Зарегистрирован без резолюции {ts}",
        "SUBMISSION_UNKNOWN": f"Результат регистрации не определён {ts}",
    }.get(status, f"{status} {ts}")
    try:
        from shared.xlsx_lock import xlsx_lock
        with xlsx_lock(path, timeout=60):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            ws.append([asud_id or "", okrug or "", fio or "", link_str, status_text])
            format_registry_before_save(ws, path, changed_row=ws.max_row)
            wb.save(path)
            wb.close()
        log.info(f"  → {os.path.basename(path)}: "
                 f"{asud_id or '—'} | {okrug or '—'} | {fio or '—'}")
    except Exception as e:
        log.warning(f"Не удалось записать в {path}: {e}")


def load_excel(file_path):
    """Читает Лист2. Колонки: A=Link, B=Subject, C=TextBody, D=Тип (index).

    ФИО корреспондента извлекается из TextBody.
    Если не найдено — ставится заглушка (unknown_correspondent) и флаг corr_found=False
    → такой документ потом останется в черновиках.
    """
    sheet_name = settings.get("sheet_name", cfg.DEFAULTS["sheet_name"])
    unknown = settings.get("unknown_correspondent", cfg.DEFAULTS["unknown_correspondent"])

    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        log.warning(f"Лист '{sheet_name}' не найден, использую активный: {wb.active.title}")
        ws = wb.active

    rows = []
    skipped = 0
    unknown_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
        if not row or len(row) < 4:
            skipped += 1
            continue
        link = row[0]
        subject = row[1]
        body = row[2]
        type_idx = row[3]

        if not subject:
            skipped += 1
            continue
        try:
            type_idx = int(type_idx) if type_idx is not None else 0
        except (ValueError, TypeError):
            type_idx = 0
        if type_idx == 0 or type_idx not in cfg.DOC_TYPE_MAP:
            skipped += 1
            continue

        clean_subject = re.sub(r'^(FW:|RE:|Fwd:)\s*', '',
                               str(subject).strip(), flags=re.IGNORECASE)
        body_clean = _clean_body(body) if body else clean_subject

        fio, fio_src = extract_fio_from_text(body)
        if fio:
            correspondent = fio
            corr_found = True
        else:
            correspondent = unknown
            corr_found = False
            unknown_rows.append((row_idx, clean_subject))

        rows.append({
            "row_idx": row_idx,
            "содержание": body_clean,
            "корреспондент": correspondent,
            "корр_найден": corr_found,
            "корр_источник": fio_src,
            "тема": clean_subject,
            "тип_индекс": type_idx,
            "тип_название": cfg.DOC_TYPE_MAP[type_idx],
            "link": link,
        })
    wb.close()

    log.info(f"Загружено: {len(rows)}, пропущено: {skipped}")
    log.info(f"  ФИО найдено: {sum(1 for r in rows if r['корр_найден'])}")
    log.info(f"  ФИО НЕ найдено: {len(unknown_rows)} (уйдут в черновики)")
    for ri, subj in unknown_rows:
        log.info(f"    → Row {ri}: {subj[:60]}")

    return rows


# ================= FORM FILLING =================

def fill_text(driver, text):
    """Заполняет краткое содержание (textarea) — JS-set.

    Приоритет: textarea внутри #asudik-form-description (стабильный id из дампа).
    Fallback: первая видимая textarea на странице.
    """
    try:
        area = None
        try:
            container = driver.find_element(By.ID, "asudik-form-description")
            cands = container.find_elements(By.TAG_NAME, "textarea")
            for c in cands:
                if c.is_displayed():
                    area = c
                    break
        except Exception:
            pass
        if not area:
            areas = driver.find_elements(By.TAG_NAME, "textarea")
            visible = [a for a in areas if a.is_displayed()]
            area = visible[0] if visible else None
        if area:
            js_set_value(driver, area, text)
            log.info("Краткое содержание заполнено (JS)")
        else:
            log.warning("Textarea для краткого содержания не найдена")
    except Exception as e:
        log.error(f"Ошибка заполнения содержания: {e}")


def fill_corr_number(driver, link=None, override=None):
    """Заполняет 'Номер у корреспондента'.
    override (если задан) — использовать как есть (для ZHKH: реальный № обращения).
    Иначе — формат 'б/н <link>' (текущее поведение)."""
    if override:
        value = str(override).strip()
    else:
        if isinstance(link, (datetime, date)):
            link_str = link.strftime("%d.%m.%Y %H-%M-%S")
        elif link:
            link_str = str(link).strip()
        else:
            link_str = ""
        value = f"б/н {link_str}" if link_str else "б/н"

    # Приоритет — стабильный id #IncomingView-crspRegNumber (по дампу).
    inp = None
    try:
        container = driver.find_element(By.ID, "IncomingView-crspRegNumber")
        cands = container.find_elements(By.CSS_SELECTOR, "input[type='text']")
        for c in cands:
            if c.is_displayed():
                inp = c
                break
    except Exception:
        pass
    if not inp:
        inp = find_input_near_label(driver, "Номер у корреспондента")
    if not inp:
        log.warning("Поле 'Номер у корреспондента' не найдено")
        return
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        js_set_value(driver, inp, value)
        log.info(f"Номер (JS): {value}")
    except Exception as e:
        log.warning(f"Номер: ошибка {e}")


def fill_corr_date(driver, override=None):
    """Заполняет 'Дата у корреспондента'.
    override (если задан) — использовать как есть (для ZHKH: реальная дата
    получения обращения, формат DD.MM.YYYY). Иначе — сегодня.

    Приоритет — стабильный id #IncomingView-crspRegDate (по дампу).
    Fallback — поиск по label-text + parent-walk.
    """
    today = override.strip() if override else date.today().strftime("%d.%m.%Y")
    # Защита: override может прийти со временем («29.06.2026 15:36») — поле
    # даты в АСУД принимает только DD.MM.YYYY, иначе остаётся пустым.
    _m = re.match(r'(\d{2}\.\d{2}\.\d{4})', today)
    if _m:
        today = _m.group(1)
    inp = None
    try:
        container = driver.find_element(By.ID, "IncomingView-crspRegDate")
        cands = container.find_elements(By.CSS_SELECTOR, "input")
        for c in cands:
            if c.is_displayed() and c.get_attribute("readonly") is None:
                inp = c
                break
    except Exception:
        pass
    if not inp:
        labels = driver.find_elements(By.XPATH,
            "//*[normalize-space(text())='Дата у корреспондента']")
        for label in labels:
            try:
                if not label.is_displayed():
                    continue
                for level in range(1, 6):
                    parent = label
                    for _ in range(level):
                        parent = parent.find_element(By.XPATH, "..")
                    inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                    visible = [i for i in inputs
                               if i.is_displayed() and i.get_attribute("readonly") is None]
                    if visible:
                        inp = visible[0]
                        break
                if inp:
                    break
            except Exception:
                continue

    if not inp:
        log.warning("Поле 'Дата у корреспондента' не найдено")
        return
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        js_set_value(driver, inp, today)
        log.info(f"Дата (JS): {today}")
    except Exception as e:
        log.warning(f"Дата: ошибка {e}")


def fill_delivery_method(driver):
    """Выбирает 'Электронная почта' в 'Способ получения'.

    Приоритет — стабильный id #IncomingView-deliveryType (по дампу).
    Fallback — поиск по label.
    """
    target_text = settings.get("delivery_method", "Электронная почта")
    trigger = None
    try:
        container = driver.find_element(By.ID, "IncomingView-deliveryType")
        cands = container.find_elements(By.CSS_SELECTOR, "input")
        for c in cands:
            if c.is_displayed():
                trigger = c
                break
        if not trigger:
            for sel in ["div[class*='trigger']", "img[class*='trigger']"]:
                try:
                    el = container.find_element(By.CSS_SELECTOR, sel)
                    if el.is_displayed():
                        trigger = el
                        break
                except Exception:
                    continue
    except Exception:
        pass

    if not trigger:
        labels = driver.find_elements(By.XPATH,
            "//*[normalize-space(text())='Способ получения']")
        for label in labels:
            try:
                if not label.is_displayed():
                    continue
                for level in range(1, 8):
                    parent = label
                    for _ in range(level):
                        parent = parent.find_element(By.XPATH, "..")
                    inputs = parent.find_elements(By.CSS_SELECTOR, "input[type='text']")
                    for i in inputs:
                        if i.is_displayed():
                            trigger = i
                            break
                    if trigger:
                        break
                    for sel in ["div[class*='trigger']", "img[class*='trigger']"]:
                        try:
                            el = parent.find_element(By.CSS_SELECTOR, sel)
                            if el.is_displayed():
                                trigger = el
                                break
                        except Exception:
                            continue
                    if trigger:
                        break
                if trigger:
                    break
            except Exception:
                continue

    if not trigger:
        log.warning("Поле 'Способ получения' не найдено")
        return

    click(driver, trigger, "Способ получения")
    time.sleep(1.5)

    option = None
    for _ in range(3):
        candidates = driver.find_elements(By.XPATH,
            f"//*[contains(text(),'{target_text}')]")
        for c in candidates:
            try:
                if c.is_displayed() and c.tag_name.lower() != 'input':
                    option = c
                    break
            except Exception:
                continue
        if option:
            break
        time.sleep(1)

    if option:
        click(driver, option, target_text)
        log.info(f"Способ получения: {target_text}")
    else:
        log.warning(f"'{target_text}' не найдена в списке")


def _addressee_chip_present(driver, inp, surname):
    """True если рядом с input появился chip с фамилией адресата
    (или сам input содержит её). Проверяется в parent-цепочке вверх до 5 уровней.
    """
    surname_lower = surname.lower()
    try:
        val = (inp.get_attribute("value") or "").lower()
        if surname_lower in val:
            return True
    except Exception:
        pass
    try:
        parent = inp
        for _ in range(5):
            parent = parent.find_element(By.XPATH, "..")
            txts = parent.find_elements(By.XPATH,
                f".//*[contains(normalize-space(text()), '{surname}')]")
            for el in txts:
                try:
                    if el.is_displayed() and el != inp:
                        return True
                except Exception:
                    continue
    except Exception:
        pass
    return False


def _poll_addressee_chip(driver, inp, surname, timeout=1.0):
    """Поллинг до timeout секунд — ждём появления chip-а.
    Дефолт уменьшен до 1s т.к. в trust-mode мы не блокируемся на проверке."""
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        if _addressee_chip_present(driver, inp, surname):
            return True
        time.sleep(0.2)
    return False


def _add_addressee_legacy(driver, person_name):
    """Добавляет одного адресата через combobox.
    Multi-strategy click + проверка появления chip — GXT не всегда
    пропагирует value сразу после ActionChains.click.

    Приоритет input'а: внутри #addressee_grid_id (стабильный id).
    Fallback: поиск по label «Адресаты».
    """
    from shared.ui import cdp_click
    inp = None
    try:
        grid = driver.find_element(By.ID, "addressee_grid_id")
        cands = grid.find_elements(By.CSS_SELECTOR, "input")
        for c in cands:
            if c.is_displayed() and c.get_attribute("readonly") is None:
                inp = c
                break
    except Exception:
        pass
    if not inp:
        inp = find_input_near_label(driver, "Адресаты")
    if not inp:
        log.warning("Поле адресата не найдено (ни в #addressee_grid_id, ни по label)")
        return

    surname = person_name.split()[0]
    inp.click()
    js_type_combobox(driver, inp, surname)

    from shared.correspondent import match_correspondent

    all_results = []
    try:
        WebDriverWait(driver, 5).until(
            lambda d: len(find_dropdown_options(d, surname, inp)) > 0)
        all_results = find_dropdown_options(driver, surname, inp)
    except Exception:
        try:
            inp.send_keys(Keys.ENTER)
            WebDriverWait(driver, 3).until(
                lambda d: len(find_dropdown_options(d, surname, inp)) > 0)
            all_results = find_dropdown_options(driver, surname, inp)
        except Exception:
            pass

    target = None
    for r in all_results:
        try:
            if match_correspondent(r.text, person_name):
                target = r
                break
        except Exception:
            continue
    if not target and all_results:
        target = all_results[0]

    if not target:
        log.warning(f"Адресат не найден в выпадашке: {person_name}")
        return

    # Найти parent-option (как в correspondent) — для GXT клик нужен по option,
    # не по внутреннему span'у.
    parent_option = None
    try:
        parent_option = driver.execute_script("""
            let el = arguments[0];
            for (let i = 0; i < 6 && el; i++) {
                el = el.parentElement;
                if (!el) break;
                const role = el.getAttribute('role');
                const cls = (el.className || '').toString().toLowerCase();
                if (role === 'option' ||
                    /\\b(option|item|menu-item|select-option|combo-item|boundlist-item|ListItem|SelectItem)\\b/i.test(cls) ||
                    /gxt-\\w*item|x-combo-list-item|x-boundlist-item/i.test(cls)) {
                    return el;
                }
            }
            return null;
        """, target)
    except Exception:
        pass

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target)
    except Exception:
        pass

    # TRUST-MODE: один CDP-click и доверяем визуальному результату. Раньше
    # пробовали 4 стратегии (≈15 секунд на адресат). В новой АСУД UI
    # заполняется, но наш XPath не находит chip в ancestor-tree.
    cdp_target = parent_option if parent_option is not None else target
    log.debug(f"Адресат: CDP click по {'parent-option' if parent_option else 'target'}")
    click_ok = cdp_click(driver, cdp_target)
    chip_ok = _poll_addressee_chip(driver, inp, surname) if click_ok else False
    if chip_ok:
        log.info(f"Адресат добавлен (CDP, chip найден): {person_name}")
        return
    if click_ok:
        log.info(f"Адресат: CDP-click отправлен, chip не найден в DOM — "
                 f"доверяем визуальному результату")
        return

    # CDP сам не сработал — fallback ActionChains
    log.debug("Адресат: CDP не сработал, fallback ActionChains")
    try:
        ActionChains(driver).move_to_element(target).pause(0.2).click().perform()
    except Exception as e:
        log.debug(f"  AC click err: {e}")
    if _poll_addressee_chip(driver, inp, surname):
        log.info(f"Адресат добавлен (AC fallback): {person_name}")
        return

    log.warning(f"Адресат '{person_name}': клик прошёл, chip не появился — "
                f"идём дальше (возможно value в скрытом узле)")


# The legacy implementation above looked for a non-existent "chip" in the
# search input ancestors and trusted a dispatched click. Every production call
# is routed through the verified grid-row implementation below.
def add_addressee(driver, person_name):
    return _add_addressee_verified(driver, person_name, logger=log)


# ================= REGISTRATION =================

_CAPTURE_ASUD_ID_JS = r"""
// Один проход по DOM на стороне браузера. Возвращает регистрационный
// номер или null.
// Реальная структура АСУД (из лога diag): номер лежит в body innerText
// в виде "№ ОРТС/8/20890 от 05.05.2026". JS \b НЕ работает с кириллицей
// даже под /u — поэтому используем явный контекст "№ ... от".
const RE_NUM_OT = /№\s+([А-Я]{2,5}(?:\/[А-Я0-9.\-]+){2,})\s+от/u;
const RE_LOOSE = /([А-Я]{2,5}(?:\/[А-Я0-9.\-]+){2,})/u;

function looksLike(t) {
    if (!t) return false;
    t = t.trim();
    if (!t.includes('/') || t.length < 6) return false;
    if (/^\d{2}\.\d{2}\.\d{4}/.test(t)) return false;
    if (!/\d/.test(t)) return false;
    return true;
}

// Главный путь: явный паттерн "№ ... от" в body.innerText
const body = document.body.innerText || '';
let m = body.match(RE_NUM_OT);
if (m && looksLike(m[1])) return m[1];

// Fallback 1: ScreenHeader1 → <b>
const header = document.querySelector("[data-marker='ScreenHeader1']");
if (header) {
    for (const b of header.querySelectorAll('b')) {
        const t = (b.textContent || '').trim();
        if (looksLike(t)) return t;
    }
}

// Fallback 2: общий regex по body без \b
m = body.match(RE_LOOSE);
if (m && looksLike(m[1])) return m[1];

return null;
"""


_DIAG_ASUD_ID_JS = r"""
// Диагностика: что в DOM, когда capture не сработал.
const out = {header_present: false, header_text: null, header_bolds: [], body_snippet: null};
const header = document.querySelector("[data-marker='ScreenHeader1']");
if (header) {
    out.header_present = true;
    out.header_text = (header.textContent || '').trim().slice(0, 300);
    for (const b of header.querySelectorAll('b')) {
        out.header_bolds.push((b.textContent || '').trim().slice(0, 100));
    }
}
out.body_snippet = (document.body.innerText || '').slice(0, 800).replace(/\s+/g, ' ');
return out;
"""


def capture_asud_id(driver, timeout=15):
    """Читает регистрационный номер документа после регистрации.
    Один JS-вызов на итерацию (быстрый поллинг 100ms).
    Возвращает 'ОПТС/8/19892' / 'ОРТС/СЗ/ТЭС-03-УПО-02/176' или None."""
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        try:
            asud_id = driver.execute_script(_CAPTURE_ASUD_ID_JS)
            if asud_id:
                log.info(f"  asud_id: {asud_id!r}")
                return asud_id
        except Exception:
            pass
        time.sleep(0.1)

    # Диагностика: что было в DOM на момент таймаута
    try:
        diag = driver.execute_script(_DIAG_ASUD_ID_JS) or {}
        log.warning(f"  capture diag: header_present={diag.get('header_present')}")
        if diag.get('header_text'):
            log.warning(f"  header_text: {diag['header_text']!r}")
        for i, b in enumerate(diag.get('header_bolds', [])):
            log.warning(f"  header_bold[{i}]: {b!r}")
        if diag.get('body_snippet'):
            log.warning(f"  body[0:800]: {diag['body_snippet']!r}")
    except Exception:
        pass
    log.warning("Регистрационный номер не захватили — пуст в output")
    return None


def _post_register_check(driver, timeout=5):
    """После клика 'Зарегистрировать' проверяет что регистрация СОСТОЯЛАСЬ:
    'На резолюцию' должна стать видимой И НЕ disabled. Если нет — клик ушёл вхолостую.
    Возвращает элемент res_btn если успех, None если регистрация не сработала."""
    end = time.monotonic() + timeout
    while time.monotonic() < end:
        try:
            res = driver.find_element(By.ID, "header-action-btn-send_on_resolution")
            if res.is_displayed():
                cls = (res.get_attribute('class') or '').lower()
                data_dis = res.get_attribute('data-disabled')
                aria_dis = res.get_attribute('aria-disabled')
                if data_dis != '1' and aria_dis != 'true' and 'disabled' not in cls:
                    return res
        except Exception:
            pass
        time.sleep(0.1)
    return None


def _legacy_register_and_resolve(driver, index, total):
    """Регистрирует + На резолюцию + Да.

    Возвращает пару ``(registered, asud_id)``. Номер может не захватиться даже
    при состоявшейся регистрации, поэтому он не служит флагом успеха.
    """
    log.info("Регистрирую...")
    registered = False
    asud_id = None
    res_btn = None
    try:
        btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                "#header-action-btn-register, [id*='header-action-btn-register']")))
        # GXT в transition-анимации после Save может ставить
        # pointer-events:none на кнопку — клик пройдёт впустую.
        # Ждём пока pointer-events станет 'auto' (макс 2s).
        wait_pointer_events_auto(driver, btn, timeout=2)
        click(driver, btn, "Зарегистрировать")

        # Verify: 'На резолюцию' стала enabled? Если нет — повтор клика
        res_btn = _post_register_check(driver, timeout=5)
        if not res_btn:
            log.warning(f"Документ {index}/{total}: клик 'Зарегистрировать' не дал эффекта — повторяю")
            close_open_modals(driver)
            try:
                btn = driver.find_element(By.CSS_SELECTOR,
                    "#header-action-btn-register, [id*='header-action-btn-register']")
                click(driver, btn, "Зарегистрировать (retry)")
                res_btn = _post_register_check(driver, timeout=5)
            except Exception as e:
                log.error(f"Retry клик упал: {e}")
            if not res_btn:
                log.error(f"Документ {index}/{total}: НЕ зарегистрирован — пропускаю 'На резолюцию'")
                return False, None
        registered = True
    except Exception:
        try:
            btn = driver.find_element(By.XPATH, "//div[contains(text(),'Зарегистрировать')]")
            click(driver, btn, "Зарегистрировать (fallback)")
            res_btn = _post_register_check(driver, timeout=5)
            if res_btn:
                registered = True
        except Exception as e:
            log.error(f"'Зарегистрировать' не найдена: {e}")

    if not registered:
        return False, None

    # res_btn уже получен из _post_register_check выше. Захват номера до 1.5s
    # — если за это время не появился, идём дальше с пустым (записываем в xlsx
    # пусто, но регистрация по факту прошла → flow продолжается).
    asud_id = capture_asud_id(driver, timeout=1.5)
    if asud_id:
        log.info(f"Документ {index}/{total} ЗАРЕГИСТРИРОВАН: {asud_id}")
    else:
        log.warning(f"Документ {index}/{total} ЗАРЕГИСТРИРОВАН (номер не захватили)")

    if not res_btn:
        log.warning("'На резолюцию' не появилась")
        return True, asud_id

    click(driver, res_btn, "На резолюцию")

    # Да — сразу опрашиваем DOM, sleep не нужен (внутри цикла уже есть time.sleep(1) между попытками)
    # По HTML-дампу id кнопки = `confirm-dialog-btn-yes` (дефисы, не подчёркивания).
    yes_btn = None
    for _ in range(10):
        # 1) Точные id (новый с дефисами + старый с подчёркиваниями fallback)
        for sel in ("confirm-dialog-btn-yes", "confirm_dialog_btn_yes"):
            try:
                btn = driver.find_element(By.ID, sel)
                if btn.is_displayed():
                    yes_btn = btn
                    break
            except Exception:
                continue
        if yes_btn:
            break
        # 2) Substring id (GWT может добавлять префиксы/суффиксы)
        try:
            btn = driver.find_element(By.CSS_SELECTOR,
                "[id*='confirm-dialog-btn-yes'], [id*='confirm_dialog_btn_yes'],"
                " [id*='confirm'][id*='yes']")
            if btn.is_displayed():
                yes_btn = btn
                break
        except Exception:
            pass
        # 3) По тексту "Да" в любом видимом элементе
        try:
            for b in driver.find_elements(By.XPATH, "//*[normalize-space(text())='Да']"):
                if b.is_displayed():
                    yes_btn = b
                    break
        except Exception:
            pass
        if yes_btn:
            break
        time.sleep(1)

    if yes_btn:
        # Helper: re-find yes_btn freshly чтобы обойти stale element после
        # перерисовки GXT-диалога между нашими click-стратегиями.
        def _refind_yes():
            for sel in ("confirm-dialog-btn-yes", "confirm_dialog_btn_yes"):
                try:
                    b = driver.find_element(By.ID, sel)
                    if b.is_displayed():
                        return b
                except Exception:
                    continue
            try:
                b = driver.find_element(By.CSS_SELECTOR,
                    "[id*='confirm-dialog-btn-yes'], [id*='confirm_dialog_btn_yes'],"
                    " [id*='confirm'][id*='yes']")
                if b.is_displayed():
                    return b
            except Exception:
                pass
            try:
                for b in driver.find_elements(By.XPATH, "//*[normalize-space(text())='Да']"):
                    if b.is_displayed():
                        return b
            except Exception:
                pass
            return None

        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", yes_btn)
            time.sleep(0.3)
        except Exception:
            pass
        clicked = False
        # 1) ActionChains
        try:
            fresh = _refind_yes() or yes_btn
            ActionChains(driver).move_to_element(fresh).pause(0.3).click().perform()
            log.info(f"Клик 'Да' (ActionChains): id={fresh.get_attribute('id')}")
            clicked = True
        except Exception as e:
            log.debug(f"ActionChains 'Да' не сработал (попробую другие): {type(e).__name__}")
        # 2) JS
        if not clicked:
            try:
                fresh = _refind_yes() or yes_btn
                driver.execute_script("arguments[0].click();", fresh)
                log.info("Клик 'Да' (JS)")
                clicked = True
            except Exception as e:
                log.debug(f"JS 'Да' не сработал: {type(e).__name__}")
        # 3) Нативный
        if not clicked:
            try:
                fresh = _refind_yes() or yes_btn
                fresh.click()
                log.info("Клик 'Да' (native)")
                clicked = True
            except Exception as e:
                log.debug(f"Native 'Да' не сработал: {type(e).__name__}")
        # 4) Enter — GWT-диалоги часто принимают его (отправляем независимо
        # от clicked: если кнопка stale всё равно может закрыть диалог)
        try:
            ActionChains(driver).send_keys(Keys.ENTER).perform()
        except Exception:
            pass
        # Ждём пока модалка закроется (диалог 'Да' уйдёт из DOM).
        # НЕ используем visibility_of(yes_btn) — там тот же stale ref.
        try:
            WebDriverWait(driver, 5).until_not(
                lambda d: bool(_refind_yes()))
        except Exception:
            pass
        log.info(f"Документ {index}/{total} НА РЕЗОЛЮЦИИ")
    else:
        log.warning("Диалог 'Да' не появился за 10 сек")
    return True, asud_id


def register_and_resolve(driver, index, total):
    """Регистрация по подтверждаемым переходам UI.

    Возвращает ``(registered, resolved, asud_id, submission_uncertain)``.
    Последний флаг означает: событие могло уйти в АСУД, но переход интерфейса
    не подтвердился. Такой документ нельзя автоматически регистрировать снова.
    """
    outcome = run_registration(
        driver,
        timeout=max(20, cfg.DEFAULTS["timeout"]),
        retry_grace=2.5,
        capture_id=lambda current_driver: capture_asud_id(
            current_driver, timeout=1.5),
        logger=log,
    )
    if not outcome.registered:
        log.error(
            f"Документ {index}/{total}: регистрация не подтверждена: "
            f"{outcome.reason}"
        )
    elif not outcome.resolved:
        log.error(
            f"Документ {index}/{total}: зарегистрирован, но отправка "
            f"на резолюцию не подтверждена: {outcome.reason}"
        )
    else:
        log.info(f"Документ {index}/{total} НА РЕЗОЛЮЦИИ")
    return (
        outcome.registered,
        outcome.resolved,
        outcome.asud_id,
        outcome.submission_uncertain,
    )


def close_card_and_wait_main(driver):
    """Закрывает карточку и ждёт главную страницу.

    После 'На резолюцию + Да' ASUD сам закрывает карточку — pre-check
    на видимый mainscreen-create-button даёт быстрый выход без лишних
    кликов/ESC.

    Иначе (черновик, ошибка): в DOM может быть несколько элементов с
    id='header-close-btn' (скрытые модалки), берём первый видимый.
    Если ни один не видим — пробуем ESC. Reload — последняя страховка.
    """
    # Быстрый путь: главная уже на экране (типично после 'На резолюцию')
    try:
        main_btn = driver.find_element(By.ID, "mainscreen-create-button")
        if main_btn.is_displayed():
            log.info("Главная уже на экране (карточка закрылась автоматически)")
            return
    except Exception:
        pass

    closed = False
    try:
        candidates = driver.find_elements(By.ID, "header-close-btn")
    except Exception:
        candidates = []

    for btn in candidates:
        try:
            if not btn.is_displayed():
                continue
        except Exception:
            continue
        try:
            ActionChains(driver).move_to_element(btn).pause(0.2).click().perform()
            log.info("Карточка закрыта")
            closed = True
            break
        except Exception:
            try:
                driver.execute_script("arguments[0].click();", btn)
                log.info("Карточка закрыта (JS)")
                closed = True
                break
            except Exception:
                continue

    if not closed:
        # Видимого header-close-btn нет — пробуем ESC, GXT-карточки часто
        # на него реагируют.
        try:
            ActionChains(driver).send_keys(Keys.ESCAPE).perform()
            log.debug("Карточка: попытка закрытия через ESC")
        except Exception:
            pass

    # Ждём главную напрямую — это и есть «карточка ушла + список загружен»
    try:
        WebDriverWait(driver, 8).until(
            EC.element_to_be_clickable((By.ID, "mainscreen-create-button")))
    except Exception:
        log.warning("Главная не загрузилась — перезагружаю")
        driver.get(settings.get("asud_url", cfg.DEFAULTS["asud_url"]))
        wait_asud_loaded(driver)


# ================= DOCUMENT FLOW =================

# Статус последнего create_one_document — читают вызывающие (email-flow и т.п.),
# чтобы решить куда переместить .msg (Завершено / Ошибки / оставить).
# Значения: 'OK' | 'DUPLICATE' | 'DRAFT' | 'REGISTERED_ONLY' |
# 'SUBMISSION_UNKNOWN' | 'FAILED'.
# 'FAILED' — дефолт; перетирается в успешных путях.
_last_result = {"status": "UNKNOWN"}


def _document_addressees(doc_data):
    """Resolve the actual addressee list, honoring a fail-closed doc override."""
    override = doc_data.get("addressees_override")
    if override is not None:
        if isinstance(override, str) or not isinstance(override, (list, tuple)):
            raise RuntimeError("некорректный тематический маршрут адресата")
        raw = override
    else:
        raw = settings.get("addressees", cfg.DEFAULTS["addressees"])
        if isinstance(raw, str):
            raw = [raw]

    resolved = []
    for value in raw or []:
        name = str(value or "").strip()
        if name and name not in resolved:
            resolved.append(name)

    if override is not None and not resolved:
        # A matched GIS rule must never silently fall back to Basmanov.
        raise RuntimeError("тематический маршрут ГИС ЖКХ не содержит адресата")

    doc_data["assigned_addressees"] = resolved
    return resolved


def create_one_document(driver, doc_data, index, total):
    """Создаёт один входящий документ."""
    _last_result["status"] = "FAILED"
    log.info(f"{'='*50}")
    log.info(f"ДОКУМЕНТ {index}/{total}: {doc_data['тема'][:60]}")
    log.info(f"Корреспондент: {doc_data['корреспондент']} "
             f"({'найден ' + (doc_data['корр_источник'] or '') if doc_data['корр_найден'] else 'ЗАГЛУШКА'})")
    log.info(f"Тип: {doc_data['тип_название']}")

    # Business routing is resolved before opening a card. A malformed matched
    # GIS rule must fail before any UI side effect, never fall back to Basmanov.
    try:
        document_addressees = _document_addressees(doc_data)
    except RuntimeError as exc:
        log.error(f"Адресат не определён безопасно: {exc}")
        raise

    # [1/7] Кнопка создания
    el = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
        EC.presence_of_element_located((By.ID, "mainscreen-create-button")))
    click(driver, el, "Создать документ")

    # [2/7] Входящий документ
    wait_and_click(driver, By.XPATH,
        "//div[contains(text(),'Входящий документ')]", "Входящий документ")

    # [3/7] Вид — приоритет по стабильному id (полное название = id).
    # По HTML-дампу: <div id="Письма, заявления и жалобы граждан, акционеров">...</div>
    # Fallback — text-search по короткому префиксу как раньше.
    subtype = doc_data.get("тип_название", "Письма, заявления и жалобы граждан, акционеров")
    type_clicked = False
    try:
        el = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, subtype)))
        click(driver, el, f"тип: {subtype}")
        type_clicked = True
    except Exception:
        log.debug(f"Тип по id '{subtype}' не найден, fallback на text-search")
    if not type_clicked:
        short = subtype[:30]
        wait_and_click(driver, By.XPATH,
            f"//div[contains(text(),'{short}')] | //td[contains(text(),'{short}')]", subtype)

    wait_and_click(driver, By.XPATH,
        "//button[contains(text(),'Создать документ')] | //div[contains(text(),'Создать документ')]",
        "Создать документ")

    # Ждём пока форма документа отрендерится — появится textarea (краткое содержание)
    try:
        WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            lambda d: any(t.is_displayed() for t in d.find_elements(By.TAG_NAME, "textarea")))
    except Exception:
        log.warning("Textarea формы не появилась — продолжаю по таймауту")

    # [4/7] Заполнение формы
    fill_text(driver, doc_data["содержание"])
    correspondent_ready = fill_correspondent_field(
        driver,
        doc_data["корреспондент"],
        kind=doc_data.get("корреспондент_тип", "person"),
    )
    if not correspondent_ready:
        _last_result["status"] = "FAILED"
        log.error("Корреспондент не выбран и не создан — документ остановлен")
        close_open_modals(driver)
        close_card_and_wait_main(driver)
        raise RuntimeError("поле Корреспондент не подтверждено")
    fill_corr_number(driver, doc_data.get("link"),
                      override=doc_data.get("номер_обращения"))
    fill_corr_date(driver, override=doc_data.get("дата_обращения"))

    if doc_data.get("addressees_override") is not None:
        log.info(
            "Тематический маршрут ГИС ЖКХ%s: %s",
            (f" {doc_data.get('zhkh_topic_code')}"
             if doc_data.get("zhkh_topic_code") else ""),
            ", ".join(document_addressees),
        )

    for addr in document_addressees:
        if not add_addressee(driver, addr):
            _last_result["status"] = "FAILED"
            log.error("Адресат не подтверждён — документ остановлен до сохранения")
            close_open_modals(driver)
            close_card_and_wait_main(driver)
            raise RuntimeError("поле Адресат не подтверждено")

    fill_delivery_method(driver)

    # [5/7] Сохранение
    try:
        save_btn = WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.element_to_be_clickable((By.ID, "header-save-btn")))
        click(driver, save_btn, "Сохранить")
        # Quick-check: показывает ли АСУД "уже зарегистрирован" warning?
        # Если да — пропускаем без долгого ожидания register-кнопки.
        time.sleep(1.5)
        if is_duplicate_warning(driver):
            log.warning(f"Документ {index}/{total}: АСУД говорит УЖЕ ЗАРЕГИСТРИРОВАН — пропускаю")
            close_open_modals(driver)
            _last_result["status"] = "DUPLICATE"
            close_card_and_wait_main(driver)
            return None  # caller увидит None и не запишет в output
        # Ждём кнопку 'Зарегистрировать' — признак что save прошёл и форма ушла в режим регистрации
        WebDriverWait(driver, cfg.DEFAULTS["timeout"]).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
                "#header-action-btn-register, [id*='header-action-btn-register']")))
        log.info(f"Документ {index}/{total} сохранён")
    except Exception as e:
        log.error(f"Ошибка сохранения: {e}")
        _last_result["status"] = "FAILED"
        close_open_modals(driver)
        close_card_and_wait_main(driver)
        raise RuntimeError("Сохранение документа не подтверждено") from e

    # [6/7] Прикрепление
    outlook_dir = settings.get("outlook_dir", cfg.DEFAULTS["outlook_dir"])
    dummy_path = doc_data.get("файл")
    attach_path = find_msg_by_link(doc_data.get("link"), outlook_dir, dummy_path)
    if attach_path:
        log.info(f"Прикрепляю: {os.path.basename(attach_path)}")
        attached = attach_content(driver, attach_path)
        wait_modal_closed(driver)
        if attached:
            log.info(f"Документ {index}/{total}: вложение прикреплено ✓")
        else:
            log.warning(f"Документ {index}/{total}: вложение НЕ прикреплено ✗")
            if doc_data.get("require_attachment"):
                _last_result["status"] = "FAILED"
                close_card_and_wait_main(driver)
                raise RuntimeError(
                    f"обязательное вложение не прикрепилось: {attach_path}")
    else:
        log.info("Нет файла — пропускаю")
        if doc_data.get("require_attachment"):
            _last_result["status"] = "FAILED"
            close_card_and_wait_main(driver)
            raise RuntimeError("обязательное вложение не найдено")

    # [7/7] Регистрация (если ФИО найдено) или черновик
    asud_id = None
    if doc_data["корр_найден"]:
        registered, resolved, asud_id, submission_uncertain = (
            register_and_resolve(driver, index, total)
        )
        # После успешной регистрации — реальный (не dummy) .msg → Завершено/
        # Черновики НЕ переносим: файл нужен для ручной доработки.
        if registered and resolved and attach_path and attach_path != dummy_path:
            move_to_done(attach_path, outlook_dir)
        if submission_uncertain:
            _last_result["status"] = "SUBMISSION_UNKNOWN"
        elif registered and resolved:
            _last_result["status"] = "OK"
        elif registered:
            _last_result["status"] = "REGISTERED_ONLY"
        else:
            _last_result["status"] = "FAILED"
        if submission_uncertain:
            close_card_and_wait_main(driver)
            log.error(
                "АСУД мог принять регистрацию, но переход не подтверждён; "
                "автоматически повторять этот документ нельзя"
            )
            return asud_id
        if not registered:
            close_card_and_wait_main(driver)
            raise RuntimeError("Регистрация документа не подтверждена")
        if not resolved:
            close_card_and_wait_main(driver)
            log.error(
                "Документ уже зарегистрирован, но резолюция не подтверждена; "
                "автоматически повторять регистрацию нельзя"
            )
            return asud_id
    else:
        log.warning(f"Row {doc_data['row_idx']}: ФИО НЕ найдено — "
                    f"оставляю в ЧЕРНОВИКАХ для ручной доработки "
                    f"(тема: {doc_data['тема'][:60]}). "
                    f"Файл НЕ перемещаю — лежит на месте.")
        _last_result["status"] = "DRAFT"

    close_card_and_wait_main(driver)
    return asud_id


# ================= MAIN =================

settings = {}


def main():
    global settings
    settings = cfg.load()
    cfg.setup_file_logger("mix")
    cfg.keep_system_awake(True)

    log.info("=" * 50)
    log.info("АСУД ИК — MIX (auto-create + smart-routing)")
    log.info("=" * 50)

    base_dir = cfg.get_base_dir()

    # Если xlsx уже выбран извне (через app.py) — используем его
    excel_path = os.environ.get('ASUD_XLSX')
    xlsx_files = [] if excel_path else \
        [f for f in os.listdir(base_dir) if f.lower().endswith('.xlsx')]

    if excel_path:
        log.info(f"Файл (через app): {os.path.basename(excel_path)}")
    elif not xlsx_files:
        log.error(f"Нет .xlsx в {base_dir}")
        input("Enter...")
        sys.exit(1)
    elif len(xlsx_files) == 1:
        excel_path = os.path.join(base_dir, xlsx_files[0])
        log.info(f"Файл: {xlsx_files[0]}")
    else:
        print(f"\nНайдено {len(xlsx_files)} xlsx-файлов:")
        for i, f in enumerate(xlsx_files, 1):
            print(f"  {i}. {f}")
        choice = input("Выбери номер: ").strip()
        try:
            excel_path = os.path.join(base_dir, xlsx_files[int(choice) - 1])
        except (ValueError, IndexError):
            log.error("Неверный выбор")
            sys.exit(1)

    # Папка с .msg — приоритет:
    #   1) ASUD_OUTLOOK_DIR env (от пресета в settings.json)
    #   2) settings.outlook_dir / DEFAULTS — default для prompt'а
    #   3) интерактивный ввод
    env_outlook = os.environ.get('ASUD_OUTLOOK_DIR', '').strip()
    if env_outlook:
        settings["outlook_dir"] = env_outlook
        log.info(f"Папка вложений из пресета: {env_outlook}")
    else:
        default_outlook = settings.get("outlook_dir", "") or cfg.DEFAULTS.get("outlook_dir", "")
        print(f"\nПапка с .msg-файлами (поиск рекурсивно по подпапкам).")
        if default_outlook:
            print(f"Enter — использовать: {default_outlook}")
        else:
            print(f"Enter — без поиска .msg (все вложения как пустышка)")
        user_dir = input("Путь: ").strip().strip('"').strip("'")
        if user_dir:
            settings["outlook_dir"] = user_dir
    outlook_dir = settings.get("outlook_dir") or ""
    if outlook_dir and not os.path.isdir(outlook_dir):
        log.warning(f"Папка '{outlook_dir}' не существует — "
                    f"все вложения уйдут как пустышки")
    elif outlook_dir:
        log.info(f"Папка вложений: {outlook_dir}")

    # Пустышка (для случаев когда .msg по link не найден)
    msg_path = get_dummy_msg(base_dir)
    if msg_path:
        log.info(f"Пустышка: {os.path.basename(msg_path)}")

    # Данные
    docs = load_excel(excel_path)
    for doc in docs:
        doc["файл"] = msg_path

    if not docs:
        log.error("Нет данных!")
        input("Enter...")
        sys.exit(1)

    # Resume: проверяем state-файл
    processed = load_state(excel_path)
    if processed:
        done_in_current = [d for d in docs
                           if _link_key(d.get("link")) in processed]
        if done_in_current:
            print(f"\nВ state-файле {len(done_in_current)} ранее обработанных документов.")
            print("  Enter / 'да'  — ПРОПУСТИТЬ их, продолжить с остальных (по умолчанию)")
            print("  'нет'         — обработать ВСЁ заново (дубли в АСУД! не рекомендуется)")
            print("  'сброс'       — обнулить state и обработать всё (для полного старта)")
            ans = input("Что делаем? [да]: ").strip().lower()
            if ans in ("сброс", "reset"):
                save_state(excel_path, set())
                processed = set()
                log.info("State обнулён — обрабатываю всё заново")
            elif ans in ("нет", "н", "n", "no"):
                log.info("Обрабатываю всё заново (state будет дополнен)")
            else:
                before = len(docs)
                docs = [d for d in docs
                        if _link_key(d.get("link")) not in processed]
                log.info(f"Пропускаю {before - len(docs)} обработанных, "
                         f"осталось {len(docs)}")
                if not docs:
                    log.info("Все строки уже обработаны — нечего делать.")
                    input("Enter...")
                    sys.exit(0)

    known = sum(1 for d in docs if d["корр_найден"])
    unknown = len(docs) - known
    print(f"\nПервые 5:")
    for i, d in enumerate(docs[:5], 1):
        flag = 'OK' if d["корр_найден"] else '!!'
        print(f"  {i}. [{d['тип_индекс']}] {flag} {d['корреспондент'][:30]} | {d['тема'][:50]}")
    mode_label = os.environ.get('ASUD_MODE', 'mix')
    print(f"\nВсего к обработке: {len(docs)}  (ФИО: {known}, заглушка: {unknown})")
    print(f"режим: {mode_label.upper()}  —  создание + регистрация + На резолюцию + .msg по Link")

    confirm = input("Начать? (да/нет): ").strip().lower()
    if confirm not in ("да", "д", "y", "yes", ""):
        print("Отменено.")
        sys.exit(0)

    # Браузер
    driver_path = os.path.join(base_dir, "msedgedriver.exe")
    if not os.path.exists(driver_path):
        log.error(f"msedgedriver.exe не найден в {base_dir}")
        input("Enter...")
        sys.exit(1)

    options = cfg.build_edge_options()

    service = EdgeService(executable_path=driver_path)
    driver = webdriver.Edge(service=service, options=options)
    set_driver_timeout(driver, settings.get("asud_load_timeout_sec",
                                              cfg.DEFAULTS["asud_load_timeout_sec"]))

    try:
        url = settings.get("asud_url", cfg.DEFAULTS["asud_url"])
        log.info(f"Открываю {url}")
        driver.get(url)
        wait_asud_loaded(driver)

        # Output xlsx для clean-resolutions: ОПТС | Округ | ФИО | Link
        output_path = _output_xlsx_path(excel_path)
        _ensure_output_xlsx(output_path)
        log.info(f"Output xlsx: {output_path}")

        done_count, err_count = 0, 0
        for i, doc in enumerate(docs, 1):
            try:
                asud_id = create_one_document(driver, doc, i, len(docs))
                flow_status = _last_result.get("status", "FAILED")
                if flow_status in {"REGISTERED_ONLY", "SUBMISSION_UNKNOWN"}:
                    # Терминальный ручной исход: ссылка фиксируется, чтобы
                    # повторный запуск не создавал дубль документа.
                    key = _link_key(doc.get("link"))
                    if key:
                        processed.add(key)
                        save_state(excel_path, processed)
                    _append_output_row(
                        output_path, doc, asud_id,
                        status=flow_status,
                    )
                    log.error(
                        f"Документ {i}: {flow_status} — требуется ручная "
                        "проверка; автоматический повтор запрещён"
                    )
                    err_count += 1
                    continue
                # Помечаем как обработанный сразу после успеха (до следующей итерации)
                key = _link_key(doc.get("link"))
                if key:
                    processed.add(key)
                    save_state(excel_path, processed)
                # Запись в output xlsx (даже если asud_id не захватили — ставим Link)
                _append_output_row(output_path, doc, asud_id, status="OK")
                done_count += 1
            except Exception as e:
                log.error(f"ОШИБКА документ {i}: {e}")
                err_count += 1
                driver.get(url)
                wait_asud_loaded(driver)
                continue

        elapsed_seconds = time.monotonic() - start_time
        elapsed = timedelta(seconds=int(elapsed_seconds))
        avg = timedelta(seconds=int(elapsed_seconds / done_count)) if done_count else None

        summary = [
            "",
            "=" * 60,
            f"ГОТОВО!",
            f"  Обработано:   {done_count} / {len(docs)}",
            f"  Ошибок:       {err_count}",
            f"  В черновиках: {unknown}",
            f"  Затрачено:    {elapsed}" + (f"  (в среднем {avg}/док)" if avg else ""),
            "=" * 60,
        ]
        for line in summary:
            log.info(line)
            print(line)

        if unknown:
            log.warning(f"Проверьте {unknown} документов в черновиках — "
                        f"ФИО не извлечено автоматически")
        if err_count:
            log.warning(f"{err_count} документов упали с ошибкой. "
                        f"Перезапуск скрипта продолжит с них "
                        f"(уже обработанные запомнены в state-файле).")
        input("\nEnter для закрытия...")

    except Exception as e:
        log.error(f"Ошибка: {e}")
        input("Enter...")
    finally:
        try:
            driver.quit()
            log.info("Браузер закрыт")
        except Exception:
            pass
        cfg.keep_system_awake(False)


if __name__ == "__main__":
    main()
