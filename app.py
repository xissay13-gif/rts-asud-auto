"""
app.py — Единая точка входа для АСУД-автоматизации создания документов.

Поддерживаемые режимы:
  • mix         — создание + регистрация + На резолюцию + .msg по Link
  • auto-create — создание + регистрация + На резолюцию (без .msg по Link)
  • smart       — создание ТОЛЬКО как черновик + .msg по Link (без регистрации,
                  корреспондент = «Неизвестный Неизвестный Неизвестный»)
  • email       — создание прямо из .msg-писем (без xlsx-реестра): рекурсивно
                  обходит папку, парсит письма, ФИО абонента берёт из тела.
  • sbis        — регистрация PDF из СБИС: ФЛ по полному ФИО, ЮЛ по полному
                  названию с прочерками в имени и отчестве.

Выдача резолюций — отдельный exe, ветка clean-resolutions.

Запуск:
  python app.py                       # auto-detect режима по xlsx
  python app.py --mode=mix
  python app.py --mode=auto-create
  python app.py --mode=smart
  python app.py --mode=email
  python app.py --mode=sbis --folder=path
  python app.py --xlsx=path.xlsx --mode=...
  python app.py --folder=path         # → авто-режим email
  python app.py --headless            # фоновый режим (Edge без GUI)

Auto-detect:
  • Передан --mode=sbis                                      → sbis
  • Передан --folder или --mode=email                         → email
  • Лист содержит колонку 'Link' (старый mix-формат)          → mix
  • Лист 'результат' (новый формат) или Subject/корреспондент → auto-create
  Smart НЕ определяется автоматически (тот же xlsx что у mix), нужен --mode=smart
"""

import argparse
import logging
import os
import sys

import openpyxl

from shared import config as cfg

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S',
    handlers=[logging.StreamHandler()],
)
log = logging.getLogger("asud.app")


_MODE_DESCRIPTIONS = {
    'mix':         'Создание + регистрация + На резолюцию + .msg по Link',
    'auto-create': 'Создание + регистрация + На резолюцию (без .msg)',
    'smart':       'Создание как черновик + .msg (без регистрации, фикс. корреспондент)',
    'email':       'Создание прямо из .msg-писем (без xlsx-реестра)',
    'sbis':        'Регистрация файлов из выгрузки СБИС',
    'zhkh-daemon': 'Постоянный мониторинг реестров → Басманов выдаёт резолюции Халецкой',
}
_MODES = ['mix', 'auto-create', 'smart', 'email', 'sbis', 'zhkh-daemon']


def detect_mode(xlsx_path):
    """Определяет режим по структуре xlsx (mix или auto-create)."""
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        # Лист 'результат' — новый auto-create формат
        if 'результат' in wb.sheetnames:
            wb.close()
            return 'auto-create'
        ws = wb.active
        headers = [str(c.value or '').lower()
                   for c in next(ws.iter_rows(max_row=1))]
        wb.close()
        # Колонка 'link' → mix-flow (есть .msg по ссылке)
        if any('link' in h for h in headers):
            return 'mix'
    except Exception as e:
        log.warning(f"Не удалось прочитать {xlsx_path} для auto-detect: {e}")

    return 'auto-create'


def pick_mode(xlsx_path):
    """Интерактивный выбор режима с подсказкой auto-detect.
    Email-режим в этом меню не показываем — он работает с папкой, не с xlsx."""
    xlsx_modes = ['mix', 'auto-create', 'smart']
    auto = detect_mode(xlsx_path)
    print(f"\nРеестр: {os.path.basename(xlsx_path)}")
    print("Какой процесс запустить?")
    for i, m in enumerate(xlsx_modes, 1):
        marker = '  ← рекомендую (auto-detect)' if m == auto else ''
        print(f"  {i}. {m:11} — {_MODE_DESCRIPTIONS[m]}{marker}")
    print(f"\n[Enter] = {auto}  (рекомендуется по реестру)")
    choice = input(f"Номер режима (1-{len(xlsx_modes)}) или Enter: ").strip()
    if not choice:
        return auto
    try:
        return xlsx_modes[int(choice) - 1]
    except (ValueError, IndexError):
        log.warning(f"Неверный выбор '{choice}' — использую {auto}")
        return auto


def pick_source():
    """Интерактивный верхне-уровневый выбор источника.
    Вызывается когда юзер не передал ни --mode, ни --xlsx, ни --folder."""
    print("\nЧто обработать?")
    print("  1. Реестр (.xlsx)        — режимы mix / auto-create / smart")
    print("  2. Папку с .msg-письмами — режим email")
    print("  3. Файлы из выгрузки СБИС — режим sbis")
    choice = input("Номер (1-3) [1]: ").strip()
    if choice == '2':
        return 'email'
    if choice == '3':
        return 'sbis'
    return 'xlsx'


def pick_preset(presets):
    """Меню пресетов сценариев. Возвращает выбранный preset dict или None.
    Если пользователь нажал Enter — берётся первый."""
    print("\nВыбери сценарий:")
    for i, p in enumerate(presets, 1):
        name = p.get("name", "?")
        folder = p.get("folder", "")
        print(f"  {i}. {name}  ({folder})")
    print(f"[Enter] = 1")
    choice = input(f"Номер (1-{len(presets)}) или Enter: ").strip()
    if not choice:
        return presets[0]
    try:
        return presets[int(choice) - 1]
    except (ValueError, IndexError):
        log.warning(f"Неверный выбор '{choice}'")
        return None


def find_preset(presets, selector):
    """Find a preset by stable id or exact display name."""
    needle = str(selector or "").strip().casefold()
    if not needle:
        return None
    for preset in presets:
        if not isinstance(preset, dict):
            continue
        candidates = (preset.get("id", ""), preset.get("name", ""))
        if any(str(value).strip().casefold() == needle for value in candidates):
            return preset
    return None

def main():
    parser = argparse.ArgumentParser(
        description="АСУД ИК — автоматизация документооборота")
    parser.add_argument('--mode', choices=_MODES,
                        help="Режим работы (если не задан — auto-detect по xlsx)")
    parser.add_argument('--xlsx', help="Путь к реестру (если не задан — спрашиваем)")
    parser.add_argument('--folder', help="Папка-источник для режимов email/sbis")
    parser.add_argument('--preset',
                        help="Preset id or exact name from settings.json")
    parser.add_argument('--surname',
                        help="Фамилия корреспондента для всей партии в режиме sbis; "
                             "если не задана — берётся из имени файла")
    parser.add_argument('--reset-state', action='store_true',
                        help="Сбросить нумерацию и список обработанных файлов sbis")
    parser.add_argument('--watch', action='store_true',
                        help="Непрерывный мониторинг папки (только email-режим). "
                             "Ctrl+C — остановка после текущего документа")
    parser.add_argument('--headless', action='store_true',
                        help="Запустить Edge без GUI (фоновый режим, требует Стадии 1б)")
    args = parser.parse_args()

    base_dir = cfg.get_base_dir()
    log.info(f"Базовая папка: {base_dir}")
    if args.headless:
        os.environ['ASUD_HEADLESS'] = '1'
        log.info("Режим: HEADLESS (Edge без GUI)")

    # === Пресеты сценариев (из settings.json) ================================
    # Если флагов нет И в settings.json есть presets — показываем меню.
    # Иначе fallback на старое поведение (интерактивный выбор источника).
    settings_data = cfg.load()
    presets = settings_data.get("presets") or []
    # Глобальный flag delete_after_done (если есть в settings.json) тоже подхватываем
    if settings_data.get("delete_after_done"):
        os.environ.setdefault('ASUD_DELETE_AFTER_DONE', '1')
    no_flags = not (args.mode or args.xlsx or args.folder or args.preset)
    preset = None
    if args.preset:
        preset = find_preset(presets, args.preset)
        if preset is None and args.preset.casefold() == "sbis":
            legacy_folder = str(
                (settings_data.get("sbis") or {}).get("input_dir") or ""
            ).strip()
            if legacy_folder:
                preset = {
                    "id": "sbis",
                    "name": "SBIS",
                    "mode": "sbis",
                    "folder": legacy_folder,
                }
                log.warning(
                    "Legacy sbis.input_dir is in use; move the path "
                    "to presets[].folder")
        if preset is None:
            log.error(f"Preset '{args.preset}' was not found in settings.json")
            sys.exit(1)
    elif no_flags and presets:
        preset = pick_preset(presets)
        if preset is None:
            log.error("Пресет не выбран — выход")
            sys.exit(1)

    if preset is not None:
        log.info(f"Пресет: {preset.get('name', '?')}")
        # Заполняем args из пресета и продолжаем обычный flow
        if preset.get("folder"):
            args.folder = preset["folder"]
        elif preset.get("folders"):
            # Multi-folder preset: пробрасываем через env как JSON.
            # daemon_main распарсит и будет ходить по списку (опционально
            # с round-robin).
            import json as _json
            os.environ['ASUD_EMAIL_FOLDERS_JSON'] = _json.dumps(
                preset["folders"], ensure_ascii=False)
            # Принудительно включаем watch+email-режим — multi-folder работает
            # только в daemon-loop'е.
            args.watch = True
            args.folder = preset["folders"][0].get("dir") if isinstance(
                preset["folders"][0], dict) else preset["folders"][0]
            # ASUD_EMAIL_ROUND_ROBIN — daemon_main прочитает
            if preset.get("round_robin"):
                os.environ['ASUD_EMAIL_ROUND_ROBIN'] = '1'
        if preset.get("mode"):
            args.mode = preset["mode"]
        # xlsx из пресета — авто-подстановка реестра для mix/auto-create/smart
        if preset.get("xlsx") and not args.xlsx:
            args.xlsx = preset["xlsx"]
        # outlook_dir из пресета — папка с .msg для mix-вложений (по Link).
        # Пробрасываем через env, mix.main() прочитает.
        if preset.get("outlook_dir"):
            os.environ['ASUD_OUTLOOK_DIR'] = preset["outlook_dir"]
        # ASUD_EMAIL_PROCESS_MODE — для email-flow: 'mix' (текущая логика) или
        # 'smart' (всегда черновик + фикс. корреспондент).
        os.environ['ASUD_EMAIL_PROCESS_MODE'] = preset.get("mode", "mix")
        # ASUD_OUTPUT_SUFFIX — суффикс в имени per-date xlsx (для разделения
        # реестров при параллельных запусках двух .bat).
        if preset.get("output_suffix"):
            os.environ['ASUD_OUTPUT_SUFFIX'] = preset["output_suffix"]
        # delete_after_done — удалять обработанные .msg вместо переноса в
        # Завершено/. Удобно для непрерывного daemon-режима.
        if preset.get("delete_after_done"):
            os.environ['ASUD_DELETE_AFTER_DONE'] = '1'

    # === ZHKH-DAEMON (непрерывный второй проход) ============================
    # Этот режим не использует обычный источник (xlsx/email/sbis), поэтому
    # запускаем его до интерактивного меню выбора источника.
    if args.mode == 'zhkh-daemon':
        log.info("Режим: zhkh-daemon (мониторинг реестров под Басмановым)")
        os.environ['ASUD_MODE'] = 'zhkh-daemon'
        from flows.zhkh_daemon import main as flow_main
        flow_main()
        return

    # === Определяем источник: sbis vs email vs xlsx ==========================
    if args.mode == 'sbis':
        source = 'sbis'
    elif args.folder or args.mode == 'email':
        source = 'email'
    elif args.xlsx or args.mode in ('mix', 'auto-create', 'smart'):
        # Если пришёл флаг 'smart' или 'mix' С --folder — это email-источник
        # (пресет проставил оба), переопределяем
        if args.folder:
            source = 'email'
        else:
            source = 'xlsx'
    else:
        # Ничего не указано — спрашиваем
        source = pick_source()

    # === СБИС-источник ======================================================
    if source == 'sbis':
        if args.folder:
            os.environ['ASUD_SBIS_DIR'] = args.folder
            log.info(f"Режим: sbis (папка: {args.folder})")
        else:
            log.info("Режим: sbis")
        if args.surname:
            os.environ['ASUD_SBIS_SURNAME'] = args.surname
        if args.reset_state:
            os.environ['ASUD_SBIS_RESET_STATE'] = '1'
        os.environ['ASUD_MODE'] = 'sbis'
        from flows.sbis import main as flow_main
        flow_main()
        return

    # === EMAIL-источник =====================================================
    if source == 'email':
        if args.folder:
            os.environ['ASUD_EMAIL_FOLDER'] = args.folder
            log.info(f"Режим: email (папка: {args.folder})")
        else:
            log.info("Режим: email")
        os.environ['ASUD_MODE'] = 'email'
        if args.watch:
            os.environ['ASUD_WATCH'] = '1'
            log.info("Режим: WATCH (непрерывный мониторинг, Ctrl+C для остановки)")
            from flows.email import daemon_main as flow_main
        else:
            from flows.email import main as flow_main
        flow_main()
        return

    # === XLSX-источник: mix / auto-create / smart ===========================
    # Если xlsx не указан — выбираем интерактивно
    xlsx_path = args.xlsx
    if not xlsx_path:
        # Сначала пробуем выбрать из base_dir
        candidates = [f for f in os.listdir(base_dir)
                      if f.lower().endswith('.xlsx') and not f.startswith('~')]
        if len(candidates) == 1:
            xlsx_path = os.path.join(base_dir, candidates[0])
            log.info(f"Найден единственный xlsx: {candidates[0]}")
        elif len(candidates) > 1:
            print("\nДоступные реестры:")
            for i, name in enumerate(candidates, 1):
                print(f"  {i}. {name}")
            choice = input(f"Выбери (1-{len(candidates)}): ").strip()
            try:
                xlsx_path = os.path.join(base_dir, candidates[int(choice) - 1])
            except (ValueError, IndexError):
                log.error("Неверный выбор")
                sys.exit(1)
        else:
            log.error(f"Не нашёл .xlsx в {base_dir}")
            sys.exit(1)

    # Определяем режим: явный флаг → интерактивный выбор → auto-detect
    if args.mode:
        mode = args.mode
        log.info(f"Режим: {mode} (через --mode)")
    else:
        mode = pick_mode(xlsx_path)
        log.info(f"Режим выбран: {mode}")

    # Запуск соответствующего flow
    if mode == 'mix':
        from flows.mix import main as flow_main
    elif mode == 'auto-create':
        from flows.auto_create import main as flow_main
    elif mode == 'smart':
        from flows.smart import main as flow_main
    else:
        log.error(f"Неизвестный режим: {mode}")
        sys.exit(1)

    # Каждый flow.main() сам читает xlsx — передаём через env-переменную
    # (минимальный contract без переписывания main каждого flow).
    # ASUD_MODE передаём для отображения в превью flow'а.
    os.environ['ASUD_XLSX'] = xlsx_path
    os.environ['ASUD_MODE'] = mode
    flow_main()


if __name__ == "__main__":
    main()
