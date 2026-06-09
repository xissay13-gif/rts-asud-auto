"""
shared/xlsx_status.py — Хелперы для отметки статусов в реестре резолюций.

Реестры (Registered/*_резолюции.xlsx) содержат колонки:
  "Отписано Халецкой"  — заполняется когда Басманов выдал резолюцию Халецкой
                          (второй проход ГИСЖКХ)
  "Отписано в округ"   — заполняется когда Халецкая выдала резолюцию окружному
                          начальнику (clean-resolutions, основной сценарий)

Формат значения: дата DD.MM.YYYY (когда отписали).

Используется и для пометки (после успешной обработки), и для skip-already-done
(чтобы recovery после crash/сна не дублировал отписки).
"""

import os
import logging
from datetime import date

import openpyxl

from shared.xlsx_lock import xlsx_lock

log = logging.getLogger("asud")

COL_HALETSKAYA = "Отписано Халецкой"
COL_OKRUG     = "Отписано в округ"
STATUS_COLUMNS = (COL_HALETSKAYA, COL_OKRUG)


def _find_asud_col(headers_lower):
    """Возвращает 0-based индекс колонки с asud_id (Номер/ОПТС). None если нет."""
    keys = ('номер', 'опт', 'орт', 'асуд', 'asud', 'регистрацион')
    for i, h in enumerate(headers_lower):
        for k in keys:
            if k in h:
                return i
    return None


def get_done_asud_ids(xlsx_path, column_name):
    """Возвращает set asud_id у которых колонка column_name непустая.
    Используется для skip-already-done в начале прохода.
    """
    done = set()
    if not xlsx_path or not os.path.isfile(xlsx_path):
        return done
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return done
        headers = [str(c or '').strip() for c in rows[0]]
        headers_lower = [h.lower() for h in headers]
        asud_col = _find_asud_col(headers_lower)
        try:
            status_col = headers.index(column_name)
        except ValueError:
            wb.close()
            return done
        if asud_col is None:
            wb.close()
            return done
        for r in rows[1:]:
            if len(r) <= max(asud_col, status_col):
                continue
            aid = str(r[asud_col] or '').strip()
            val = str(r[status_col] or '').strip() if r[status_col] is not None else ''
            if aid and val:
                done.add(aid)
        wb.close()
    except Exception as e:
        log.debug(f"get_done_asud_ids({xlsx_path!r}, {column_name!r}): {e}")
    return done


def mark_status(xlsx_path, asud_id, column_name, value=None):
    """Помечает статус для строки с заданным asud_id в xlsx.

    Поиск строки — по колонке Номер/ОПТС (а не по 1-й, т.к. она может
    отличаться у разных форматов реестра).
    Если column_name отсутствует в шапке — добавляет.
    value=None → подставляется сегодняшняя дата DD.MM.YYYY.

    Возвращает True если запись сохранилась.
    """
    if not xlsx_path or not os.path.isfile(xlsx_path):
        return False
    if value is None:
        value = date.today().strftime("%d.%m.%Y")
    try:
        with xlsx_lock(xlsx_path):
            return _mark_status_locked(xlsx_path, asud_id, column_name, value)
    except TimeoutError as e:
        log.warning(f"mark_status({asud_id}, {column_name}): {e}")
        return False
    except PermissionError as e:
        log.warning(f"mark_status({asud_id}, {column_name}): xlsx занят "
                    f"(возможно, открыт в Excel): {e}")
        return False
    except Exception as e:
        log.warning(f"mark_status({asud_id}, {column_name}): {e}")
        return False


def _mark_status_locked(xlsx_path, asud_id, column_name, value):
    """Внутренняя часть mark_status — выполняется уже под lock'ом."""
    try:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(max_row=1))]
        headers_lower = [h.lower() for h in headers]
        asud_col = _find_asud_col(headers_lower)
        if asud_col is None:
            log.warning(f"mark_status: в {xlsx_path} нет колонки Номер/ОПТС")
            wb.close()
            return False
        # Колонка статуса — есть в шапке? иначе добавляем
        try:
            status_col = headers.index(column_name)
        except ValueError:
            status_col = len(headers)
            ws.cell(row=1, column=status_col + 1).value = column_name
            ws.cell(row=1, column=status_col + 1).font = openpyxl.styles.Font(bold=True)
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(status_col + 1)].width = 18
        # Найти строку с этим asud_id
        target_row = None
        for r in ws.iter_rows(min_row=2):
            if asud_col >= len(r):
                continue
            if str(r[asud_col].value or '').strip() == asud_id:
                target_row = r[0].row
                break
        if target_row is None:
            log.debug(f"mark_status: asud_id {asud_id!r} не найден в {xlsx_path}")
            wb.close()
            return False
        ws.cell(row=target_row, column=status_col + 1).value = value
        wb.save(xlsx_path)
        wb.close()
        return True
    except Exception:
        # caller (mark_status) ловит и логирует
        raise
