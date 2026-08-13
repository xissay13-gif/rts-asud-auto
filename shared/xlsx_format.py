"""Единое оформление накопительных XLSX-реестров АСУД.

OpenPyXL не поддерживает настоящий Excel AutoFit. Поэтому ширина колонок
оценивается по тексту: короткие значения расширяют колонку, а длинные
ограничиваются максимальной шириной и переносятся внутри ячейки.
"""

import os
import unicodedata
from copy import copy

from openpyxl.utils import get_column_letter


MIN_COLUMN_WIDTH = 8
MAX_COLUMN_WIDTH = 60
HEADER_ROW_HEIGHT = 30

# В течение одного запуска полностью форматируем каждый реестр один раз.
# При последующих append достаточно оформить только новую строку.
_formatted_registry_paths = set()


def _path_key(path):
    if not path:
        return None
    return os.path.normcase(os.path.abspath(os.fspath(path)))


def _display_width(value):
    """Приблизительная ширина самой длинной строки значения в символах."""
    if value is None:
        return 0
    text = str(value).replace("\t", "    ")
    lines = text.splitlines() or [text]

    def line_width(line):
        return sum(
            2 if unicodedata.east_asian_width(char) in ("W", "F") else 1
            for char in line
        )

    return max((line_width(line) for line in lines), default=0)


def _wrap_cell(cell, header=False):
    if cell.value is None:
        return
    alignment = copy(cell.alignment)
    alignment.wrap_text = True
    alignment.vertical = "center" if header else "top"
    if alignment.horizontal is None:
        alignment.horizontal = "left"
    cell.alignment = alignment


def _bounded_width(current, desired, min_width, max_width):
    try:
        current = float(current or 0)
    except (TypeError, ValueError):
        current = 0
    return min(max_width, max(min_width, current, desired))


def format_registry_worksheet(
        ws, min_width=MIN_COLUMN_WIDTH, max_width=MAX_COLUMN_WIDTH,
        header_row=1):
    """Форматирует всю заполненную область листа и подбирает ширину колонок."""
    desired_widths = {}
    for row in ws.iter_rows(
            min_row=1, max_row=max(1, ws.max_row),
            min_col=1, max_col=max(1, ws.max_column)):
        for cell in row:
            if cell.value is None:
                continue
            _wrap_cell(cell, header=(cell.row == header_row))
            desired_widths[cell.column] = max(
                desired_widths.get(cell.column, 0),
                _display_width(cell.value) + 2,
            )

    for col_idx in range(1, max(1, ws.max_column) + 1):
        letter = get_column_letter(col_idx)
        dimension = ws.column_dimensions[letter]
        dimension.width = _bounded_width(
            dimension.width,
            desired_widths.get(col_idx, min_width),
            min_width,
            max_width,
        )

    if 1 <= header_row <= ws.max_row:
        current_height = ws.row_dimensions[header_row].height or 0
        ws.row_dimensions[header_row].height = max(
            current_height, HEADER_ROW_HEIGHT)


def format_registry_row(
        ws, row_idx, min_width=MIN_COLUMN_WIDTH,
        max_width=MAX_COLUMN_WIDTH, header_row=1):
    """Оформляет добавленную/изменённую строку и при необходимости расширяет колонки."""
    if not row_idx or row_idx < 1:
        return
    for col_idx in range(1, max(1, ws.max_column) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is None:
            continue
        _wrap_cell(cell, header=(row_idx == header_row))
        letter = get_column_letter(col_idx)
        dimension = ws.column_dimensions[letter]
        dimension.width = _bounded_width(
            dimension.width,
            _display_width(cell.value) + 2,
            min_width,
            max_width,
        )


def format_registry_before_save(ws, path=None, changed_row=None):
    """Форматирует реестр перед save без повторного обхода всей истории.

    Первый save данного файла за запуск форматирует весь лист. Последующие
    операции форматируют только добавленную или изменённую строку.
    """
    key = _path_key(path)
    if key is None or key not in _formatted_registry_paths:
        format_registry_worksheet(ws)
        if key is not None:
            _formatted_registry_paths.add(key)
        return
    if changed_row is not None:
        format_registry_row(ws, changed_row)
