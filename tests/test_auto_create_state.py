import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import openpyxl

from flows import auto_create


def _outcome(*, registered=False, resolved=False, uncertain=False,
             asud_id=None):
    return SimpleNamespace(
        registered=registered,
        resolved=resolved,
        submission_uncertain=uncertain,
        asud_id=asud_id,
        reason="test",
    )


class AutoCreateStateTests(unittest.TestCase):
    def _workbook(self, root):
        path = Path(root) / "input.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "результат"
        ws.append(["№", "обращение", "фио"])
        ws.append([1, "Первое обращение", "Иванов Иван Иванович"])
        ws.append([2, None, None])
        ws.append([3, "Второе обращение", "Петров Пётр Петрович"])
        wb.save(path)
        wb.close()
        return path

    def test_load_excel_adds_content_key_and_source_row_metadata(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(tmp)

            docs = auto_create.load_excel(source)

        self.assertEqual([doc["_source_row"] for doc in docs], [2, 4])
        self.assertTrue(all(
            doc["_source_key"].startswith("row-sha256:") for doc in docs
        ))
        self.assertNotEqual(docs[0]["_source_key"], docs[1]["_source_key"])

    def test_inserting_empty_row_keeps_source_keys_stable(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(tmp)
            before = {
                doc["содержание"]: doc["_source_key"]
                for doc in auto_create.load_excel(source)
            }

            wb = openpyxl.load_workbook(source)
            wb["результат"].insert_rows(2)
            wb.save(source)
            wb.close()

            after_docs = auto_create.load_excel(source)
            after = {
                doc["содержание"]: doc["_source_key"] for doc in after_docs
            }

        self.assertEqual(after, before)
        self.assertEqual([doc["_source_row"] for doc in after_docs], [3, 5])

    def test_changing_row_content_changes_source_key(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(tmp)
            before = auto_create.load_excel(source)[0]["_source_key"]

            wb = openpyxl.load_workbook(source)
            wb["результат"]["B2"] = "Первое обращение — изменено"
            wb.save(source)
            wb.close()

            after = auto_create.load_excel(source)[0]["_source_key"]

        self.assertNotEqual(after, before)

    def test_identical_rows_receive_distinct_ordinals(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(tmp)
            wb = openpyxl.load_workbook(source)
            ws = wb["результат"]
            ws.append([1, "Первое обращение", "Иванов Иван Иванович"])
            wb.save(source)
            wb.close()

            matches = [
                doc for doc in auto_create.load_excel(source)
                if doc["содержание"] == "Первое обращение"
            ]

        self.assertEqual(len(matches), 2)
        first_base, first_ordinal = matches[0]["_source_key"].rsplit(":", 1)
        second_base, second_ordinal = matches[1]["_source_key"].rsplit(":", 1)
        self.assertEqual(first_base, second_base)
        self.assertEqual((first_ordinal, second_ordinal), ("1", "2"))

    def test_pending_sidecar_is_atomic_and_terminal_after_restart(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(tmp)
            doc = auto_create.load_excel(source)[0]
            state = auto_create.load_registration_state(source)

            auto_create._record_document_state(
                source, state, doc, "SUBMISSION_PENDING")

            state_path = auto_create._registration_state_path(source)
            self.assertEqual(state_path.parent, source.parent)
            self.assertTrue(state_path.is_file())
            self.assertFalse(state_path.with_name(state_path.name + ".tmp").exists())

            restarted = auto_create.load_registration_state(source)
            self.assertTrue(auto_create._document_is_terminal(restarted, doc))
            self.assertEqual(
                auto_create._document_state(restarted, doc)["status"],
                "SUBMISSION_PENDING",
            )

    def test_every_protected_status_is_skipped_after_restart(self):
        doc = {"_source_key": "лист!7", "_source_row": 7}
        for status in auto_create.TERMINAL_REGISTRATION_STATUSES:
            with self.subTest(status=status):
                state = {
                    "version": 1,
                    "documents": {"лист!7": {"status": status}},
                }
                self.assertTrue(auto_create._document_is_terminal(state, doc))

    def test_corrupt_sidecar_fails_closed(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(tmp)
            state_path = auto_create._registration_state_path(source)
            state_path.write_text("{broken", encoding="utf-8")

            with self.assertRaisesRegex(RuntimeError, "Автоповтор заблокирован"):
                auto_create.load_registration_state(source)

    def test_write_ahead_is_saved_before_registration_state_machine_runs(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(tmp)
            doc = auto_create.load_excel(source)[0]
            state = auto_create.load_registration_state(source)
            expected = _outcome(
                registered=True,
                resolved=True,
                asud_id="ОРТС/8/1",
            )

            def write_ahead():
                auto_create._record_document_state(
                    source, state, doc, "SUBMISSION_PENDING")

            def assert_barrier_then_run(*_args, **_kwargs):
                disk_state = auto_create.load_registration_state(source)
                self.assertEqual(
                    auto_create._document_state(disk_state, doc)["status"],
                    "SUBMISSION_PENDING",
                )
                return expected

            with patch.object(
                    auto_create, "run_registration",
                    side_effect=assert_barrier_then_run) as run_mock:
                actual = auto_create.register_and_resolve(
                    object(), 1, 1, before_register=write_ahead)

            self.assertIs(actual, expected)
            run_mock.assert_called_once()

    def test_failed_write_ahead_prevents_registration_call(self):
        with patch.object(auto_create, "run_registration") as run_mock:
            with self.assertRaisesRegex(RuntimeError, "disk full"):
                auto_create.register_and_resolve(
                    object(),
                    1,
                    1,
                    before_register=lambda: (_ for _ in ()).throw(
                        RuntimeError("disk full")
                    ),
                )

        run_mock.assert_not_called()

    def test_outcome_mapping_preserves_uncertain_and_partial_states(self):
        cases = (
            (_outcome(registered=True, resolved=True), "OK"),
            (_outcome(registered=True, resolved=False), "REGISTERED_ONLY"),
            (_outcome(uncertain=True), "SUBMISSION_UNKNOWN"),
            (_outcome(), "RETRYABLE"),
        )
        for outcome, expected in cases:
            with self.subTest(expected=expected):
                self.assertEqual(
                    auto_create._registration_status(outcome), expected)

    def test_definitely_not_attempted_atomically_clears_pending_for_retry(self):
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(tmp)
            doc = auto_create.load_excel(source)[0]
            state = auto_create.load_registration_state(source)
            auto_create._record_document_state(
                source, state, doc, "SUBMISSION_PENDING")

            retry_status = auto_create._registration_status(_outcome())
            auto_create._record_document_state(
                source,
                state,
                doc,
                retry_status,
                error="клик точно не отправлен",
            )

            restarted = auto_create.load_registration_state(source)
            entry = auto_create._document_state(restarted, doc)
            state_path = auto_create._registration_state_path(source)
            tmp_path_exists = state_path.with_name(
                state_path.name + ".tmp"
            ).exists()

        self.assertEqual(retry_status, "RETRYABLE")
        self.assertEqual(entry["status"], "RETRYABLE")
        self.assertFalse(auto_create._document_is_terminal(restarted, doc))
        self.assertFalse(tmp_path_exists)


if __name__ == "__main__":
    unittest.main()
