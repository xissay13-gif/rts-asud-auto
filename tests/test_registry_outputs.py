import tempfile
import unittest
from pathlib import Path

import openpyxl

from flows import auto_create, email, mix
from shared.xlsx_format import MAX_COLUMN_WIDTH
from shared.xlsx_status import mark_status


class RegistryOutputFormattingTests(unittest.TestCase):
    def _assert_formatted_headers(self, path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        try:
            for cell in ws[1]:
                self.assertTrue(cell.alignment.wrap_text, cell.coordinate)
                self.assertTrue(cell.font.bold, cell.coordinate)
            for col_idx in range(1, ws.max_column + 1):
                letter = openpyxl.utils.get_column_letter(col_idx)
                self.assertLessEqual(
                    ws.column_dimensions[letter].width,
                    MAX_COLUMN_WIDTH,
                    letter,
                )
        finally:
            wb.close()

    def test_mix_and_sbis_registry_headers_are_formatted(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "mix.xlsx"
            mix._ensure_output_xlsx(path)
            self._assert_formatted_headers(path)

    def test_email_and_smart_registry_headers_are_formatted(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "email.xlsx"
            email._ensure_dated_xlsx(path)
            self._assert_formatted_headers(path)

    def test_auto_create_registry_headers_are_formatted(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "auto-create.xlsx"
            auto_create._ensure_output_xlsx(path)
            self._assert_formatted_headers(path)

    def test_added_status_column_is_wrapped_and_bounded(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "status.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Номер"])
            ws.append(["ОРТС/8/2026/1"])
            wb.save(path)
            wb.close()

            self.assertTrue(mark_status(
                str(path),
                "ОРТС/8/2026/1",
                "Очень длинное название новой статусной колонки",
                "13.08.2026",
            ))

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            try:
                self.assertTrue(ws["B1"].alignment.wrap_text)
                self.assertTrue(ws["B2"].alignment.wrap_text)
                self.assertLessEqual(
                    ws.column_dimensions["B"].width,
                    MAX_COLUMN_WIDTH,
                )
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()
