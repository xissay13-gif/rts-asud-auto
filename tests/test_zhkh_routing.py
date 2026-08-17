import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from flows import email as email_flow
from flows import mix, zhkh_daemon
from shared import config as cfg
from shared.zhkh_routing import (
    normalize_topic_title,
    resolve_addressee_override,
)


BASMANOV = "Басманов Александр Владимирович"
ZHUKOV = "Жуков Иван Сергеевич"

ROUTED_TOPICS = {
    "2.1": "Отсутствие отопления",
    "2.2": "Отсутствие водоснабжения",
    "2.3": "Нарушение температурного режима подачи воды",
    "2.16": "Порыв труб",
    "2.25": "Затопление подвала",
    "5.1": "Качество оказания коммунальных услуг",
    "5.2": "Сроки оказания коммунальных услуг",
    "23": "Некачественная поставка ресурса",
}

ROUTES = [
    {
        "topics": ROUTED_TOPICS,
        "addressees": [ZHUKOV],
    }
]


class _DummyMessage:
    def close(self):
        pass


class _Element:
    def is_displayed(self):
        return True


class _ImmediateWait:
    """Browserless wait: every ASUD control requested by the flow exists."""

    def __init__(self, _driver, _timeout):
        pass

    def until(self, _predicate):
        return _Element()


def _zhkh_payload(topic):
    return {
        "фамилия": "Иванов",
        "имя": "Иван",
        "отчество": "Иванович",
        "фио": "Иванов Иван Иванович",
        "номер_обращения": "55-2026-1",
        "дата_обращения": "17.08.2026",
        "планируемая_дата": None,
        "тема_обращения": topic,
        "адрес": "г. Омск, ул. Тестовая, д. 1",
        "телефон": None,
        "email": None,
    }


def _parse_zhkh_email(topic, routes=ROUTES):
    values = {
        "subject": f"55-2026-1 — {topic or 'Без тематики'}",
        "body": "Структурированное обращение ГИС ЖКХ",
    }
    old_settings = email_flow.settings
    email_flow.settings = {
        "unknown_correspondent": "Неизвестный Неизвестный Неизвестный",
        "default_type_idx": 8,
        "zhkh_addressee_routes": routes,
    }
    try:
        with (
            patch.object(
                email_flow.extract_msg,
                "openMsg",
                return_value=_DummyMessage(),
            ),
            patch.object(
                email_flow,
                "_safe_field",
                side_effect=lambda _msg, attr, *_args: values[attr],
            ),
            patch.object(email_flow, "_msg_link", return_value="zhkh-1"),
            patch.object(
                email_flow, "_msg_date_prefix", return_value="2026-08-17"
            ),
            patch.object(
                email_flow,
                "parse_zhkh_body",
                return_value=_zhkh_payload(topic),
            ),
            patch.object(email_flow, "okrug_from_textbody", return_value=None),
            patch.object(email_flow, "_compute_zhkh_deadline", return_value=None),
        ):
            return email_flow._parse_one_msg(
                "zhkh-routing.msg", process_mode="mix"
            )
    finally:
        email_flow.settings = old_settings


class ZhkhRoutingPolicyTests(unittest.TestCase):
    def test_normalization_is_unicode_case_and_whitespace_stable(self):
        self.assertEqual(
            normalize_topic_title(
                "  НЕКАЧЕСТВЕННАЯ\u00a0\u00a0ПОСТАВКА   РЕСУРСА  "
            ),
            "некачественная поставка ресурса",
        )
        self.assertEqual(normalize_topic_title("  Ёлка  "), "елка")
        self.assertEqual(normalize_topic_title(None), "")

    def test_all_eight_exact_topics_route_only_to_zhukov(self):
        for code, title in ROUTED_TOPICS.items():
            with self.subTest(code=code, title=title):
                addressees, matched_code = resolve_addressee_override(
                    title, ROUTES
                )
                self.assertEqual(addressees, [ZHUKOV])
                self.assertEqual(matched_code, code)
                self.assertNotIn(BASMANOV, addressees)

    def test_normalized_exact_title_still_routes(self):
        addressees, code = resolve_addressee_override(
            "  СРОКИ\u00a0  ОКАЗАНИЯ  КОММУНАЛЬНЫХ УСЛУГ ", ROUTES
        )

        self.assertEqual(addressees, [ZHUKOV])
        self.assertEqual(code, "5.2")

    def test_similar_missing_or_code_prefixed_titles_do_not_match(self):
        nonmatches = (
            "Порыв трубы",
            "Качество оказания коммунальных услуг населению",
            "Срок оказания коммунальных услуг",
            "Отсутствие отопления и водоснабжения",
            "2.1 Отсутствие отопления",
            "2.10 Отсутствие отопления",
            "23.1 Некачественная поставка ресурса",
            "Некачественная поставка ресурсов",
            "Другая тема",
            "",
            None,
        )
        for title in nonmatches:
            with self.subTest(title=title):
                self.assertEqual(
                    resolve_addressee_override(title, ROUTES),
                    (None, None),
                )

    def test_duplicate_normalized_title_is_a_configuration_error(self):
        conflicting = [
            {
                "topics": {"5.2": "Сроки оказания коммунальных услуг"},
                "addressees": [ZHUKOV],
            },
            {
                "topics": {"other": "  СРОКИ  ОКАЗАНИЯ КОММУНАЛЬНЫХ УСЛУГ "},
                "addressees": [BASMANOV],
            },
        ]

        with self.assertRaises(ValueError):
            resolve_addressee_override(
                "Сроки оказания коммунальных услуг", conflicting
            )

    def test_matched_rule_with_malformed_addressees_fails_closed(self):
        malformed_values = ([], [""], "Жуков Иван Сергеевич", [42])
        for malformed in malformed_values:
            with self.subTest(addressees=malformed):
                routes = [
                    {
                        "topics": {"2.1": "Отсутствие отопления"},
                        "addressees": malformed,
                    }
                ]
                with self.assertRaises(ValueError):
                    resolve_addressee_override(
                        "Отсутствие отопления", routes
                    )

    def test_shipped_default_and_example_contain_the_eight_routes(self):
        sources = [cfg.DEFAULTS["zhkh_addressee_routes"]]
        example_path = Path(__file__).resolve().parents[1] / "settings.json.example"
        with example_path.open(encoding="utf-8") as stream:
            sources.append(json.load(stream)["zhkh_addressee_routes"])

        for source_name, routes in zip(("DEFAULTS", "settings example"), sources):
            for code, title in ROUTED_TOPICS.items():
                with self.subTest(source=source_name, code=code):
                    self.assertEqual(
                        resolve_addressee_override(title, routes),
                        ([ZHUKOV], code),
                    )


class ZhkhEmailRoutingTests(unittest.TestCase):
    def test_email_propagates_topic_override_code_and_chain_flag(self):
        doc = _parse_zhkh_email(ROUTED_TOPICS["5.2"])

        self.assertEqual(doc["тема_обращения"], ROUTED_TOPICS["5.2"])
        self.assertEqual(doc["addressees_override"], [ZHUKOV])
        self.assertEqual(doc["zhkh_topic_code"], "5.2")
        self.assertTrue(doc["skip_legacy_resolution_chain"])

    def test_unmatched_topic_keeps_global_default_route(self):
        doc = _parse_zhkh_email("Цены (тарифы) на коммунальные услуги")

        self.assertEqual(
            doc["тема_обращения"],
            "Цены (тарифы) на коммунальные услуги",
        )
        self.assertIsNone(doc["addressees_override"])
        self.assertIsNone(doc["zhkh_topic_code"])
        self.assertFalse(doc["skip_legacy_resolution_chain"])

    def test_matched_malformed_route_does_not_fall_back_to_basmanov(self):
        malformed_routes = [
            {
                "topics": {"5.2": ROUTED_TOPICS["5.2"]},
                "addressees": [],
            }
        ]

        self.assertIsNone(
            _parse_zhkh_email(ROUTED_TOPICS["5.2"], malformed_routes)
        )


class MixAddresseeOverrideTests(unittest.TestCase):
    def _run_mix(self, override):
        driver = object()
        doc = {
            "row_idx": 1,
            "тема": "ГИС ЖКХ 55-2026-1",
            "содержание": "ГИС ЖКХ 55-2026-1, тестовый адрес",
            "корреспондент": "Иванов Иван Иванович",
            "корреспондент_тип": "person",
            "корр_источник": "zhkh",
            "корр_найден": True,
            "тип_название": "Письма, заявления и жалобы граждан, акционеров",
            "link": None,
            "файл": None,
            "addressees_override": override,
        }
        with (
            patch.object(
                mix,
                "settings",
                {"addressees": [BASMANOV], "outlook_dir": ""},
            ),
            patch.object(mix, "WebDriverWait", _ImmediateWait),
            patch.object(mix.time, "sleep", return_value=None),
            patch.object(mix, "wait_and_click", return_value=True),
            patch.object(mix, "click", return_value=True) as click_mock,
            patch.object(mix, "fill_text"),
            patch.object(mix, "fill_correspondent_field", return_value=True),
            patch.object(mix, "fill_corr_number"),
            patch.object(mix, "fill_corr_date"),
            patch.object(mix, "add_addressee", return_value=True) as add_mock,
            patch.object(mix, "fill_delivery_method"),
            patch.object(mix, "is_duplicate_warning", return_value=False),
            patch.object(mix, "find_msg_by_link", return_value=None),
            patch.object(
                mix,
                "register_and_resolve",
                return_value=(True, True, "ОРТС/8/2026/1", False),
            ),
            patch.object(mix, "close_card_and_wait_main"),
        ):
            result = mix.create_one_document(driver, doc, index=1, total=1)
        return driver, doc, result, add_mock, click_mock

    def test_override_replaces_basmanov_and_records_actual_addressee(self):
        driver, doc, result, add_mock, _click_mock = self._run_mix([ZHUKOV])

        self.assertEqual(result, "ОРТС/8/2026/1")
        add_mock.assert_called_once_with(driver, ZHUKOV)
        self.assertEqual(doc["assigned_addressees"], [ZHUKOV])
        self.assertNotIn(BASMANOV, doc["assigned_addressees"])

    def test_absent_override_preserves_global_basmanov_default(self):
        driver, doc, _result, add_mock, _click_mock = self._run_mix(None)

        add_mock.assert_called_once_with(driver, BASMANOV)
        self.assertEqual(doc["assigned_addressees"], [BASMANOV])

    def test_empty_override_aborts_before_save_instead_of_falling_back(self):
        with self.assertRaises(RuntimeError):
            self._run_mix([])


class AddresseeRegistryTests(unittest.TestCase):
    @staticmethod
    def _read_rows(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        try:
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(value or "") for value in rows[0]]
            return headers, [dict(zip(headers, row)) for row in rows[1:]]
        finally:
            wb.close()

    def test_new_email_registry_contains_addressee_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "registry.xlsx"
            email_flow._ensure_dated_xlsx(path)

            headers, _rows = self._read_rows(path)

        self.assertIn("Адресат", headers)

    def test_existing_registry_is_migrated_without_losing_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "legacy.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append([
                "Номер",
                "Link",
                "Округ",
                "Subject",
                "Body",
                "Дата получения",
                "Планируемая дата",
                "Статус",
                "Отписано Халецкой",
                "Отписано в округ",
            ])
            ws.append([
                "ОРТС/8/2026/old",
                "old-link",
                "ЦАО",
                "Старая тема",
                "Старое содержание",
                "17.08.2026",
                "20.08.2026",
                "Зарегистрирован",
                "",
                "",
            ])
            wb.save(path)
            wb.close()

            email_flow._append_dated_row(
                str(path),
                {
                    "тема": "Новая тема",
                    "содержание": "Новое содержание",
                    "assigned_addressees": [BASMANOV],
                },
                "ОРТС/8/2026/new",
                status="OK",
            )
            headers, rows = self._read_rows(path)

        self.assertIn("Адресат", headers)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["Номер"], "ОРТС/8/2026/old")
        self.assertIn(rows[0]["Адресат"], (None, ""))
        self.assertEqual(rows[1]["Адресат"], BASMANOV)

    def test_direct_zhukov_route_is_recorded_and_closes_legacy_chain(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "registry.xlsx"
            email_flow._ensure_dated_xlsx(path)
            doc = {
                "link": "zhkh-1",
                "тема": "ГИС ЖКХ 55-2026-1",
                "содержание": "ГИС ЖКХ 55-2026-1",
                "дата_обращения": "17.08.2026",
                "планируемая_дата": "20.08.2026",
                "assigned_addressees": [ZHUKOV],
                "skip_legacy_resolution_chain": True,
            }

            email_flow._append_dated_row(
                str(path), doc, "ОРТС/8/2026/1", status="OK"
            )
            _headers, rows = self._read_rows(path)

        marker = f"Не требуется — прямой адресат: {ZHUKOV}"
        self.assertEqual(rows[-1]["Адресат"], ZHUKOV)
        self.assertEqual(rows[-1]["Отписано Халецкой"], marker)
        self.assertEqual(rows[-1]["Отписано в округ"], marker)


class ZhkhDaemonAddresseeFilterTests(unittest.TestCase):
    @staticmethod
    def _write_registry(path, headers, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        wb.save(path)
        wb.close()

    def test_explicit_non_basmanov_rows_are_skipped_but_blank_is_legacy(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "routing.xlsx"
            self._write_registry(
                path,
                ["Номер", "Планируемая дата", "Адресат", "Отписано Халецкой"],
                [
                    ["B-explicit", "20.08.2026", BASMANOV, ""],
                    ["B-legacy-blank", "20.08.2026", "", ""],
                    ["Z-explicit", "20.08.2026", ZHUKOV, ""],
                    ["OTHER-explicit", "20.08.2026", "Иванов Иван Иванович", ""],
                ],
            )

            todo = zhkh_daemon._read_todo(
                path, expected_addressee=BASMANOV
            )

        self.assertEqual(
            {item["asud_id"] for item in todo},
            {"B-explicit", "B-legacy-blank"},
        )

    def test_registry_without_addressee_column_remains_backward_compatible(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "legacy.xlsx"
            self._write_registry(
                path,
                ["Номер", "Планируемая дата", "Отписано Халецкой"],
                [
                    ["LEGACY-1", "20.08.2026", ""],
                    ["LEGACY-2", "21.08.2026", ""],
                ],
            )

            todo = zhkh_daemon._read_todo(path)

        self.assertEqual(
            {item["asud_id"] for item in todo},
            {"LEGACY-1", "LEGACY-2"},
        )


if __name__ == "__main__":
    unittest.main()
