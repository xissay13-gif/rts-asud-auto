import json
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl

from flows import email as email_flow
from flows.gis_api import GisApiOutcome, gis_external_guid


def _doc(msg_path):
    return {
        "файл": str(msg_path),
        "корр_источник": "zhkh",
        "корр_найден": True,
        "корреспондент": "Иванов Иван Иванович",
        "корреспондент_тип": "person",
        "номер_обращения": "55-2026-12345",
        "дата_обращения": "21.08.2026",
        "планируемая_дата": "28.08.2026",
        "тема": "ГИС ЖКХ 55-2026-12345",
        "тема_обращения": "Другая тема",
        "содержание": "Проверочное обращение",
        "округ_прогноз": "ЦАО",
        "link": "55-2026-12345.msg",
        "тип_индекс": 8,
        "skip_asud_registration": False,
    }


def _recovery_settings(branch_id="branch-recovery"):
    return {
        "asud_api": {
            "enabled": True,
            "mode": "live-one",
            "branch_id": branch_id,
        }
    }


def _write_recovery_sidecars(
    msg_path,
    *,
    external_guid,
    status,
    object_id="object-recovery",
    registration_number="АСУД/2026/recovery",
):
    claim = Path(str(msg_path) + ".asud_api.claim")
    state = Path(str(msg_path) + ".asud_api_state.json")
    claim.write_text(
        json.dumps({"version": 1, "external_guid": external_guid}),
        encoding="utf-8",
    )
    state.write_text(
        json.dumps(
            {
                "version": 1,
                "status": status,
                "phase": "TERMINAL",
                "external_guid": external_guid,
                "object_id": object_id,
                "registration_number": registration_number,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return claim, state


class GisApiEmailFlowTests(unittest.TestCase):
    def setUp(self):
        self.old_settings = email_flow.settings
        self.addCleanup(setattr, email_flow, "settings", self.old_settings)

    def test_explicit_api_backend_never_falls_back_to_selenium(self):
        with tempfile.TemporaryDirectory() as tmp:
            msg = Path(tmp) / "api.msg"
            msg.write_bytes(b"msg")
            doc = _doc(msg)
            email_flow.settings = {"asud_api": {"enabled": True}}

            with (
                patch.dict(
                    os.environ,
                    {"ASUD_EMAIL_REGISTRATION_BACKEND": "asud_api"},
                    clear=False,
                ),
                patch(
                    "flows.gis_api.should_use_gis_api", return_value=True
                ),
                patch(
                    "flows.gis_api.process_gis_document",
                    return_value=GisApiOutcome(
                        "DRY_RUN", external_guid="guid-1"
                    ),
                ) as api_mock,
                patch.object(
                    email_flow.mix_flow, "create_one_document"
                ) as selenium_mock,
                patch.object(email_flow, "move_to_done") as move_mock,
            ):
                result = email_flow._process_doc(
                    None,
                    doc,
                    tmp,
                    tmp,
                    1,
                    1,
                    in_daemon=False,
                    output_suffix="ГИСЖКХ_API_TEST",
                )

            self.assertEqual(result, ("DRY_RUN", None, None))
            api_mock.assert_called_once()
            selenium_mock.assert_not_called()
            move_mock.assert_not_called()
            self.assertFalse(
                Path(str(msg) + ".asud_terminal.json").exists()
            )

    def test_api_success_upserts_registry_and_keeps_msg_with_marker(self):
        with tempfile.TemporaryDirectory() as tmp:
            msg = Path(tmp) / "success.msg"
            msg.write_bytes(b"msg")
            doc = _doc(msg)
            email_flow.settings = {"asud_api": {"enabled": True}}
            outcome = GisApiOutcome(
                "OK",
                object_id="object-1",
                registration_number="АСУД/2026/1",
                external_guid="guid-1",
                state_path=str(msg) + ".asud_api_state.json",
            )

            with (
                patch.dict(
                    os.environ,
                    {"ASUD_EMAIL_REGISTRATION_BACKEND": "asud_api"},
                    clear=False,
                ),
                patch(
                    "flows.gis_api.should_use_gis_api", return_value=True
                ),
                patch(
                    "flows.gis_api.process_gis_document",
                    return_value=outcome,
                ),
                patch.object(
                    email_flow.mix_flow, "create_one_document"
                ) as selenium_mock,
                patch.object(email_flow, "move_to_done") as move_mock,
            ):
                status, asud_id, registry = email_flow._process_doc(
                    None,
                    doc,
                    tmp,
                    tmp,
                    1,
                    1,
                    in_daemon=False,
                    output_suffix="ГИСЖКХ_API_TEST",
                )

            self.assertEqual(status, "OK")
            self.assertEqual(asud_id, "АСУД/2026/1")
            self.assertTrue(Path(registry).exists())
            self.assertTrue(msg.exists())
            marker = Path(str(msg) + ".asud_terminal.json")
            self.assertTrue(marker.exists())
            with marker.open(encoding="utf-8") as stream:
                marker_data = json.load(stream)
            self.assertEqual(marker_data["status"], "OK")
            selenium_mock.assert_not_called()
            move_mock.assert_not_called()

            workbook = openpyxl.load_workbook(registry)
            worksheet = workbook.active
            headers = [cell.value for cell in worksheet[1]]
            row = {
                header: worksheet.cell(2, index + 1).value
                for index, header in enumerate(headers)
            }
            workbook.close()
            self.assertEqual(row["Номер"], "АСУД/2026/1")
            self.assertEqual(row["GIS Номер обращения"], "55-2026-12345")
            self.assertEqual(row["External GUID"], "guid-1")
            self.assertEqual(email_flow._list_root_msgs(tmp), [])

    def test_api_registry_upsert_does_not_duplicate_same_guid(self):
        with tempfile.TemporaryDirectory() as tmp:
            registry = str(Path(tmp) / "registry.xlsx")
            msg = Path(tmp) / "row.msg"
            msg.write_bytes(b"msg")
            doc = _doc(msg)

            email_flow._upsert_api_dated_row(
                registry, doc, "АСУД/2026/1", "guid-1"
            )
            email_flow._upsert_api_dated_row(
                registry, doc, "АСУД/2026/1", "guid-1"
            )

            workbook = openpyxl.load_workbook(registry)
            self.assertEqual(workbook.active.max_row, 2)
            workbook.close()

    def test_terminal_api_outcome_never_enters_resolution_registry(self):
        with tempfile.TemporaryDirectory() as tmp:
            msg = Path(tmp) / "uncertain.msg"
            msg.write_bytes(b"msg")
            doc = _doc(msg)
            email_flow.settings = {"asud_api": {"enabled": True}}
            outcome = GisApiOutcome(
                "SUBMISSION_UNKNOWN",
                external_guid="guid-1",
                state_path=str(msg) + ".asud_api_state.json",
                message="timeout_after_submit",
            )

            with (
                patch.dict(
                    os.environ,
                    {"ASUD_EMAIL_REGISTRATION_BACKEND": "asud_api"},
                    clear=False,
                ),
                patch(
                    "flows.gis_api.should_use_gis_api", return_value=True
                ),
                patch(
                    "flows.gis_api.process_gis_document",
                    return_value=outcome,
                ),
                patch.object(email_flow, "_xlsx_path") as xlsx_mock,
                patch.object(
                    email_flow.mix_flow, "create_one_document"
                ) as selenium_mock,
            ):
                result = email_flow._process_doc(
                    None,
                    doc,
                    tmp,
                    tmp,
                    1,
                    1,
                    in_daemon=False,
                    output_suffix="ГИСЖКХ_API_TEST",
                )

            self.assertEqual(result[0], "SUBMISSION_UNKNOWN")
            self.assertTrue(msg.exists())
            self.assertTrue(
                Path(str(msg) + ".asud_terminal.json").exists()
            )
            xlsx_mock.assert_not_called()
            selenium_mock.assert_not_called()

    def test_scanner_recovers_claim_plus_state_only_in_api_backend(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            recovery_msg = root / "recovery.msg"
            claim_only_msg = root / "claim-only.msg"
            state_only_msg = root / "state-only.msg"
            clean_msg = root / "clean.msg"
            for item in (
                recovery_msg,
                claim_only_msg,
                state_only_msg,
                clean_msg,
            ):
                item.write_bytes(b"msg")

            _write_recovery_sidecars(
                recovery_msg,
                external_guid="guid-recovery",
                status="OK",
            )
            Path(str(claim_only_msg) + ".asud_api.claim").write_text(
                "{}", encoding="utf-8"
            )
            Path(str(state_only_msg) + ".asud_api_state.json").write_text(
                "{}", encoding="utf-8"
            )

            with patch.dict(
                os.environ,
                {"ASUD_EMAIL_REGISTRATION_BACKEND": "selenium"},
                clear=False,
            ):
                self.assertEqual(
                    email_flow._list_root_msgs(tmp),
                    [str(clean_msg)],
                )

            with patch.dict(
                os.environ,
                {"ASUD_EMAIL_REGISTRATION_BACKEND": "asud_api"},
                clear=False,
            ):
                self.assertEqual(
                    email_flow._list_root_msgs(tmp),
                    [str(clean_msg), str(recovery_msg)],
                )

    def test_recovery_ok_finishes_xlsx_and_marker_without_network(self):
        with tempfile.TemporaryDirectory() as tmp:
            msg = Path(tmp) / "recovery-ok.msg"
            msg.write_bytes(b"msg")
            doc = _doc(msg)
            branch_id = "branch-ok"
            external_guid = gis_external_guid(
                doc["номер_обращения"], branch_id
            )
            _write_recovery_sidecars(
                msg,
                external_guid=external_guid,
                status="OK",
                registration_number="АСУД/2026/OK",
            )
            email_flow.settings = _recovery_settings(branch_id)

            with (
                patch.dict(
                    os.environ,
                    {"ASUD_EMAIL_REGISTRATION_BACKEND": "asud_api"},
                    clear=False,
                ),
                patch("flows.gis_api.AsudApiClient") as client_mock,
                patch.object(
                    email_flow.mix_flow, "create_one_document"
                ) as selenium_mock,
            ):
                result = email_flow._process_doc(
                    None,
                    doc,
                    tmp,
                    tmp,
                    1,
                    1,
                    in_daemon=False,
                    output_suffix="ГИСЖКХ_API_TEST",
                )
                # Simulate a crash after the XLSX replace but before the
                # terminal marker survived.  Replaying the durable API state
                # must update the same row, never append a duplicate.
                Path(str(msg) + ".asud_terminal.json").unlink()
                replayed = email_flow._process_doc(
                    None,
                    doc,
                    tmp,
                    tmp,
                    1,
                    1,
                    in_daemon=False,
                    output_suffix="ГИСЖКХ_API_TEST",
                )

            self.assertEqual(result[0:2], ("OK", "АСУД/2026/OK"))
            self.assertEqual(replayed, result)
            client_mock.assert_not_called()
            selenium_mock.assert_not_called()
            self.assertTrue(Path(result[2]).is_file())
            marker = Path(str(msg) + ".asud_terminal.json")
            self.assertTrue(marker.is_file())
            self.assertEqual(
                json.loads(marker.read_text(encoding="utf-8"))["status"],
                "OK",
            )

            workbook = openpyxl.load_workbook(result[2])
            worksheet = workbook.active
            headers = [cell.value for cell in worksheet[1]]
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            workbook.close()
            self.assertEqual(len(rows), 1)
            recovered = dict(zip(headers, rows[0]))
            self.assertEqual(recovered["External GUID"], external_guid)
            self.assertEqual(recovered["Номер"], "АСУД/2026/OK")

    def test_recovery_registered_only_marks_terminal_without_xlsx_or_network(self):
        with tempfile.TemporaryDirectory() as tmp:
            msg = Path(tmp) / "recovery-registered-only.msg"
            msg.write_bytes(b"msg")
            doc = _doc(msg)
            branch_id = "branch-registered-only"
            external_guid = gis_external_guid(
                doc["номер_обращения"], branch_id
            )
            _write_recovery_sidecars(
                msg,
                external_guid=external_guid,
                status="REGISTERED_ONLY",
            )
            email_flow.settings = _recovery_settings(branch_id)

            with (
                patch.dict(
                    os.environ,
                    {"ASUD_EMAIL_REGISTRATION_BACKEND": "asud_api"},
                    clear=False,
                ),
                patch("flows.gis_api.AsudApiClient") as client_mock,
                patch.object(email_flow, "_xlsx_path") as xlsx_mock,
            ):
                result = email_flow._process_doc(
                    None,
                    doc,
                    tmp,
                    tmp,
                    1,
                    1,
                    in_daemon=False,
                    output_suffix="ГИСЖКХ_API_TEST",
                )

            self.assertEqual(result[0], "REGISTERED_ONLY")
            client_mock.assert_not_called()
            xlsx_mock.assert_not_called()
            marker = Path(str(msg) + ".asud_terminal.json")
            self.assertTrue(marker.is_file())
            self.assertEqual(
                json.loads(marker.read_text(encoding="utf-8"))["status"],
                "REGISTERED_ONLY",
            )

    def test_recovery_guid_mismatch_is_manual_without_xlsx_or_network(self):
        with tempfile.TemporaryDirectory() as tmp:
            msg = Path(tmp) / "recovery-guid-mismatch.msg"
            msg.write_bytes(b"msg")
            doc = _doc(msg)
            branch_id = "branch-guid-mismatch"
            expected_guid = gis_external_guid(
                doc["номер_обращения"], branch_id
            )
            self.assertNotEqual(expected_guid, "wrong-guid")
            _write_recovery_sidecars(
                msg,
                external_guid="wrong-guid",
                status="OK",
            )
            email_flow.settings = _recovery_settings(branch_id)

            with (
                patch.dict(
                    os.environ,
                    {"ASUD_EMAIL_REGISTRATION_BACKEND": "asud_api"},
                    clear=False,
                ),
                patch("flows.gis_api.AsudApiClient") as client_mock,
                patch.object(email_flow, "_xlsx_path") as xlsx_mock,
            ):
                result = email_flow._process_doc(
                    None,
                    doc,
                    tmp,
                    tmp,
                    1,
                    1,
                    in_daemon=False,
                    output_suffix="ГИСЖКХ_API_TEST",
                )

            self.assertEqual(result[0], "MANUAL_REVIEW")
            client_mock.assert_not_called()
            xlsx_mock.assert_not_called()
            marker = Path(str(msg) + ".asud_terminal.json")
            self.assertTrue(marker.is_file())
            self.assertEqual(
                json.loads(marker.read_text(encoding="utf-8"))["status"],
                "MANUAL_REVIEW",
            )

    def test_recovery_corrupt_state_is_terminal_without_xlsx_or_network(self):
        with tempfile.TemporaryDirectory() as tmp:
            msg = Path(tmp) / "recovery-corrupt.msg"
            msg.write_bytes(b"msg")
            doc = _doc(msg)
            branch_id = "branch-corrupt"
            external_guid = gis_external_guid(
                doc["номер_обращения"], branch_id
            )
            claim = Path(str(msg) + ".asud_api.claim")
            state = Path(str(msg) + ".asud_api_state.json")
            claim.write_text(
                json.dumps({"external_guid": external_guid}),
                encoding="utf-8",
            )
            state.write_text("{not-json", encoding="utf-8")
            email_flow.settings = _recovery_settings(branch_id)

            with (
                patch.dict(
                    os.environ,
                    {"ASUD_EMAIL_REGISTRATION_BACKEND": "asud_api"},
                    clear=False,
                ),
                patch("flows.gis_api.AsudApiClient") as client_mock,
                patch.object(email_flow, "_xlsx_path") as xlsx_mock,
            ):
                result = email_flow._process_doc(
                    None,
                    doc,
                    tmp,
                    tmp,
                    1,
                    1,
                    in_daemon=False,
                    output_suffix="ГИСЖКХ_API_TEST",
                )

            self.assertEqual(result[0], "SUBMISSION_UNKNOWN")
            client_mock.assert_not_called()
            xlsx_mock.assert_not_called()
            marker = Path(str(msg) + ".asud_terminal.json")
            self.assertTrue(marker.is_file())
            self.assertEqual(
                json.loads(marker.read_text(encoding="utf-8"))["status"],
                "SUBMISSION_UNKNOWN",
            )

    def test_normal_backend_is_unchanged(self):
        with tempfile.TemporaryDirectory() as tmp:
            msg = Path(tmp) / "selenium.msg"
            msg.write_bytes(b"msg")
            doc = _doc(msg)
            email_flow.settings = {}
            email_flow.mix_flow._last_result = {"status": "DUPLICATE"}

            with (
                patch.dict(
                    os.environ,
                    {"ASUD_EMAIL_REGISTRATION_BACKEND": "selenium"},
                    clear=False,
                ),
                patch.object(
                    email_flow.mix_flow,
                    "create_one_document",
                    return_value="АСУД/old",
                ) as selenium_mock,
                patch.object(email_flow, "move_to_done") as move_mock,
                patch(
                    "flows.gis_api.process_gis_document"
                ) as api_mock,
            ):
                result = email_flow._process_doc(
                    object(), doc, tmp, tmp, 1, 1, in_daemon=False
                )

            self.assertEqual(result[0], "DUPLICATE")
            selenium_mock.assert_called_once()
            move_mock.assert_called_once()
            api_mock.assert_not_called()

    def test_unknown_backend_aborts_instead_of_falling_into_selenium(self):
        with tempfile.TemporaryDirectory() as tmp:
            msg = Path(tmp) / "unknown-backend.msg"
            msg.write_bytes(b"msg")
            doc = _doc(msg)

            with (
                patch.dict(
                    os.environ,
                    {"ASUD_EMAIL_REGISTRATION_BACKEND": "asud-api"},
                    clear=False,
                ),
                patch.object(
                    email_flow.mix_flow, "create_one_document"
                ) as selenium_mock,
            ):
                with self.assertRaisesRegex(ValueError, "selenium or asud_api"):
                    email_flow._process_doc(
                        object(), doc, tmp, tmp, 1, 1, in_daemon=False
                    )

            selenium_mock.assert_not_called()

    def test_api_one_shot_does_not_require_driver_or_start_edge(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            mailbox = root / "mailbox"
            mailbox.mkdir()
            msg = mailbox / "one.msg"
            msg.write_bytes(b"msg")
            doc = _doc(msg)
            main_settings = {
                "email_folder": str(mailbox),
                "asud_api": {"enabled": True},
            }

            with (
                patch.dict(
                    os.environ,
                    {
                        "ASUD_EMAIL_REGISTRATION_BACKEND": "asud_api",
                        "ASUD_EMAIL_PROCESS_MODE": "mix",
                        "ASUD_OUTPUT_SUFFIX": "ГИСЖКХ_API_TEST",
                    },
                    clear=False,
                ),
                patch.object(email_flow.cfg, "load", return_value=main_settings),
                patch.object(email_flow.cfg, "setup_file_logger"),
                patch.object(email_flow.cfg, "keep_system_awake"),
                patch.object(email_flow.cfg, "get_base_dir", return_value=tmp),
                patch.object(
                    email_flow,
                    "_wait_for_folder",
                    return_value=(True, str(mailbox)),
                ),
                patch.object(email_flow, "load_emails", return_value=[doc]),
                patch.object(
                    email_flow,
                    "_process_doc",
                    return_value=("DRY_RUN", None, None),
                ) as process_mock,
                patch.object(email_flow.webdriver, "Edge") as edge_mock,
                patch("builtins.input", side_effect=["", "", ""]),
            ):
                email_flow.main()

            process_mock.assert_called_once()
            self.assertIsNone(process_mock.call_args.args[0])
            edge_mock.assert_not_called()

    def test_api_preflight_rejects_two_documents_before_any_api_or_edge(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            mailbox = root / "mailbox"
            mailbox.mkdir()
            first = mailbox / "one.msg"
            second = mailbox / "two.msg"
            first.write_bytes(b"one")
            second.write_bytes(b"two")
            docs = [_doc(first), _doc(second)]
            docs[1]["номер_обращения"] = "55-2026-99999"

            with (
                patch.dict(
                    os.environ,
                    {"ASUD_EMAIL_REGISTRATION_BACKEND": "asud_api"},
                    clear=False,
                ),
                patch.object(
                    email_flow.cfg,
                    "load",
                    return_value={
                        "email_folder": str(mailbox),
                        "asud_api": {"enabled": True},
                    },
                ),
                patch.object(email_flow.cfg, "setup_file_logger"),
                patch.object(email_flow.cfg, "keep_system_awake"),
                patch.object(email_flow.cfg, "get_base_dir", return_value=tmp),
                patch.object(
                    email_flow,
                    "_wait_for_folder",
                    return_value=(True, str(mailbox)),
                ),
                patch.object(email_flow, "load_emails", return_value=docs),
                patch.object(email_flow, "_process_doc") as process_mock,
                patch.object(email_flow.webdriver, "Edge") as edge_mock,
                patch("builtins.input", side_effect=["", ""]),
            ):
                email_flow.main()

            process_mock.assert_not_called()
            edge_mock.assert_not_called()

    def test_api_daemon_is_refused_before_edge_start(self):
        with (
            patch.dict(
                os.environ,
                {"ASUD_EMAIL_REGISTRATION_BACKEND": "asud_api"},
                clear=False,
            ),
            patch.object(email_flow.cfg, "load", return_value={}),
            patch.object(email_flow.cfg, "setup_file_logger"),
            patch.object(email_flow.cfg, "keep_system_awake"),
            patch.object(email_flow.webdriver, "Edge") as edge_mock,
        ):
            with self.assertRaisesRegex(SystemExit, "2"):
                email_flow.daemon_main()

        edge_mock.assert_not_called()


if __name__ == "__main__":
    unittest.main()
