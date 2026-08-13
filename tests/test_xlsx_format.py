import tempfile
import unittest
from pathlib import Path

import openpyxl

from shared.xlsx_format import (
    MAX_COLUMN_WIDTH,
    MIN_COLUMN_WIDTH,
    format_registry_before_save,
    format_registry_worksheet,
)


class XlsxFormatTests(unittest.TestCase):
    def test_autosizes_columns_and_wraps_long_text(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Коротко", "Очень длинное содержание"])
        ws.append(["1", "Текст " * 40])

        format_registry_worksheet(ws)

        self.assertTrue(ws["A1"].alignment.wrap_text)
        self.assertTrue(ws["B2"].alignment.wrap_text)
        self.assertEqual(ws["B2"].alignment.vertical, "top")
        self.assertGreaterEqual(ws.column_dimensions["A"].width,
                                MIN_COLUMN_WIDTH)
        self.assertEqual(ws.column_dimensions["B"].width,
                         MAX_COLUMN_WIDTH)
        self.assertGreaterEqual(ws.row_dimensions[1].height, 30)

    def test_new_row_can_expand_existing_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "реестр.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Номер", "Статус"])
            format_registry_before_save(ws, path, changed_row=1)
            initial_width = ws.column_dimensions["B"].width

            ws.append(["ОРТС/1", "Зарегистрирован 13.08.2026 12:00"])
            format_registry_before_save(ws, path, changed_row=2)

            self.assertTrue(ws["B2"].alignment.wrap_text)
            self.assertGreater(ws.column_dimensions["B"].width,
                               initial_width)
            self.assertLessEqual(ws.column_dimensions["B"].width,
                                 MAX_COLUMN_WIDTH)


if __name__ == "__main__":
    unittest.main()
